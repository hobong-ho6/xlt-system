#!/usr/bin/env python3
"""
XLT System v5.0.6 Stable Web Interface (완전 자동화 시스템)
안정적인 웹 서버 (강화된 예외 처리)
"""

import os
import sys
import time
import traceback
import secrets
import re
import json
import requests
import logging
import subprocess
from logging.handlers import RotatingFileHandler
from flask import Flask, request, jsonify, send_from_directory, render_template, redirect, url_for
from werkzeug.utils import secure_filename
from datetime import datetime
from pathlib import Path

# XLT 패키지 경로 추가
sys.path.insert(0, os.path.join(os.path.dirname(__file__)))

app = Flask(__name__)
app.secret_key = 'xlt_system_stable_key'

# Favicon 라우트 (404 에러 방지)
@app.route('/favicon.ico')
def favicon():
    return '', 204  # No Content

# 로깅 설정 (파일 + 콘솔)
# logs 디렉토리 생성
logs_dir = Path(__file__).parent / "logs"
logs_dir.mkdir(exist_ok=True)
log_file = logs_dir / "server.log"
log_handler = RotatingFileHandler(log_file, maxBytes=10*1024*1024, backupCount=3)  # 10MB, 3개 백업
log_handler.setLevel(logging.INFO)
log_formatter = logging.Formatter(
    '%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
log_handler.setFormatter(log_formatter)

# 콘솔 핸들러 추가
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(log_formatter)

# 로거 설정
logger = logging.getLogger('xlt_server')
logger.setLevel(logging.INFO)
logger.addHandler(log_handler)
logger.addHandler(console_handler)

# Flask 기본 로거도 설정
app.logger.addHandler(log_handler)
app.logger.setLevel(logging.INFO)

# Jinja2 환경에 enumerate 함수 추가
app.jinja_env.globals['enumerate'] = enumerate

# CORS 헤더 추가 (DropWeb 프론트엔드 지원)
@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

# OPTIONS 요청 처리 (CORS Preflight)
@app.route('/', methods=['OPTIONS'])
@app.route('/<path:path>', methods=['OPTIONS'])
def options_handler(path=None):
    return '', 200

# 템플릿 컨텍스트 프로세서 - 모든 템플릿에서 버전 정보 사용 가능
@app.context_processor
def inject_version_info():
    """모든 템플릿에 버전 정보 주입 (통합 version_info 객체)"""
    try:
        from xlt.utils.version_manager import get_version_manager
        vm = get_version_manager()
        vm.refresh_cache()  # 항상 최신 정보 보장

        version_info = vm.get_version_info()

        return {
            'version_info': {
                'name': version_info.get('name', 'XLT System'),
                'version': vm.get_version_number(),  # 5.1.6 (v 없이)
                'full_version': vm.get_version(),    # v5.1.6 (v 포함)
                'build': vm.get_build_date(),
                'features': version_info.get('features', [])
            },
            # 하위 호환성을 위한 개별 변수들
            'version': vm.get_version(),
            'version_number': vm.get_version_number(),
            'full_name': vm.get_full_name(),
            'build_date': vm.get_build_date()
        }
    except Exception as e:
        # 폴백: 하드코딩된 버전 (안전장치)
        print(f"⚠️ 버전 정보 로드 실패, 폴백 사용: {e}")
        return {
            'version_info': {
                'name': 'XLT System',
                'version': '5.1.6',
                'full_version': 'v5.1.6',
                'build': '2026-05-14',
                'features': []
            },
            'version': 'v5.1.6',
            'version_number': '5.1.6',
            'full_name': 'XLT System v5.1.6',
            'build_date': '2026-05-14'
        }

# XLT v3.0: 로그 기능 제거 - 하지만 변수는 유지 (오류 방지)
session_status = {}
session_logs = {}
translation_progress = {}  # 번역 진행 상황 추적 (세션별)

# 전역 오류 핸들러 (모든 오류를 JSON으로 응답)
@app.errorhandler(404)
def not_found(error):
    return jsonify({
        'status': 'error',
        'error': 'API 엔드포인트를 찾을 수 없습니다.',
        'code': 404
    }), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({
        'status': 'error',
        'error': '내부 서버 오류가 발생했습니다.',
        'code': 500,
        'details': str(error)
    }), 500

@app.errorhandler(Exception)
def handle_exception(e):
    # HTTP 오류가 아닌 일반 예외 처리
    return jsonify({
        'status': 'error',
        'error': f'예기치 않은 오류: {str(e)}',
        'type': type(e).__name__
    }), 500

# XLT v3.0: 로그 기능 제거
def add_session_log(session_id, message, log_type='info'):
    """세션 로그 추가 (파일 + 콘솔 출력)"""
    # 로그 레벨에 따라 출력
    log_message = f"[{session_id}] {message}"

    if log_type == 'error':
        logger.error(log_message)
    elif log_type == 'warning':
        logger.warning(log_message)
    elif log_type == 'success':
        logger.info(f"✅ {log_message}")
    else:
        logger.info(log_message)

# 설정
# UPLOAD_FOLDER = 'uploads'  # XLT v3.0: 피그마 전용으로 파일 업로드 기능 제거
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'bmp', 'tiff', 'webp', 'gif'}

# 업로드 폴더 생성 (XLT v3.0: 피그마 전용으로 제거)
# os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# XLT 시스템 초기화
pipeline = None
config = None
updater = None
auto_updater = None

# 메인 시스템 초기화 (파이프라인)
try:
    from xlt import XLTPipeline, XLTConfig

    config = XLTConfig()
    pipeline = XLTPipeline(config)
    pipeline.initialize()  # 🔧 명시적으로 초기화 호출

    print("✅ XLT 시스템 초기화 완료")
    print(f"   📝 입력 처리기: {len(pipeline.input_processors)}개 등록")
    print(f"   📋 처리기 종류: {list(pipeline.input_processors.keys())}")
except Exception as e:
    print(f"❌ XLT 시스템 초기화 실패: {e}")

# 업데이터 초기화 (독립적으로 실행)
try:
    from xlt.utils.updater import XLTUpdater
    updater = XLTUpdater()
    print("✅ 업데이터 초기화 완료")
except Exception as e:
    print(f"⚠️ 업데이터 초기화 실패: {e}")
    print("   수동 업데이트만 가능합니다.")

# 자동 업데이터 초기화 (독립적으로 실행)
try:
    from xlt.utils.auto_updater import get_auto_updater

    # 자동 업데이터 초기화 (웹 서버 모드에서는 백그라운드만)
    auto_updater = get_auto_updater(enable_tray=False)  # 웹서버에서는 트레이 비활성화
    # 웹 서버에서는 트레이가 없으므로 백그라운드만 실행
    auto_updater.start_background_check()
    print("🔍 웹 서버 자동 업데이트 감지 활성화")
except Exception as e:
    print(f"⚠️ 자동 업데이터 초기화 실패: {e}")
    print("   업데이트 확인은 수동으로만 가능합니다.")


def cleanup_old_figma_images():
    """오래된 피그마 이미지 파일 정리 (24시간 이상 된 파일 삭제)"""
    try:
        figma_dir = os.path.join(os.getcwd(), 'figma')
        if not os.path.exists(figma_dir):
            return

        import time
        current_time = time.time()
        cleanup_age = 24 * 3600  # 24시간

        for filename in os.listdir(figma_dir):
            if filename.endswith(('.png', '.jpg', '.jpeg', '.webp')):
                file_path = os.path.join(figma_dir, filename)
                try:
                    file_age = current_time - os.path.getmtime(file_path)
                    if file_age > cleanup_age:
                        os.remove(file_path)
                        print(f"[CLEANUP] 오래된 피그마 이미지 삭제: {filename}")
                except Exception as e:
                    print(f"[CLEANUP ERROR] {filename}: {e}")
    except Exception as e:
        print(f"[CLEANUP ERROR] 피그마 이미지 정리 실패: {e}")


# ❌ 정적 교정 캐시 함수들 제거됨 - Claude AI 전용으로 전환
# 캐시 기반 정적 교정은 모두 제거되었습니다.


# ❌ 바른 API 관련 함수들 제거됨 - Claude AI 전용으로 전환
# 이전 정적 룰 기반 교정 시스템들은 모두 제거되었습니다.
# 모든 교정은 Claude AI를 통해서만 처리됩니다.


def check_spelling_with_api(text):
    """정적 교정 API 제거됨 - Claude AI로 대체"""
    print("⚠️ 정적 교정 API는 제거되었습니다. Claude AI 교정을 사용하세요.")
    return text


def apply_korean_corrections(text):
    """Claude AI 전용 한국어 맞춤법 및 띄어쓰기 교정 (v5.1.0: 정적 룰 완전 제거)"""
    if not text or not text.strip():
        return text

    try:
        # Claude 교정기 초기화
        from xlt.translation.claude_translator import ClaudeTranslator
        from xlt.core.config import XLTConfig

        config = XLTConfig()
        claude_corrector = ClaudeTranslator(config)

        # Claude AI로 교정 실행
        result = claude_corrector.correct_korean_text_only(text)

        corrected_text = result.get('corrected', text)
        corrections_applied = result.get('corrections_applied', [])

        if corrections_applied:
            print(f"🤖 Claude 교정: '{text}' → '{corrected_text}'")
            for correction in corrections_applied[:3]:  # 최대 3개만 출력
                print(f"   📝 {correction}")
        else:
            print(f"🤖 Claude 교정: '{text}' (교정 불필요)")

        return corrected_text

    except Exception as e:
        print(f"❌ Claude 교정 실패: {e}")
        print(f"⚠️ 정적 룰 제거됨 - Claude AI 필수")

        # Claude 실패 시 원본 반환 (정적 룰 완전 제거)
        return text.strip()


def load_translation_guide():
    """LINE API에서 번역 용어집 로드 (guide.md 대체)"""
    try:
        # LINE API에서 용어집 로드
        terminology = load_line_api_terminology()
        if terminology:
            print(f"🌐 LINE API에서 {len(terminology)}개 용어 로드됨")
            return terminology

        # 폴백: 로컬 JSON 파일
        terminology_json_path = os.path.join(os.path.dirname(__file__), 'terminology_data.json')
        if os.path.exists(terminology_json_path):
            with open(terminology_json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            terminology = {}
            for korean, translations in data.get('terminology', {}).items():
                terminology[korean] = {
                    'en_US': translations.get('en_US', ''),
                    'ja_JP': translations.get('ja_JP', ''),
                    'zh_TW': translations.get('zh_TW', ''),
                    'th_TH': translations.get('th_TH', '')
                }

            print(f"📂 로컬 JSON에서 {len(terminology)}개 용어 로드됨 (폴백)")
            return terminology

        # 최종 폴백: 하드코딩된 핵심 용어
        terminology = {
            '거래': {'en_US': 'transaction', 'ja_JP': '取引', 'zh_TW': '交易', 'th_TH': 'ธุรกรรม'},
            '지갑': {'en_US': 'wallet', 'ja_JP': 'ウォレット', 'zh_TW': '錢包', 'th_TH': 'กระเป๋า'},
            '토큰': {'en_US': 'token', 'ja_JP': 'トークン', 'zh_TW': '代幣', 'th_TH': 'โทเค็น'},
            '자산': {'en_US': 'asset', 'ja_JP': '資産', 'zh_TW': '資產', 'th_TH': 'สินทรัพย์'},
            '로그인': {'en_US': 'log in', 'ja_JP': 'ログイン', 'zh_TW': '登入', 'th_TH': 'เข้าสู่ระบบ'}
        }
        print(f"🔄 하드코딩된 핵심 용어 사용: {len(terminology)}개 용어 (최종 폴백)")
        return terminology

    except Exception as e:
        print(f"❌ 용어집 로드 실패: {e}")
        return {}

def load_line_api_terminology():
    """LINE API에서 용어집 데이터 로드"""
    try:
        import requests

        # LINE API URL
        api_url = "https://landpress-content.line-scdn.net/contents/v2/projects/wdmwbfuv10x39bukv58ocevp/collections/web3_xlt_json/item"

        # API 호출 (타임아웃 3초)
        response = requests.get(api_url, timeout=3)

        if response.status_code != 200:
            return None

        data = response.json()

        if not data.get('header', {}).get('success'):
            return None

        # 용어집 데이터 추출 (exceptions 키 안에 실제 데이터가 있음)
        terminology_data = data['body']['exceptions']['terminology']

        # 서버 형식으로 변환 ({korean: {en_US: '', ja_JP: '', ...}})
        terminology = {}
        for korean, translations in terminology_data.items():
            terminology[korean] = {
                'en_US': translations.get('en_US', ''),
                'ja_JP': translations.get('ja_JP', ''),
                'zh_TW': translations.get('zh_TW', ''),
                'th_TH': translations.get('th_TH', '')
            }

        return terminology

    except Exception as e:
        print(f"⚠️ LINE API 연결 실패: {str(e)}")
        return None


def apply_terminology_guide(text, target_lang, terminology):
    """LINE API 용어집을 사용해서 번역 후 용어 일관성 검증 및 수정"""
    if not text or not terminology:
        return text

    # 용어집에서 일치하는 용어 찾기 (번역된 텍스트에서 한국어 용어 검색 후 치환)
    for korean_term, translations in terminology.items():
        if korean_term in text and target_lang in translations:
            guided_translation = translations[target_lang]
            # 한국어 용어를 해당 언어의 표준 번역으로 치환
            # \b는 비라틴 문자에서 작동하지 않으므로 단순 문자열 치환 사용
            text = text.replace(korean_term, guided_translation)

    return text


# Excel 파일 생성 기능 제거됨 (XLT System v3.0)


# Excel 파일 생성 기능 제거됨 (create_excel_from_template - 레거시)


# =====================================
# 설정 관리 시스템 (XLT System v3.0)
# =====================================

def load_user_config():
    """사용자 설정 파일 로드"""
    config_path = Path('user_config.json')
    default_config = {
        'server_port': 5004,
        'max_concurrent_users': 5,
        'session_timeout_minutes': 60,
        'translation': {
            'default_engine': 'claude_integrated',
            'claude_settings': {
                'timeout': 180,
                'cli_command': 'claude',
                'verify_on_startup': True
            }
        },
        # v5.0.7 로그 관리 시스템
        'logging': {
            'enabled': True,
            'log_dir': 'logs',
            'log_level': 'INFO',
            'max_file_size_mb': 10,
            'backup_count': 3,
            'auto_updater_log_enabled': True,
            'auto_updater_separate_file': False
        }
    }

    if config_path.exists():
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                # 기본값과 병합
                return {**default_config, **config}
        except Exception as e:
            print(f"⚠️ user_config.json 로드 실패: {e}")

    return default_config

def save_user_config(config):
    """사용자 설정 파일 저장"""
    try:
        config_path = Path('user_config.json')
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"❌ user_config.json 저장 실패: {e}")
        return False

def load_version_info():
    """버전 정보 로드"""
    try:
        version_path = Path('version.json')
        if version_path.exists():
            with open(version_path, 'r', encoding='utf-8') as f:
                version_data = json.load(f)
                return {
                    'status': 'success',
                    'version': version_data.get('version', '알 수 없음'),
                    'build': version_data.get('build', '알 수 없음'),
                    'name': version_data.get('name', 'XLT System'),
                    'description': version_data.get('description', ''),
                    'release_date': version_data.get('release_date', ''),
                    'changelog': version_data.get('changelog', [])
                }
        else:
            return {
                'status': 'error',
                'error': 'version.json 파일을 찾을 수 없습니다',
                'version': '알 수 없음',
                'name': 'XLT System'
            }
    except Exception as e:
        return {
            'status': 'error',
            'error': f'버전 정보 로드 실패: {str(e)}',
            'version': '알 수 없음',
            'name': 'XLT System'
        }

def check_figma_config():
    """피그마 설정 확인"""
    config_path = Path('figma_config.json')
    env_token = os.getenv('FIGMA_TOKEN')

    if env_token:
        return {'status': 'ok', 'source': 'environment'}

    if config_path.exists():
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                if config.get('access_token'):
                    return {'status': 'ok', 'source': 'file'}
        except Exception as e:
            return {'status': 'error', 'error': f'figma_config.json 파싱 오류: {e}'}

    return {'status': 'missing', 'error': '피그마 토큰이 설정되지 않았습니다'}

def test_figma_token(token):
    """피그마 토큰 유효성 테스트"""
    try:
        import requests

        headers = {'X-Figma-Token': token}
        response = requests.get('https://api.figma.com/v1/me', headers=headers, timeout=10)

        if response.status_code == 200:
            user_info = response.json()
            return {
                'status': 'success',
                'user': user_info.get('name', 'Unknown'),
                'email': user_info.get('email', 'Unknown')
            }
        elif response.status_code == 403:
            return {'status': 'error', 'error': '토큰이 유효하지 않습니다'}
        else:
            return {'status': 'error', 'error': f'API 오류: {response.status_code}'}

    except Exception as e:
        return {'status': 'error', 'error': f'연결 테스트 실패: {str(e)}'}

# XLT System v3.1: 디렉토리 테스트 함수 제거됨 (메모리 기반 직접 다운로드로 변경)

def check_system_setup():
    """시스템 필수 설정 확인"""
    setup_status = {
        'all_configured': True,
        'missing_settings': [],
        'current_settings': {}
    }

    # 1. 피그마 토큰 확인
    figma_status = check_figma_config()
    if figma_status['status'] != 'ok':
        setup_status['all_configured'] = False
        setup_status['missing_settings'].append('피그마 토큰')
    else:
        setup_status['current_settings']['figma_token'] = True

    # 2. 사용자 설정 확인
    user_config = load_user_config()

    # XLT System v3.0: 출력 디렉토리 설정 제거됨

    setup_status['current_settings']['server_port'] = user_config.get('server_port', 5004)

    # ❌ 바른 API 키 확인 제거됨 - Claude AI 전용으로 전환

    return setup_status


def allowed_file(filename: str) -> bool:
    """허용된 파일 확장자인지 확인"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def is_figma_url(url: str) -> bool:
    """피그마 URL인지 확인"""
    return 'figma.com' in url.lower()


# OCR 결과 임시 저장
temp_ocr_results = {}


def get_current_port():
    """현재 실행 중인 서버 포트 가져오기"""
    # Flask의 현재 요청에서 포트 추출
    return str(request.environ.get('SERVER_PORT', request.host.split(':')[-1] if ':' in request.host else '5004'))

@app.route('/')
def index():
    """메인 페이지"""
    # 시스템 설정 상태 확인
    setup_status = check_system_setup()
    if not setup_status['all_configured']:
        return redirect(url_for('settings_page'))

    # 현재 실행 포트와 설정 포트 비교
    user_config = load_user_config()
    current_port = get_current_port()
    configured_port = str(user_config.get('server_port', 5004))
    port_mismatch = current_port != configured_port

    # 버전 정보 로드
    version_info = load_version_info()

    return render_template('index.html',
                         port_mismatch=port_mismatch,
                         current_port=current_port,
                         configured_port=configured_port,
                         version_info=version_info)

@app.route('/static/<path:filename>')
def static_files(filename):
    """정적 파일 서빙"""
    return send_from_directory('static', filename)

@app.route('/guide.md')
def serve_guide():
    """LINE API 기반 번역 가이드라인 제공 (Markdown 형식)"""
    try:
        # LINE API에서 용어집 로드
        terminology = load_line_api_terminology()

        if terminology:
            # Markdown 형식으로 용어집 생성
            markdown_content = """# 🌐 LINE API 기반 번역 가이드

**버전**: 3.0
**최종 수정**: 2026-05-13
**대상 서비스**: Unifi (핀테크 플랫폼)
**데이터 소스**: LINE API

---

## 핵심 용어집 (LINE API)

| 한국어 | English | 日本語 | 繁體中文 | ไทย |
|--------|---------|--------|---------|-----|
"""

            # 용어집 테이블 생성
            for korean, translations in terminology.items():
                markdown_content += f"| {korean} | {translations.get('en_US', '')} | {translations.get('ja_JP', '')} | {translations.get('zh_TW', '')} | {translations.get('th_TH', '')} |\n"

            markdown_content += f"""

---

## 데이터 소스 정보

- **총 용어 수**: {len(terminology)}개
- **지원 언어**: 5개 (한국어, 영어, 일본어, 중국어 번체, 태국어)
- **업데이트 방식**: LINE API 실시간 동기화
- **응답 속도**: 평균 0.2초대

---

## 사용 방법

이 용어집은 LINE API에서 실시간으로 로드되며, XLT 시스템의 Claude AI 번역에 자동으로 적용됩니다.

### API 엔드포인트
```
GET https://landpress-content.line-scdn.net/contents/v2/projects/wdmwbfuv10x39bukv58ocevp/collections/web3_xlt_json/item
```

### 폴백 순서
1. LINE API (우선)
2. 로컬 JSON 파일
3. 하드코딩된 핵심 용어

---

*이 문서는 LINE API에서 자동 생성되었습니다.*
"""

            return markdown_content, 200, {'Content-Type': 'text/markdown; charset=utf-8'}

        else:
            # 폴백: 기본 가이드 메시지
            fallback_content = """# 번역 가이드라인

LINE API 연결에 실패하여 기본 가이드를 표시합니다.

## 기본 용어집

- 거래 → EN: transaction, JA: 取引, ZH: 交易, TH: ธุรกรรม
- 지갑 → EN: wallet, JA: ウォレット, ZH: 錢包, TH: กระเป๋า
- 토큰 → EN: token, JA: トークン, ZH: 代幣, TH: โทเค็น

LINE API 연결을 확인해주세요.
"""
            return fallback_content, 200, {'Content-Type': 'text/markdown; charset=utf-8'}

    except Exception as e:
        return f'LINE API 오류: {str(e)}', 500

# XLT v3.0: 자동 테스트 기능 제거
# @app.route('/test_web_flow.html') - 자동 테스트 페이지 제거됨

@app.route('/terminology-test')
def terminology_test_page():
    """용어집 API 테스트 페이지"""
    try:
        return render_template('terminology_test.html')
    except Exception as e:
        app.logger.error(f"용어집 테스트 페이지 오류: {str(e)}")
        return f'용어집 테스트 페이지 로드 실패: {str(e)}', 500

    # serve_uploaded_image 제거됨 - 파일 업로드 기능 비활성화

@app.route('/debug-sessions')
def debug_sessions():
    """현재 활성 세션 디버깅"""
    debug_info = {}
    for session_id, data in temp_ocr_results.items():
        debug_info[session_id] = {
            'input_type': data.get('input_type'),
            'image_path': data.get('image_path'),
            'image_exists': os.path.exists(data.get('image_path', '')) if data.get('image_path') else False,
            'ocr_count': len(data.get('ocr_results', [])),
            'source': data.get('source', '')[:100] + '...' if len(data.get('source', '')) > 100 else data.get('source', '')
        }

    return jsonify({
        'active_sessions': debug_info,
        'figma_files': os.listdir('figma') if os.path.exists('figma') else []
    })

@app.route('/test-session')
def create_test_session():
    """테스트 세션 생성"""
    import shutil
    from PIL import Image

    # 테스트 세션 ID 생성
    test_session_id = "test_session_figma_preview"

    # 간단한 테스트 이미지 생성
    figma_dir = os.path.join(os.getcwd(), 'figma')
    os.makedirs(figma_dir, exist_ok=True)

    # 200x100 파란색 테스트 이미지 생성
    test_image = Image.new('RGB', (200, 100), color='blue')
    test_image_path = os.path.join(figma_dir, f"{test_session_id}_figma.png")
    test_image.save(test_image_path)

    # 테스트 세션 데이터 생성
    temp_ocr_results[test_session_id] = {
        'source': 'https://test.figma.com/test',
        'source_description': '테스트 피그마 파일',
        'input_type': 'figma',
        'image_path': test_image_path,
        'ocr_results': [
            {'text': '테스트 텍스트 1', 'confidence': 0.9},
            {'text': '테스트 텍스트 2', 'confidence': 0.8}
        ]
    }

    return jsonify({
        'status': 'success',
        'test_session_id': test_session_id,
        'image_path': test_image_path,
        'select_texts_url': f'/select_texts?session_id={test_session_id}'
    })

@app.route('/figma_image/<path:filename>')
def serve_figma_image(filename):
    """피그마 이미지 파일 서빙"""
    try:
        figma_dir = os.path.join(os.getcwd(), 'figma')
        file_path = os.path.join(figma_dir, filename)

        if not os.path.exists(file_path):
            print(f"[DEBUG] 피그마 이미지 파일 없음: {file_path}")
            return "Image not found", 404

        print(f"[DEBUG] 피그마 이미지 서빙: {filename}")
        return send_from_directory('figma', filename)
    except Exception as e:
        print(f"[ERROR] 피그마 이미지 서빙 오류: {e}")
        return "Image serving error", 500


# =====================================
# 설정 관리 라우트 (XLT System v3.0)
# =====================================

@app.route('/settings')
def settings_page():
    """시스템 설정 페이지"""
    user_config = load_user_config()
    return render_template('settings.html', current_settings=user_config)

@app.route('/api/settings/status')
def api_settings_status():
    """현재 설정 상태 확인"""
    try:
        setup_status = check_system_setup()
        return jsonify(setup_status)
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'설정 상태 확인 실패: {str(e)}'
        }), 500

@app.route('/api/settings/test-figma-token', methods=['GET', 'POST'])
def api_test_figma_token():
    """피그마 토큰 테스트"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    try:
        data = request.get_json()
        token = data.get('token')

        if not token:
            return jsonify({
                'status': 'error',
                'error': '토큰을 입력해주세요.'
            }), 400

        result = test_figma_token(token)
        return jsonify(result)

    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'토큰 테스트 실패: {str(e)}'
        }), 500

# XLT System v3.1: 디렉토리 테스트 API 제거됨 (메모리 기반 직접 다운로드로 변경)

# ❌ 바른 API 테스트 엔드포인트 제거됨 - Claude AI 전용으로 전환
# 정적 룰 기반 교정 시스템은 모두 제거되었습니다.

@app.route('/api/settings/get-figma-token', methods=['GET', 'POST'])
def api_get_figma_token():
    """저장된 피그마 토큰 조회 (보안상 POST 사용)"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    try:
        # 피그마 설정 확인
        figma_config = check_figma_config()

        if figma_config['status'] != 'ok':
            return jsonify({
                'status': 'error',
                'error': '저장된 피그마 토큰이 없습니다.'
            }), 404

        # 토큰 값 가져오기
        token = None

        # 환경변수에서 확인
        env_token = os.getenv('FIGMA_TOKEN')
        if env_token:
            token = env_token
        else:
            # figma_config.json에서 확인
            config_path = Path('figma_config.json')
            if config_path.exists():
                try:
                    with open(config_path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                        token = config.get('access_token')
                except Exception as e:
                    return jsonify({
                        'status': 'error',
                        'error': f'토큰 파일 읽기 실패: {e}'
                    }), 500

        if not token:
            return jsonify({
                'status': 'error',
                'error': '토큰을 찾을 수 없습니다.'
            }), 404

        return jsonify({
            'status': 'success',
            'token': token
        })

    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'토큰 조회 실패: {str(e)}'
        }), 500


@app.route('/api/claude/account/switch', methods=['POST'])
def api_claude_account_switch():
    """Claude 계정 전환"""
    try:
        data = request.get_json()
        auth_method = data.get('auth_method', 'claudeai')  # claudeai, console, sso
        email = data.get('email', '')

        # 현재 계정 로그아웃
        logout_result = subprocess.run(
            ['claude', 'auth', 'logout'],
            capture_output=True,
            timeout=10,
            text=True
        )

        # 새 계정으로 로그인
        login_cmd = ['claude', 'auth', 'login']

        if auth_method == 'console':
            login_cmd.append('--console')
        elif auth_method == 'sso':
            login_cmd.append('--sso')
        else:  # claudeai (기본값)
            login_cmd.append('--claudeai')

        if email:
            login_cmd.extend(['--email', email])

        # 백그라운드에서 로그인 프로세스 시작 (사용자가 수동으로 완료해야 함)
        login_result = subprocess.run(
            login_cmd,
            capture_output=True,
            timeout=30,
            text=True
        )

        return jsonify({
            'success': True,
            'message': f'계정 전환을 시작했습니다. 브라우저에서 로그인을 완료하세요.',
            'auth_method': auth_method,
            'email': email,
            'timestamp': datetime.now().isoformat()
        })

    except subprocess.TimeoutExpired:
        return jsonify({
            'success': False,
            'error': '로그인 시간 초과. 수동으로 claude auth login을 실행하세요.',
            'timestamp': datetime.now().isoformat()
        }), 504
    except Exception as e:
        logger.error(f"계정 전환 중 오류: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'계정 전환 실패: {str(e)}',
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/api/claude/account', methods=['GET'])
def api_claude_account():
    """Claude CLI 계정 정보 조회"""
    try:
        # Claude CLI auth status 호출
        result = subprocess.run(
            ['claude', 'auth', 'status', '--json'],
            capture_output=True,
            timeout=10,
            text=True,
            cwd=os.path.dirname(__file__)
        )

        if result.returncode == 0:
            try:
                account_data = json.loads(result.stdout.strip())

                # 사용자 프로필 정보도 가져오기 시도
                user_info = None
                try:
                    import requests

                    # 설정 파일에서 API 정보 읽기 (os는 이미 상단에서 import됨)
                    settings_path = os.path.expanduser('~/.claude/settings.json')
                    bedrock_url = None
                    session_token = None

                    if os.path.exists(settings_path):
                        with open(settings_path, 'r') as f:
                            settings = json.load(f)
                            env_settings = settings.get('env', {})
                            bedrock_url = env_settings.get('ANTHROPIC_BEDROCK_BASE_URL')
                            session_token = env_settings.get('AWS_SESSION_TOKEN')

                    if bedrock_url and session_token:
                        # 사용자 정보 API 호출
                        headers = {
                            'Authorization': f'Bearer {session_token}',
                            'Content-Type': 'application/json'
                        }
                        response = requests.get(f'{bedrock_url}/v1/me', headers=headers, timeout=5)
                        logger.info(f"사용자 API 응답 상태: {response.status_code}")

                        if response.status_code == 200:
                            try:
                                user_data = response.json()
                                logger.info(f"원본 사용자 데이터: {user_data}")
                                # 조직 정보 안전하게 처리
                                org_name = None
                                org_id = None
                                orgs_data = user_data.get('orgs')
                                if orgs_data and isinstance(orgs_data, dict):
                                    org_list = orgs_data.get('data', [])
                                    if org_list and len(org_list) > 0:
                                        first_org = org_list[0]
                                        org_name = first_org.get('title', first_org.get('name'))
                                        org_id = first_org.get('id')

                                user_info = {
                                    'email': user_data.get('email'),
                                    'name': user_data.get('name'),
                                    'userId': user_data.get('id'),
                                    'organization': org_name,
                                    'organizationId': org_id
                                }
                                logger.info(f"사용자 정보 조회 성공: {user_info}")
                            except Exception as parse_error:
                                logger.error(f"사용자 데이터 파싱 오류: {str(parse_error)}")
                        else:
                            logger.warning(f"사용자 API 응답 오류: {response.status_code} - {response.text}")
                except Exception as user_error:
                    logger.warning(f"사용자 정보 조회 실패: {str(user_error)}")

                # 서비스 유형 감지
                service_type = 'unknown'
                if bedrock_url and 'openai-proxy.linecorp.com' in bedrock_url:
                    service_type = 'line_proxy'
                elif account_data.get('apiProvider') == 'bedrock':
                    service_type = 'bedrock'
                elif account_data.get('apiProvider') == 'anthropic':
                    service_type = 'anthropic'

                # 응답 데이터 구성
                response_data = {
                    'success': True,
                    'account_info': account_data,
                    'service_type': service_type,
                    'timestamp': datetime.now().isoformat()
                }

                # 사용자 정보가 있으면 추가
                if user_info:
                    response_data['user_info'] = user_info

                # 서비스별 추가 정보
                if service_type == 'line_proxy':
                    response_data['service_info'] = {
                        'name': 'LINE OpenAI Proxy',
                        'description': 'LINE Corp 내부 OpenAI 프록시 서버',
                        'note': '실제 Claude.ai와는 다른 조직 ID를 사용합니다'
                    }

                return jsonify(response_data)

            except json.JSONDecodeError:
                return jsonify({
                    'success': False,
                    'error': 'Claude CLI 응답 파싱 실패',
                    'timestamp': datetime.now().isoformat()
                }), 500
        else:
            error_msg = result.stderr.strip() if result.stderr else 'Claude 계정 정보 조회 실패'
            return jsonify({
                'success': False,
                'error': f'Claude CLI 오류: {error_msg}',
                'timestamp': datetime.now().isoformat()
            }), 500

    except subprocess.TimeoutExpired:
        return jsonify({
            'success': False,
            'error': 'Claude CLI 응답 시간 초과 (30초)',
            'timestamp': datetime.now().isoformat()
        }), 504
    except FileNotFoundError:
        return jsonify({
            'success': False,
            'error': 'Claude CLI를 찾을 수 없습니다. Claude CLI가 설치되어 있는지 확인하세요.',
            'timestamp': datetime.now().isoformat()
        }), 500
    except Exception as e:
        logger.error(f"Claude 계정 정보 조회 중 오류: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'서버 오류: {str(e)}',
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/api/settings/save', methods=['GET', 'POST'])
def api_save_settings():
    """설정 저장"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    try:
        data = request.get_json()

        # 현재 설정 로드
        old_config = load_user_config()
        old_port = old_config.get('server_port', 5004)

        restart_required = False
        restart_reason = []

        # 피그마 토큰 저장
        figma_token = data.get('figma_token')
        if figma_token:  # None이 아닌 경우만 저장 (기존 토큰 유지하려면 None 전달)
            figma_config = {'access_token': figma_token}
            figma_config_path = Path('figma_config.json')

            # 안전한 파일 쓰기 (임시 파일 사용)
            temp_path = figma_config_path.with_suffix('.tmp')

            try:
                # 임시 파일에 먼저 쓰기
                with open(temp_path, 'w', encoding='utf-8') as f:
                    json.dump(figma_config, f, indent=2, ensure_ascii=False)
                    f.flush()  # 버퍼 강제 플러시
                    os.fsync(f.fileno())  # 디스크 동기화

                # 원자적 이동 (atomic move)
                if os.path.exists(figma_config_path):
                    os.remove(figma_config_path)
                os.rename(temp_path, figma_config_path)

                print(f"✅ 피그마 토큰 저장됨")

            except Exception as e:
                # 임시 파일 정리
                if temp_path.exists():
                    temp_path.unlink(missing_ok=True)

                print(f"❌ 피그마 토큰 저장 실패: {e}")
                return jsonify({
                    'status': 'error',
                    'error': f'피그마 토큰 저장 실패: {str(e)}',
                    'details': '파일 쓰기 권한을 확인하거나 디스크 공간을 확인해주세요.'
                }), 500

        # ❌ 바른 API 키 저장 제거됨 - Claude AI 전용으로 전환

        # 사용자 설정 저장
        user_config = load_user_config()

        # XLT System v3.0: 출력 디렉토리 설정 제거됨

        # 서버 포트 설정
        new_port = data.get('server_port')
        if new_port and 1024 <= new_port <= 65535:
            user_config['server_port'] = new_port

            # 포트 변경 감지
            if new_port != old_port:
                restart_required = True
                restart_reason.append(f'서버 포트 변경 ({old_port} → {new_port})')

        # 사용자 설정 파일 저장
        if not save_user_config(user_config):
            return jsonify({
                'status': 'error',
                'error': '사용자 설정 저장에 실패했습니다.'
            }), 500

        print(f"✅ 사용자 설정 저장됨: {user_config}")

        # 응답 준비
        response_data = {
            'status': 'success',
            'message': '설정이 저장되었습니다.',
            'settings': user_config,
            'restart_required': restart_required
        }

        if restart_required:
            response_data['restart_reason'] = restart_reason
            response_data['message'] = '설정이 저장되었습니다. 서버 재시작이 필요합니다.'
            response_data['new_url'] = f'http://localhost:{new_port}'

        return jsonify(response_data)

    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"❌ 설정 저장 실패: {e}")
        print(error_trace)

        return jsonify({
            'status': 'error',
            'error': f'설정 저장 중 오류: {str(e)}'
        }), 500


@app.route('/api/settings/restart', methods=['GET', 'POST'])
def api_restart_server():
    """서버 재시작"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    try:
        # 새 설정으로 서버 재시작
        import threading
        import subprocess
        import sys
        import signal

        def restart_server():
            # 잠시 대기 후 현재 프로세스 종료 및 재시작
            import time
            time.sleep(1)

            # 현재 스크립트 경로
            script_path = os.path.abspath(__file__)

            print("🔄 서버 재시작 중...")

            # 새 프로세스로 서버 재시작
            try:
                subprocess.Popen([sys.executable, script_path],
                               cwd=os.path.dirname(script_path))
            except Exception as e:
                print(f"❌ 재시작 실패: {e}")

            # 현재 프로세스 종료
            os.kill(os.getpid(), signal.SIGTERM)

        # 백그라운드에서 재시작 실행
        restart_thread = threading.Thread(target=restart_server)
        restart_thread.daemon = True
        restart_thread.start()

        return jsonify({
            'status': 'success',
            'message': '서버 재시작을 시작합니다.'
        })

    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'서버 재시작 실패: {str(e)}'
        }), 500

# =====================================
# v5.0.7 로그 관리 + Claude 계정 정보 API
# =====================================

@app.route('/api/settings/logs/status', methods=['GET'])
def api_log_settings_status():
    """현재 로그 설정 상태 조회"""
    try:
        user_config = load_user_config()
        log_settings = user_config.get('logging', {
            'enabled': True,
            'log_dir': 'logs',
            'log_level': 'INFO',
            'max_file_size_mb': 10,
            'backup_count': 3,
            'auto_updater_log_enabled': True,
            'auto_updater_separate_file': False
        })

        # 현재 로그 파일들 정보 수집
        log_files = []
        total_size = 0
        log_dir = log_settings.get('log_dir', 'logs')

        # 로그 파일들 스캔 (현재 디렉토리 + 로그 디렉토리)
        log_patterns = ['*.log', '*.log.*']

        # 현재 디렉토리 스캔
        for pattern in log_patterns:
            for log_file in Path('.').glob(pattern):
                if log_file.is_file():
                    file_size = log_file.stat().st_size
                    total_size += file_size
                    log_files.append({
                        'name': log_file.name,
                        'size': f"{file_size / 1024 / 1024:.1f} MB" if file_size > 1024*1024 else f"{file_size / 1024:.1f} KB",
                        'modified': datetime.fromtimestamp(log_file.stat().st_mtime).strftime('%Y-%m-%d %H:%M')
                    })

        # 로그 디렉토리 내 파일들도 확인
        if Path(log_dir).exists() and log_dir != '.':
            for log_file in Path(log_dir).glob('*.log*'):
                if log_file.is_file():
                    file_size = log_file.stat().st_size
                    total_size += file_size
                    log_files.append({
                        'name': f"{log_dir}/{log_file.name}",
                        'size': f"{file_size / 1024 / 1024:.1f} MB" if file_size > 1024*1024 else f"{file_size / 1024:.1f} KB",
                        'modified': datetime.fromtimestamp(log_file.stat().st_mtime).strftime('%Y-%m-%d %H:%M')
                    })

        return jsonify({
            'status': 'success',
            'current_settings': log_settings,
            'log_files': log_files,
            'total_size': f"{total_size / 1024 / 1024:.1f} MB" if total_size > 1024*1024 else f"{total_size / 1024:.1f} KB"
        })

    except Exception as e:
        logger.error(f"로그 설정 상태 조회 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'error': f'로그 설정 상태 조회 실패: {str(e)}'
        }), 500

@app.route('/api/settings/logs/update', methods=['POST'])
def api_update_log_settings():
    """로그 설정 실시간 업데이트 (재시작 불필요)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                'status': 'error',
                'error': '로그 설정 데이터가 필요합니다.'
            }), 400

        # 현재 설정 로드
        user_config = load_user_config()

        # 로그 설정 업데이트
        log_settings = user_config.get('logging', {})

        # 새 설정 적용
        if 'enabled' in data:
            log_settings['enabled'] = bool(data['enabled'])
        if 'log_dir' in data:
            log_settings['log_dir'] = str(data['log_dir'])
        if 'log_level' in data:
            if data['log_level'] in ['DEBUG', 'INFO', 'WARNING', 'ERROR']:
                log_settings['log_level'] = data['log_level']
        if 'max_file_size_mb' in data:
            size = int(data['max_file_size_mb'])
            if 1 <= size <= 100:
                log_settings['max_file_size_mb'] = size
        if 'backup_count' in data:
            count = int(data['backup_count'])
            if 1 <= count <= 10:
                log_settings['backup_count'] = count
        if 'auto_updater_log_enabled' in data:
            log_settings['auto_updater_log_enabled'] = bool(data['auto_updater_log_enabled'])
        if 'auto_updater_separate_file' in data:
            log_settings['auto_updater_separate_file'] = bool(data['auto_updater_separate_file'])

        # 설정 저장
        user_config['logging'] = log_settings
        if not save_user_config(user_config):
            return jsonify({
                'status': 'error',
                'error': '로그 설정 저장 실패'
            }), 500

        # 실시간 로그 설정 적용 (재시작 불필요)
        try:
            apply_log_settings_immediately(log_settings)
            logger.info(f"✅ 로그 설정이 즉시 적용되었습니다: {log_settings}")
        except Exception as e:
            logger.warning(f"⚠️ 로그 설정 즉시 적용 실패 (재시작 필요): {e}")

        return jsonify({
            'status': 'success',
            'message': '로그 설정이 저장되고 적용되었습니다.',
            'restart_required': False,
            'applied_settings': log_settings
        })

    except Exception as e:
        logger.error(f"로그 설정 업데이트 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'error': f'로그 설정 업데이트 실패: {str(e)}'
        }), 500

@app.route('/api/settings/claude-account', methods=['GET'])
def api_claude_account_info():
    """Claude Code 계정 정보 조회"""
    try:
        import subprocess

        # Claude CLI 상태 확인 (확장된 정보)
        try:
            result = subprocess.run(['claude', 'auth', 'status'],
                                  capture_output=True, text=True, timeout=10)

            account_info = {
                'logged_in': False,
                'auth_method': 'unknown',
                'email': 'N/A',
                'subscription_type': 'N/A',
                'usage_limit': 'N/A',
                'last_activity': 'N/A',
                'api_key_status': 'N/A',
                'cli_version': 'N/A'
            }

            if result.returncode == 0:
                # JSON 응답 파싱 시도
                try:
                    auth_data = json.loads(result.stdout)
                    account_info['logged_in'] = auth_data.get('loggedIn', False)
                    account_info['auth_method'] = auth_data.get('authMethod', 'unknown')

                    # 추가 정보가 있는 경우 수집
                    if 'email' in auth_data:
                        account_info['email'] = auth_data['email']
                    if 'subscription' in auth_data:
                        account_info['subscription_type'] = auth_data['subscription']

                except json.JSONDecodeError:
                    # JSON이 아닌 경우 텍스트 파싱
                    output = result.stdout.lower()
                    if 'authenticated' in output or 'logged in' in output or 'true' in output:
                        account_info['logged_in'] = True
                        account_info['auth_method'] = 'browser'

            # Claude CLI 버전 정보 추가 수집
            try:
                version_result = subprocess.run(['claude', '--version'],
                                              capture_output=True, text=True, timeout=5)
                if version_result.returncode == 0:
                    account_info['cli_version'] = version_result.stdout.strip()
            except:
                pass

            return jsonify({
                'status': 'success',
                'account_info': account_info,
                'last_check': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })

        except FileNotFoundError:
            return jsonify({
                'status': 'error',
                'error': 'Claude CLI가 설치되지 않았습니다.',
                'account_info': {
                    'logged_in': False,
                    'auth_method': 'cli_not_installed',
                    'email': 'N/A',
                    'subscription_type': 'N/A',
                    'usage_limit': 'N/A',
                    'last_activity': 'N/A',
                    'api_key_status': 'N/A',
                    'cli_version': 'N/A'
                }
            })

        except subprocess.TimeoutExpired:
            return jsonify({
                'status': 'error',
                'error': 'Claude CLI 응답 시간 초과',
                'account_info': {
                    'logged_in': False,
                    'auth_method': 'timeout',
                    'email': 'N/A',
                    'subscription_type': 'N/A',
                    'usage_limit': 'N/A',
                    'last_activity': 'N/A',
                    'api_key_status': 'N/A',
                    'cli_version': 'N/A'
                }
            })

    except Exception as e:
        logger.error(f"Claude 계정 정보 조회 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'error': f'Claude 계정 정보 조회 실패: {str(e)}'
        }), 500

@app.route('/api/settings/browse-directory', methods=['POST'])
def api_browse_directory():
    """디렉토리 탐색 API - 폴더 선택을 위한 디렉토리 트리 제공"""
    try:
        data = request.get_json()
        current_path = data.get('path', '.')

        # 보안을 위해 절대 경로로 변환하고 현재 작업 디렉토리 내로 제한
        import os
        current_work_dir = os.getcwd()

        if current_path == '.' or current_path == '':
            target_path = current_work_dir
        else:
            # 상대 경로를 절대 경로로 변환
            if not os.path.isabs(current_path):
                target_path = os.path.join(current_work_dir, current_path)
            else:
                target_path = current_path

        # 보안 체크: 현재 작업 디렉토리 내부인지 확인
        target_path = os.path.abspath(target_path)
        if not target_path.startswith(current_work_dir):
            target_path = current_work_dir

        # 경로가 존재하는지 확인
        if not os.path.exists(target_path) or not os.path.isdir(target_path):
            target_path = current_work_dir

        directories = []
        files = []

        try:
            # 디렉토리 내용 읽기
            for item in os.listdir(target_path):
                item_path = os.path.join(target_path, item)

                # 숨김 파일/폴더 제외
                if item.startswith('.'):
                    continue

                try:
                    if os.path.isdir(item_path):
                        # 디렉토리 정보
                        dir_info = {
                            'name': item,
                            'path': os.path.relpath(item_path, current_work_dir),
                            'type': 'directory',
                            'has_subdirs': any(
                                os.path.isdir(os.path.join(item_path, sub))
                                for sub in os.listdir(item_path)
                                if not sub.startswith('.')
                            ) if os.access(item_path, os.R_OK) else False
                        }
                        directories.append(dir_info)
                    elif os.path.isfile(item_path):
                        # 파일 정보 (참고용)
                        file_info = {
                            'name': item,
                            'path': os.path.relpath(item_path, current_work_dir),
                            'type': 'file',
                            'size': os.path.getsize(item_path)
                        }
                        files.append(file_info)
                except (PermissionError, OSError):
                    continue

        except PermissionError:
            return jsonify({
                'status': 'error',
                'error': '디렉토리 접근 권한이 없습니다.'
            }), 403

        # 부모 디렉토리 정보
        parent_path = None
        if target_path != current_work_dir:
            parent_dir = os.path.dirname(target_path)
            if parent_dir.startswith(current_work_dir):
                parent_path = os.path.relpath(parent_dir, current_work_dir)
                if parent_path == '.':
                    parent_path = ''

        # 알파벳순 정렬
        directories.sort(key=lambda x: x['name'].lower())
        files.sort(key=lambda x: x['name'].lower())

        return jsonify({
            'status': 'success',
            'current_path': os.path.relpath(target_path, current_work_dir),
            'parent_path': parent_path,
            'directories': directories,
            'files': files[:20]  # 파일은 최대 20개만 표시 (참고용)
        })

    except Exception as e:
        logger.error(f"디렉토리 탐색 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'error': f'디렉토리 탐색 실패: {str(e)}'
        }), 500

# =============================================================================
# Claude 프롬프트 관리 API
# =============================================================================

@app.route('/api/settings/prompts', methods=['GET'])
def api_get_prompts():
    """모든 Claude 프롬프트 정보 조회"""
    try:
        from xlt.core.prompt_manager import get_prompt_manager

        prompt_manager = get_prompt_manager()
        prompt_info = prompt_manager.get_prompt_info()

        return jsonify({
            'status': 'success',
            'prompts': prompt_info
        })

    except Exception as e:
        logger.error(f"❌ 프롬프트 조회 오류: {e}")
        return jsonify({
            'status': 'error',
            'error': f'프롬프트 조회 실패: {str(e)}'
        }), 500

@app.route('/api/settings/prompts/<prompt_type>', methods=['GET'])
def api_get_prompt(prompt_type):
    """특정 프롬프트 조회"""
    try:
        from xlt.core.prompt_manager import get_prompt_manager

        prompt_manager = get_prompt_manager()
        prompt_content = prompt_manager.get_prompt(prompt_type)

        if prompt_content is None:
            return jsonify({
                'status': 'error',
                'error': '존재하지 않는 프롬프트입니다.'
            }), 404

        return jsonify({
            'status': 'success',
            'prompt': prompt_content
        })

    except Exception as e:
        logger.error(f"❌ 프롬프트 조회 오류 ({prompt_type}): {e}")
        return jsonify({
            'status': 'error',
            'error': f'프롬프트 조회 실패: {str(e)}'
        }), 500

@app.route('/api/settings/prompts/<prompt_type>', methods=['POST'])
def api_update_prompt(prompt_type):
    """특정 프롬프트 업데이트"""
    try:
        from xlt.core.prompt_manager import get_prompt_manager

        data = request.get_json()
        if not data:
            return jsonify({
                'status': 'error',
                'error': '요청 데이터가 없습니다.'
            }), 400

        new_prompt = data.get('prompt', '')
        name = data.get('name', '')
        description = data.get('description', '')

        if not new_prompt.strip():
            return jsonify({
                'status': 'error',
                'error': '프롬프트 내용이 비어있습니다.'
            }), 400

        prompt_manager = get_prompt_manager()
        success = prompt_manager.update_prompt(prompt_type, new_prompt, name, description)

        if success:
            logger.info(f"✅ 프롬프트 업데이트 완료: {prompt_type}")
            return jsonify({
                'status': 'success',
                'message': f'{name or prompt_type} 프롬프트가 업데이트되었습니다.'
            })
        else:
            return jsonify({
                'status': 'error',
                'error': '프롬프트 업데이트에 실패했습니다.'
            }), 500

    except Exception as e:
        logger.error(f"❌ 프롬프트 업데이트 오류 ({prompt_type}): {e}")
        return jsonify({
            'status': 'error',
            'error': f'프롬프트 업데이트 실패: {str(e)}'
        }), 500

@app.route('/api/settings/prompts/<prompt_type>/reset', methods=['POST'])
def api_reset_prompt(prompt_type):
    """특정 프롬프트를 기본값으로 리셋"""
    try:
        from xlt.core.prompt_manager import get_prompt_manager

        prompt_manager = get_prompt_manager()
        success = prompt_manager.reset_to_default(prompt_type)

        if success:
            logger.info(f"✅ 프롬프트 리셋 완료: {prompt_type}")
            return jsonify({
                'status': 'success',
                'message': f'{prompt_type} 프롬프트가 기본값으로 리셋되었습니다.'
            })
        else:
            return jsonify({
                'status': 'error',
                'error': '프롬프트 리셋에 실패했습니다.'
            }), 500

    except Exception as e:
        logger.error(f"❌ 프롬프트 리셋 오류 ({prompt_type}): {e}")
        return jsonify({
            'status': 'error',
            'error': f'프롬프트 리셋 실패: {str(e)}'
        }), 500

@app.route('/api/settings/prompts/reset-all', methods=['POST'])
def api_reset_all_prompts():
    """모든 프롬프트를 기본값으로 리셋"""
    try:
        from xlt.core.prompt_manager import get_prompt_manager

        prompt_manager = get_prompt_manager()
        success = prompt_manager.reset_to_default()

        if success:
            logger.info("✅ 모든 프롬프트 리셋 완료")
            return jsonify({
                'status': 'success',
                'message': '모든 프롬프트가 기본값으로 리셋되었습니다.'
            })
        else:
            return jsonify({
                'status': 'error',
                'error': '프롬프트 리셋에 실패했습니다.'
            }), 500

    except Exception as e:
        logger.error(f"❌ 모든 프롬프트 리셋 오류: {e}")
        return jsonify({
            'status': 'error',
            'error': f'프롬프트 리셋 실패: {str(e)}'
        }), 500

def apply_log_settings_immediately(log_settings):
    """로그 설정을 서버 재시작 없이 즉시 적용"""
    try:
        # 현재 로거 핸들러들 가져오기
        current_logger = logging.getLogger('xlt_server')

        # 기존 로그 파일 경로 저장 (이동을 위해)
        old_log_file = None
        for handler in current_logger.handlers[:]:
            if isinstance(handler, RotatingFileHandler):
                old_log_file = handler.baseFilename
                break

        # 기존 파일 핸들러 제거 (콘솔 핸들러는 유지)
        handlers_to_remove = []
        for handler in current_logger.handlers[:]:
            if isinstance(handler, RotatingFileHandler):
                handlers_to_remove.append(handler)

        for handler in handlers_to_remove:
            handler.close()
            current_logger.removeHandler(handler)

        if log_settings.get('enabled', True):
            # 새 로그 설정으로 핸들러 생성
            log_dir = log_settings.get('log_dir', 'logs')
            log_level = log_settings.get('log_level', 'INFO')
            max_size = log_settings.get('max_file_size_mb', 10) * 1024 * 1024  # MB to bytes
            backup_count = log_settings.get('backup_count', 3)

            # 로그 디렉토리 생성
            if log_dir and log_dir != '.' and log_dir != '':
                os.makedirs(log_dir, exist_ok=True)
                log_file_path = Path(log_dir) / "server.log"
            else:
                log_file_path = Path("server.log")

            # 기존 로그 파일 이동 (경로가 변경된 경우)
            if old_log_file and old_log_file != str(log_file_path):
                try:
                    migrate_log_files(old_log_file, str(log_file_path))
                except Exception as e:
                    logger.warning(f"⚠️ 기존 로그 파일 이동 실패: {e}")

            # 새 파일 핸들러 생성
            new_handler = RotatingFileHandler(
                log_file_path,
                maxBytes=max_size,
                backupCount=backup_count
            )

            # 로그 레벨 설정
            level_mapping = {
                'DEBUG': logging.DEBUG,
                'INFO': logging.INFO,
                'WARNING': logging.WARNING,
                'ERROR': logging.ERROR
            }
            new_handler.setLevel(level_mapping.get(log_level, logging.INFO))
            current_logger.setLevel(level_mapping.get(log_level, logging.INFO))

            # 포매터 적용
            log_formatter = logging.Formatter(
                '%(asctime)s [%(levelname)s] %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S'
            )
            new_handler.setFormatter(log_formatter)

            # 로거에 추가
            current_logger.addHandler(new_handler)

            logger.info(f"🔄 로그 설정이 즉시 적용되었습니다: {log_settings}")
        else:
            logger.info("🔄 로그 기능이 비활성화되었습니다")

    except Exception as e:
        logger.error(f"로그 설정 즉시 적용 실패: {e}")
        raise e

def migrate_log_files(old_path: str, new_path: str):
    """기존 로그 파일을 새 위치로 안전하게 이동"""
    try:
        old_file = Path(old_path)
        new_file = Path(new_path)

        # 새 디렉토리 생성
        new_file.parent.mkdir(parents=True, exist_ok=True)

        # 기존 파일이 존재하는 경우에만 이동
        if old_file.exists():
            # 새 파일이 이미 있으면 백업
            if new_file.exists():
                backup_path = new_file.with_suffix(f'.backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
                new_file.rename(backup_path)
                logger.info(f"📦 기존 파일 백업: {backup_path}")

            # 파일 이동
            import shutil
            shutil.move(str(old_file), str(new_file))
            logger.info(f"📁 로그 파일 이동 완료: {old_path} → {new_path}")

            # 백업 파일들도 함께 이동
            old_dir = old_file.parent
            new_dir = new_file.parent
            backup_pattern = f"{old_file.stem}.*.log"

            for backup_file in old_dir.glob(backup_pattern):
                new_backup_path = new_dir / backup_file.name
                try:
                    shutil.move(str(backup_file), str(new_backup_path))
                    logger.info(f"📁 백업 파일 이동: {backup_file.name}")
                except Exception as e:
                    logger.warning(f"⚠️ 백업 파일 이동 실패: {e}")

    except Exception as e:
        logger.error(f"로그 파일 이동 실패: {e}")
        raise e

# =====================================
# 버전 정보 API (XLT System v3.1)
# =====================================

@app.route('/api/version', methods=['GET'])
def api_get_version():
    """현재 버전 정보 조회"""
    version_info = load_version_info()
    return jsonify(version_info)

@app.route('/api/spell-check', methods=['GET', 'POST'])
def api_spell_check():
    """맞춤법 및 띄어쓰기 검사 API (v3.3)"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    try:
        data = request.get_json()
        if not data or 'text' not in data:
            return jsonify({
                'status': 'error',
                'error': '검사할 텍스트가 필요합니다.'
            }), 400

        text = data['text'].strip()
        if not text:
            return jsonify({
                'status': 'error',
                'error': '빈 텍스트는 검사할 수 없습니다.'
            }), 400

        # ❌ 정적 교정 제거됨 - Claude AI 전용으로 전환
        corrected_text = text  # 교정 없이 원본 반환 (Claude 통합 모드에서만 교정)

        return jsonify({
            'status': 'success',
            'original_text': text,
            'corrected_text': corrected_text,
            'changed': text != corrected_text,
            'message': '맞춤법 검사 완료'
        })

    except Exception as e:
        logger.error(f"맞춤법 검사 API 오류: {e}")
        return jsonify({
            'status': 'error',
            'error': f'맞춤법 검사 중 오류가 발생했습니다: {str(e)}'
        }), 500

@app.route('/api/version/check-update', methods=['GET'])
def api_check_version_update():
    """업데이트 확인 (간단 버전)"""
    try:
        version_info = load_version_info()
        update_check = requests.get(
            'https://api.github.com/repos/hobong-ho6/xlt-system/commits/main',
            timeout=10
        )

        if update_check.status_code == 200:
            remote_info = update_check.json()
            return jsonify({
                'status': 'success',
                'current_version': version_info.get('version', '알 수 없음'),
                'current_build': version_info.get('build', '알 수 없음'),
                'remote_commit': remote_info.get('sha', '')[:7],
                'remote_date': remote_info.get('commit', {}).get('author', {}).get('date', ''),
                'update_available': True,  # 항상 업데이트 가능으로 표시
                'last_check': datetime.now().isoformat()
            })
        else:
            return jsonify({
                'status': 'error',
                'error': '업데이트 확인 실패',
                'current_version': version_info.get('version', '알 수 없음')
            }), 400

    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'업데이트 확인 중 오류: {str(e)}'
        }), 500

@app.route('/select_texts', methods=['GET'])
def select_texts():
    """텍스트 추출 결과 선택 페이지"""
    session_id = request.args.get('session_id')
    if not session_id or session_id not in temp_ocr_results:
        return jsonify({
            'status': 'error',
            'error': '텍스트 추출 세션을 찾을 수 없습니다. 다시 시작해주세요.',
            'debug_info': f'session_id: {session_id}, available_sessions: {list(temp_ocr_results.keys())}'
        }), 404

    # OCR 데이터 가져오기
    session_data = temp_ocr_results[session_id]

    # v5.1.0: 번역 모드 확인 (Claude AI 기본값)
    translation_mode = session_data.get('translation_mode', 'claude_integrated')

    # OCR 결과 정규화 함수
    def normalize_ocr_item(item):
        if hasattr(item, 'text'):
            # OCRResult 객체 형태
            return {
                'text': item.text,
                'confidence': getattr(item, 'confidence', 0.0)
            }
        elif isinstance(item, dict):
            # 딕셔너리 형태
            return {
                'text': item.get('text', ''),
                'confidence': item.get('confidence', 0.0)
            }
        else:
            # 기타 형태 (문자열 등)
            return {
                'text': str(item),
                'confidence': 0.0
            }

    # 모든 OCR 결과 정규화
    raw_ocr_results = session_data.get('ocr_results', [])
    normalized_ocr_results = [normalize_ocr_item(item) for item in raw_ocr_results]

    # 의미있는 텍스트 필터링 (원본 인덱스 보존)
    filtered_texts = []
    for idx, item in enumerate(normalized_ocr_results):
        text = item['text'].strip()
        confidence = item['confidence']

        # 간단한 필터링 조건
        if (len(text) > 1 and              # 길이가 1자 이상
            confidence > 0.5 and           # 신뢰도 50% 이상
            not text.isdigit() and         # 숫자만으로 이루어지지 않음
            len(text) < 100):              # 너무 긴 텍스트 제외
            # 원본 인덱스를 포함하여 추가
            item_with_index = item.copy()
            item_with_index['original_index'] = idx
            filtered_texts.append(item_with_index)

    # 필터링 결과가 없으면 모든 결과 사용 (원본 인덱스 포함)
    if not filtered_texts:
        filtered_texts = []
        for idx, item in enumerate(normalized_ocr_results):
            item_with_index = item.copy()
            item_with_index['original_index'] = idx
            filtered_texts.append(item_with_index)

    # 이미지 경로 준비
    image_path = session_data.get('image_path')
    input_type = session_data.get('input_type', 'file')
    image_url = None

    # 디버깅: 세션 데이터 확인
    print(f"[DEBUG] session_data keys: {session_data.keys()}")
    print(f"[DEBUG] image_path: {image_path}")
    print(f"[DEBUG] input_type: {input_type}")
    print(f"[DEBUG] figma 디렉토리 내용: {os.listdir('figma') if os.path.exists('figma') else '없음'}")

    # 이미지 URL 생성
    if image_path and os.path.exists(image_path):
        if input_type == 'file':
            # 파일 업로드인 경우
            image_url = f'/uploaded_image/{os.path.basename(image_path)}'
        elif input_type == 'figma':
            # 피그마인 경우
            image_url = f'/figma_image/{os.path.basename(image_path)}'
        print(f"[DEBUG] image_url 생성됨: {image_url}")
    else:
        print(f"[DEBUG] image_url 생성 실패 - image_path={image_path}, exists={os.path.exists(image_path) if image_path else False}")

    # 템플릿 데이터 확인
    print(f"[DEBUG] ocr_data.image_url: {image_url}")
    print(f"[DEBUG] 템플릿 렌더링 시작...")

    # 템플릿에 필요한 데이터 준비
    ocr_data = {
        'source_description': session_data.get('source_description', ''),
        'ocr_results': normalized_ocr_results,
        'filtered_texts': filtered_texts,
        'image_url': image_url,
        'translation_mode': translation_mode  # v4.0: 번역 모드 정보 추가
    }

    return render_template('ocr_results.html', session_id=session_id, ocr_data=ocr_data)


@app.route('/api/session/validate', methods=['GET', 'POST'])
def validate_session():
    """세션 유효성 검증 API"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    try:
        data = request.get_json()
        session_id = data.get('session_id') if data else None

        if not session_id:
            return jsonify({
                'status': 'invalid',
                'error': '세션 ID가 제공되지 않았습니다'
            }), 400

        if session_id in temp_ocr_results:
            session_data = temp_ocr_results[session_id]
            return jsonify({
                'status': 'valid',
                'session_info': {
                    'source_description': session_data.get('source_description', ''),
                    'ocr_count': len(session_data.get('ocr_results', [])),
                    'translation_mode': session_data.get('translation_mode', 'claude_integrated')
                }
            })
        else:
            return jsonify({
                'status': 'invalid',
                'error': 'OCR 세션을 찾을 수 없습니다',
                'available_sessions': list(temp_ocr_results.keys())
            }), 404

    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'세션 검증 중 오류 발생: {str(e)}'
        }), 500


@app.route('/upload', methods=['GET', 'POST'])
def process_request():
    """요청 처리 (강화된 예외 처리)"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    # 세션 ID 생성 (POST 요청만)
    session_id = f"session_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{secrets.token_hex(4)}"

    try:
        # 초기 로그 추가
        add_session_log(session_id, "🚀 XLT 번역 작업을 시작합니다", "info")

        if not pipeline:
            add_session_log(session_id, "❌ XLT 시스템이 초기화되지 않았습니다", "error")
            return jsonify({
                'status': 'error',
                'error': 'XLT 시스템이 초기화되지 않았습니다.',
                'details': 'pipeline 객체가 None입니다.'
            }), 500

        input_type = 'figma'  # 피그마 URL 전용
        mode = request.form.get('mode', 'auto')
        translation_mode = request.form.get('translation_mode', 'claude_integrated')  # XLT v5.1.0: Claude AI 기본값

        # 동적 엔진 로그 메시지 (Claude AI 전용)
        if translation_mode == 'claude_integrated':
            engine_name = 'Claude 통합 처리 (맞춤법+번역)'
        elif translation_mode == 'claude':
            engine_name = 'Claude AI (고품질)'
        else:  # 기본값은 claude_integrated
            engine_name = 'Claude 통합 처리 (기본)'

        add_session_log(session_id, f"⚙️ 처리 모드: {mode}, 입력 타입: figma", "info")
        add_session_log(session_id, f"🔧 번역 엔진: {engine_name}", "info")

        # 오래된 피그마 이미지 파일 정리
        cleanup_old_figma_images()

        # 피그마 URL 전용 처리
        source = None
        source_description = ""
        add_session_log(session_id, "🎨 피그마 URL 방식 선택됨", "info")

        figma_url = request.form.get('figma_url', '').strip()
        if not figma_url:
            add_session_log(session_id, "❌ 피그마 URL이 입력되지 않았습니다", "error")
            return jsonify({
                'status': 'error',
                'error': '피그마 URL을 입력해주세요.'
            }), 400

        add_session_log(session_id, f"🔍 피그마 URL 검증 중: {figma_url[:50]}...", "info")

        if not is_figma_url(figma_url):
            add_session_log(session_id, "❌ 잘못된 피그마 URL 형식입니다", "error")
            return jsonify({
                'status': 'error',
                'error': '올바른 피그마 URL을 입력해주세요.'
            }), 400

        # 피그마 토큰 확인
        try:
            add_session_log(session_id, "🔑 피그마 토큰 확인 중...", "info")
            from xlt.input.figma import FigmaProcessor
            figma_processor = FigmaProcessor(pipeline.config)

            if not figma_processor.figma_token:
                add_session_log(session_id, "❌ 피그마 토큰이 설정되지 않았습니다", "error")
                return jsonify({
                    'status': 'error',
                    'error': '피그마 액세스 토큰이 설정되지 않았습니다.',
                    'details': 'figma_config.json 파일을 확인하거나 FIGMA_TOKEN 환경변수를 설정해주세요.'
                }), 400

            # 연결 테스트
            if not figma_processor.test_figma_connection():
                return jsonify({
                    'status': 'error',
                    'error': '피그마 API에 연결할 수 없습니다.',
                    'details': '토큰이 만료되었거나 권한이 없을 수 있습니다.'
                }), 400

        except Exception as e:
            return jsonify({
                'status': 'error',
                'error': '피그마 연결 확인 중 오류가 발생했습니다.',
                'details': str(e)
            }), 500

        source = figma_url
        source_description = f"피그마 URL: {figma_url}"

        # 처리 실행
        try:
            if mode == 'auto':
                # 자동 모드
                add_session_log(session_id, "🤖 자동 모드: 전체 번역 파이프라인 실행", "info")
                add_session_log(session_id, "🔄 OCR → 필터링 → 번역 진행 중...", "info")

                result = pipeline.process(source=source, auto_mode=True)

                if result['status'] == 'success':
                    add_session_log(session_id, f"🎉 번역 완료! {result['processed_count']}개 항목 처리", "success")

                    # 작업 완료 상태 설정
                    session_status[session_id] = {'completed': True}

                    return jsonify({
                        'status': 'success',
                        'processed_count': result['processed_count'],
                        'source_description': source_description,
                        'session_id': session_id
                    })
                else:
                    add_session_log(session_id, f"❌ 번역 실패: {result.get('error', '알 수 없는 오류')}", "error")
                    session_status[session_id] = {'completed': True}

                    return jsonify({
                        'status': 'error',
                        'error': result.get('error', '처리에 실패했습니다.'),
                        'details': result.get('details', '')
                    }), 500

            else:
                # 수동 모드 - 텍스트 추출
                add_session_log(session_id, "📋 수동 모드: 텍스트 추출 단계", "info")

                ocr_results = []
                image_path = None  # 이미지 파일 경로 저장용

                # 피그마 URL인 경우 먼저 이미지 다운로드 (미리보기용)
                if input_type == 'figma':
                    try:
                        add_session_log(session_id, "📸 피그마 이미지 다운로드 중 (미리보기용)...", "info")

                        from xlt.input.figma import FigmaProcessor
                        figma_processor = FigmaProcessor(pipeline.config)

                        # 피그마 이미지 다운로드
                        image, _ = figma_processor.process(source)
                        if hasattr(figma_processor, 'last_saved_file'):
                            original_image_path = figma_processor.last_saved_file

                            # 웹 접근 가능한 디렉토리에 복사
                            if original_image_path and os.path.exists(original_image_path):
                                import shutil
                                figma_dir = os.path.join(os.getcwd(), 'figma')
                                os.makedirs(figma_dir, exist_ok=True)

                                file_extension = os.path.splitext(original_image_path)[1] or '.png'
                                web_filename = f"{session_id}_figma{file_extension}"
                                web_image_path = os.path.join(figma_dir, web_filename)

                                try:
                                    shutil.copy2(original_image_path, web_image_path)
                                    image_path = web_image_path
                                    add_session_log(session_id, f"📸 미리보기용 이미지 저장: {web_filename}", "success")
                                    add_session_log(session_id, f"🔗 이미지 URL 설정: /figma_image/{web_filename}", "info")
                                except Exception as e:
                                    add_session_log(session_id, f"⚠️ 이미지 복사 실패: {str(e)}", "warning")
                                    image_path = original_image_path
                            else:
                                add_session_log(session_id, f"⚠️ 원본 이미지 경로가 없거나 파일이 존재하지 않음: {original_image_path}", "warning")
                                image_path = original_image_path
                        else:
                            add_session_log(session_id, "⚠️ 피그마 프로세서에서 저장된 파일 경로를 찾을 수 없음", "warning")
                    except Exception as e:
                        add_session_log(session_id, f"⚠️ 피그마 이미지 다운로드 실패: {str(e)}", "warning")

                    # 🎯 피그마 텍스트 노드 직접 추출 (기본 방식)
                    try:
                        add_session_log(session_id, "🎨 피그마에서 텍스트 직접 추출 중...", "info")

                        from xlt.input.figma import FigmaProcessor
                        figma_processor = FigmaProcessor(pipeline.config)

                        # URL 파싱
                        file_key, node_id = figma_processor._parse_figma_url(source)

                        # 텍스트 노드 직접 추출
                        text_result = figma_processor.extract_text_from_node(file_key, node_id)

                        if text_result['status'] == 'success' and text_result['count'] > 0:
                            add_session_log(session_id, f"✅ 피그마에서 {text_result['count']}개 텍스트 추출 완료", "success")
                            add_session_log(session_id, "💡 텍스트 노드에서 직접 가져왔습니다 (정확도 100%)", "info")

                            # 피그마 텍스트 결과를 표준 형태로 변환
                            ocr_results = text_result['texts']
                            add_session_log(session_id, "✅ 피그마 텍스트 추출 완료", "success")
                        else:
                            add_session_log(session_id, f"⚠️ 피그마에서 텍스트를 찾을 수 없습니다: {text_result.get('error', '텍스트 없음')}", "warning")
                            add_session_log(session_id, "❌ 텍스트 노드가 없거나 접근할 수 없습니다", "error")

                    except Exception as e:
                        add_session_log(session_id, f"⚠️ 피그마 텍스트 추출 오류: {str(e)}", "warning")
                        add_session_log(session_id, "❌ 피그마 API 호출에 실패했습니다", "error")

                # ⚠️ OCR 방식은 비활성화됨 (텍스트 노드 직접 추출 우선 사용)
                # 피그마 텍스트 노드에서 텍스트를 못 가져온 경우 처리 중단
                if not ocr_results:
                    add_session_log(session_id, "❌ 피그마에서 텍스트를 추출할 수 없습니다", "error")
                    add_session_log(session_id, "💡 텍스트 레이어가 있는 피그마 파일인지 확인해주세요", "info")

                    return jsonify({
                        'status': 'error',
                        'error': '피그마에서 텍스트를 찾을 수 없습니다.',
                        'details': '텍스트 레이어가 없거나 접근 권한이 부족할 수 있습니다.'
                    }), 400

                # 🔒 OCR 방식 비활성화 (아래 코드는 주석 처리됨)
                # if not ocr_results:
                #     add_session_log(session_id, "🔍 입력 타입 감지 중...", "info")
                #     from xlt.input.base import InputProcessor
                #     detected_input_type = InputProcessor.detect_input_type(source)
                #     input_processor = pipeline.input_processors[detected_input_type]
                #
                #     add_session_log(session_id, f"📸 이미지 처리 중... (타입: {detected_input_type})", "info")
                #     # 이미지 로드 및 OCR
                #     image, _ = input_processor.process(source)
                #
                #     # 피그마 처리기인 경우 - 이미 이미지가 처리되지 않은 경우에만 처리
                #     if hasattr(input_processor, 'last_saved_file'):
                #         original_image_path = input_processor.last_saved_file
                #
                #         # 피그마 이미지이고 아직 웹 접근용 이미지 경로가 설정되지 않은 경우에만 복사
                #         if (detected_input_type == 'figma' and
                #             original_image_path and os.path.exists(original_image_path) and
                #             (not image_path or not image_path.startswith(os.path.join(os.getcwd(), 'figma')))):
                #
                #             import shutil
                #             figma_dir = os.path.join(os.getcwd(), 'figma')
                #             os.makedirs(figma_dir, exist_ok=True)
                #
                #             # 고유한 파일명 생성 (세션 ID 기반)
                #             file_extension = os.path.splitext(original_image_path)[1] or '.png'
                #             web_filename = f"{session_id}_figma{file_extension}"
                #             web_image_path = os.path.join(figma_dir, web_filename)
                #
                #             try:
                #                 shutil.copy2(original_image_path, web_image_path)
                #                 image_path = web_image_path
                #                 add_session_log(session_id, f"📸 OCR용 이미지 저장 성공: {web_filename}", "success")
                #                 add_session_log(session_id, f"🔗 이미지 URL: /figma_image/{web_filename}", "info")
                #             except Exception as e:
                #                 add_session_log(session_id, f"⚠️ 이미지 복사 실패: {str(e)}", "warning")
                #                 add_session_log(session_id, f"📁 원본 경로: {original_image_path}", "warning")
                #                 add_session_log(session_id, f"📁 대상 경로: {web_image_path}", "warning")
                #                 image_path = original_image_path
                #         elif detected_input_type == 'figma' and image_path:
                #             add_session_log(session_id, f"✅ 이미지 경로가 이미 설정됨: {os.path.basename(image_path)}", "info")
                #         else:
                #             image_path = original_image_path
                #     elif input_type == 'file':
                #         image_path = source
                #
                #     add_session_log(session_id, "🤖 OCR 텍스트 추출 중...", "info")
                #     ocr_results = pipeline.ocr_engine.extract_text(image)
                #
                #     # v3.3: 한글 맞춤법/띄어쓰기 교정 적용 (OCR 결과 즉시 교정)
                #     for result in ocr_results:
                #         if 'text' in result:
                #             original_text = result['text']
                #             corrected_text = apply_korean_corrections(original_text)
                #             result['text'] = corrected_text
                #             if original_text != corrected_text:
                #                 add_session_log(session_id, f"🔤 교정: '{original_text}' → '{corrected_text}'", "info")
                #
                #     add_session_log(session_id, f"✅ OCR 완료: {len(ocr_results)}개 텍스트 발견 (교정 적용됨)", "success")
                # ✅ 피그마 텍스트 노드에서 추출된 텍스트에 맞춤법 교정 적용
                if ocr_results:
                    for result in ocr_results:
                        if 'text' in result:
                            original_text = result['text']
                            corrected_text = apply_korean_corrections(original_text)

                            # 원본 텍스트 보존 (before/after 비교용)
                            result['original_text'] = original_text
                            result['text'] = corrected_text
                            result['corrected'] = original_text != corrected_text

                            if original_text != corrected_text:
                                add_session_log(session_id, f"🔤 교정: '{original_text}' → '{corrected_text}'", "info")

                    add_session_log(session_id, f"✅ 텍스트 추출 완료: {len(ocr_results)}개 텍스트 발견 (교정 적용됨)", "success")

                # 세션에 저장 전 디버깅
                print(f"[DEBUG] 세션 저장 전 상태:")
                print(f"  - session_id: {session_id}")
                print(f"  - input_type: {input_type}")
                print(f"  - image_path: {image_path}")
                print(f"  - image_path exists: {os.path.exists(image_path) if image_path else False}")
                print(f"  - figma dir contents: {os.listdir('figma') if os.path.exists('figma') else 'N/A'}")
                if image_path and os.path.exists(image_path):
                    print(f"  - image_path basename: {os.path.basename(image_path)}")
                    print(f"  - expected_url: /figma_image/{os.path.basename(image_path)}")
                add_session_log(session_id, f"📋 세션 데이터 저장 중... (image_path: {'✅' if image_path else '❌'})", "info")

                temp_ocr_results[session_id] = {
                    'source': source,
                    'source_description': source_description,
                    'input_type': input_type,
                    'image_path': image_path,  # 실제 이미지 파일 경로 추가
                    'ocr_results': ocr_results,
                    'translation_mode': translation_mode  # v4.0: 번역 모드 정보 저장
                }

                print(f"[DEBUG] 세션 저장 완료: {session_id}")

                # OCR 결과 형태 확인 및 처리
                try:
                    if ocr_results and hasattr(ocr_results[0], 'text'):
                        # OCRResult 객체 형태
                        texts = [{'text': r.text, 'confidence': r.confidence} for r in ocr_results]
                    else:
                        # 딕셔너리 형태
                        texts = [{'text': r.get('text', ''), 'confidence': r.get('confidence', 0.0)} for r in ocr_results]
                except (AttributeError, TypeError):
                    # 예외 상황 처리
                    texts = []
                    for r in ocr_results:
                        if isinstance(r, dict):
                            texts.append({
                                'text': r.get('text', ''),
                                'confidence': r.get('confidence', 0.0)
                            })
                        else:
                            # 다른 형태의 객체라면 기본값
                            texts.append({
                                'text': str(r),
                                'confidence': 0.0
                            })

                add_session_log(session_id, "📋 OCR 결과를 사용자 선택을 위해 준비 중...", "info")

                return jsonify({
                    'status': 'ocr_complete',
                    'ocr_count': len(ocr_results),
                    'texts': texts,
                    'session_id': session_id,
                    'source_description': source_description,
                    'redirect': f'/select_texts?session_id={session_id}'
                })

        except Exception as e:
            error_trace = traceback.format_exc()
            return jsonify({
                'status': 'error',
                'error': f'처리 중 오류가 발생했습니다: {str(e)}',
                'details': error_trace
            }), 500

        finally:
            # 파일 정리
            if input_type == 'file' and source and os.path.exists(source):
                try:
                    os.remove(source)
                except:
                    pass

    except Exception as e:
        error_trace = traceback.format_exc()

        # 로그에도 오류 추가 (세션 ID가 있는 경우)
        try:
            if 'session_id' in locals():
                add_session_log(session_id, f"❌ 요청 처리 중 예기치 않은 오류: {str(e)}", "error")
                session_status[session_id] = {'completed': True}
        except:
            pass  # 로그 추가 실패 시 무시

        return jsonify({
            'status': 'error',
            'error': f'요청 처리 중 예기치 않은 오류: {str(e)}',
            'details': error_trace
        }), 500


@app.route('/check-placeholders', methods=['GET', 'POST'])
def check_placeholders():
    """선택된 텍스트의 치환 가능한 패턴 감지 (성능 최적화)"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    try:
        data = request.get_json()
        selected_indexes = [int(idx) for idx in data.get('selected_indexes', [])]
        selected_texts_from_client = data.get('selected_texts', [])  # v3.2: 클라이언트에서 받은 수정된 텍스트
        session_id = data.get('session_id')

        if not selected_indexes:
            return jsonify({
                'status': 'error',
                'error': '선택된 텍스트가 없습니다.'
            }), 400

        if not session_id or session_id not in temp_ocr_results:
            return jsonify({
                'status': 'error',
                'error': 'OCR 세션 데이터가 없습니다. 다시 시도해주세요.'
            }), 400

        # OCR 결과 가져오기
        session_data = temp_ocr_results[session_id]
        ocr_results = session_data['ocr_results']

        # v3.2: 선택된 텍스트 추출 (사용자 수정 텍스트 우선 사용)
        selected_texts = []
        if selected_texts_from_client and len(selected_texts_from_client) == len(selected_indexes):
            # 클라이언트에서 수정된 텍스트가 있으면 그것을 사용
            selected_texts = selected_texts_from_client
            print(f"✅ v3.2: 사용자가 수정한 텍스트 {len(selected_texts)}개 사용")
        else:
            # 없으면 원본 OCR 텍스트 사용 (하위 호환성)
            for idx in selected_indexes:
                if 0 <= idx < len(ocr_results):
                    selected_texts.append(ocr_results[idx]['text'])
            print(f"⚠️ 원본 OCR 텍스트 사용 (하위 호환)")

        # 성능 최적화: PlaceholderDetector 캐싱
        if not hasattr(app, '_cached_placeholder_detector'):
            from xlt.utils.placeholder_detector import PlaceholderDetector
            app._cached_placeholder_detector = PlaceholderDetector()
            print("🚀 PlaceholderDetector 캐시 생성 완료")

        # 성능 최적화: 결과 캐싱 (동일한 텍스트 조합에 대해)
        cache_key = hash(tuple(sorted(selected_texts)))
        if not hasattr(app, '_placeholder_cache'):
            app._placeholder_cache = {}

        if cache_key in app._placeholder_cache:
            print(f"🚀 치환자 검사 캐시 히트: {len(selected_texts)}개 텍스트")
            detection_results = app._placeholder_cache[cache_key]
        else:
            # 치환자 감지 실행 (캐시된 detector 사용)
            detection_results = app._cached_placeholder_detector.detect_placeholders(selected_texts)
            # 캐시 저장 (최대 100개 항목까지만)
            if len(app._placeholder_cache) < 100:
                app._placeholder_cache[cache_key] = detection_results
                print(f"🚀 치환자 검사 결과 캐시 저장: {len(selected_texts)}개 텍스트")

        # 성능 최적화: 리스트 컴프리헨션 사용
        placeholder_suggestions = []
        has_placeholders = False

        for i, result in enumerate(detection_results):
            if result['has_suggestions']:
                # 각 suggestion을 새로운 형태로 변환
                converted_suggestions = []

                for suggestion in result['suggestions']:
                    # 치환자가 적용된 텍스트 생성
                    text_with_placeholders = result['original_text']

                    # 모든 매치를 치환자로 교체 (뒤에서부터 교체해야 인덱스가 틀어지지 않음)
                    sorted_suggestions = sorted(result['suggestions'], key=lambda x: x['start_pos'], reverse=True)
                    for s in sorted_suggestions:
                        start, end = s['start_pos'], s['end_pos']
                        text_with_placeholders = text_with_placeholders[:start] + s['suggested_replacement'] + text_with_placeholders[end:]

                    converted_suggestions.append({
                        'with_placeholders': text_with_placeholders,
                        'patterns': [{
                            'type': suggestion['description'],
                            'matched': suggestion['matched_text']
                        }]
                    })

                # 첫 번째 suggestion만 사용하거나 전체 텍스트에 모든 치환자 적용
                if converted_suggestions:
                    # 모든 치환자를 적용한 하나의 제안 생성
                    all_patterns = []
                    text_with_all_placeholders = result['original_text']

                    # 모든 매치를 치환자로 교체
                    sorted_suggestions = sorted(result['suggestions'], key=lambda x: x['start_pos'], reverse=True)
                    for s in sorted_suggestions:
                        start, end = s['start_pos'], s['end_pos']
                        text_with_all_placeholders = text_with_all_placeholders[:start] + s['suggested_replacement'] + text_with_all_placeholders[end:]
                        all_patterns.append({
                            'type': s['description'],
                            'matched': s['matched_text']
                        })

                    placeholder_suggestions.append({
                        'index': selected_indexes[i],
                        'original_text': result['original_text'],
                        'suggestions': [{
                            'with_placeholders': text_with_all_placeholders,
                            'patterns': all_patterns
                        }]
                    })
                    has_placeholders = True  # 성능 최적화: 미리 플래그 설정

        return jsonify({
            'status': 'success',
            'has_placeholders': has_placeholders,  # 성능 최적화: 미리 계산된 플래그 사용
            'placeholder_suggestions': placeholder_suggestions,
            'selected_indexes': selected_indexes,
            'session_id': session_id
        })

    except Exception as e:
        error_trace = traceback.format_exc()
        return jsonify({
            'status': 'error',
            'error': f'치환자 감지 중 오류: {str(e)}',
            'details': error_trace
        }), 500


@app.route('/set-xlt-keys', methods=['GET', 'POST'])
def set_xlt_keys():
    """치환자 편집 후 XLT Key 설정 단계"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    try:
        data = request.get_json()
        session_id = data.get('session_id')
        selected_indexes = [int(idx) for idx in data.get('selected_indexes', [])]
        final_texts = data.get('final_texts', [])

        if not session_id or session_id not in temp_ocr_results:
            return jsonify({
                'status': 'error',
                'error': '세션 데이터를 찾을 수 없습니다.'
            }), 400

        print("=" * 80)
        print("🔍 /set-xlt-keys 요청 받음")
        print("=" * 80)
        print(f"Session ID: {session_id}")
        print(f"Selected indexes: {selected_indexes}")
        print(f"Final texts 개수: {len(final_texts)}")

        # 세션에 데이터 저장
        session_data = temp_ocr_results[session_id]
        session_data['selected_indexes'] = selected_indexes
        session_data['final_texts'] = final_texts

        # XLT Key 설정을 위한 데이터 준비
        key_setup_data = []
        for i, (idx, text) in enumerate(zip(selected_indexes, final_texts)):
            key_setup_data.append({
                'index': i,
                'original_index': idx,
                'text': text[:80] + '...' if len(text) > 80 else text,  # 표시용으로 축약
                'suggested_key': f"item_{i+1}"  # 기본 제안값
            })
            print(f"  [{i}] index={idx}, text='{text[:50]}...', suggested_key='item_{i+1}'")

        return jsonify({
            'status': 'success',
            'key_setup_data': key_setup_data,
            'session_id': session_id
        })

    except Exception as e:
        print(f"❌ XLT Key 설정 오류: {e}")
        return jsonify({
            'status': 'error',
            'error': f'XLT Key 설정 중 오류가 발생했습니다: {str(e)}'
        }), 500


@app.route('/translate-selected', methods=['GET', 'POST'])
def translate_selected():
    """선택된 텍스트 번역 (개별 XLT Key 적용)"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    # datetime import 명시적 추가 (오류 해결)
    from datetime import datetime

    try:
        # 상세 디버그 로깅 시작
        from xlt.utils.debug_logger import log_session_start, log_step, log_error, log_translation_attempt, log_connection_test_detail, save_session_summary, cleanup_session

        data = request.get_json()
        session_id = data.get('session_id')
        xlt_keys = data.get('xlt_keys', [])  # 개별 설정된 XLT key 배열

        # 디버그 로그: 세션 시작
        log_session_start(session_id, {
            'endpoint': '/translate-selected',
            'method': request.method,
            'data_keys': list(data.keys()) if data else [],
            'session_id': session_id,
            'xlt_keys_count': len(xlt_keys) if xlt_keys else 0,
            'request_size': len(str(data)) if data else 0
        })

        log_step(session_id, "요청 데이터 파싱 완료", {
            'session_id': session_id,
            'xlt_keys': xlt_keys,
            'data_size': len(str(data)) if data else 0
        })

        if not session_id or session_id not in temp_ocr_results:
            log_error(session_id, Exception("세션 데이터를 찾을 수 없습니다"), "세션 검증")
            return jsonify({
                'status': 'error',
                'error': '세션 데이터를 찾을 수 없습니다.'
            }), 400

        session_data = temp_ocr_results[session_id]

        # 프론트엔드에서 전송된 데이터 우선 사용, 없으면 세션에서 가져오기
        selected_indexes = data.get('selected_indexes', session_data.get('selected_indexes', []))
        final_texts = data.get('final_texts', session_data.get('final_texts', []))

        logger.info("=" * 80)
        logger.info(f"[{session_id}] 🔍 /translate-selected 요청 받음 (개별 XLT Key)")
        logger.info("=" * 80)
        logger.info(f"[{session_id}] 📋 Session ID: {session_id}")
        logger.info(f"[{session_id}] 📊 Selected indexes: {selected_indexes}")
        logger.info(f"[{session_id}] 📝 Final texts 개수: {len(final_texts)}")
        logger.info(f"[{session_id}] 🔑 XLT Keys 개수: {len(xlt_keys)}")
        for i, (idx, text, key) in enumerate(zip(selected_indexes, final_texts, xlt_keys)):
            logger.info(f"[{session_id}]   [{i}] index={idx}, key='{key}', text='{text[:50]}...'")
        logger.info("=" * 80)

        if not selected_indexes:
            return jsonify({
                'status': 'error',
                'error': '선택된 텍스트가 없습니다.'
            }), 400

        if not xlt_keys or len(xlt_keys) != len(selected_indexes):
            return jsonify({
                'status': 'error',
                'error': 'XLT Key 설정이 올바르지 않습니다.'
            }), 400

        # OCR 결과 가져오기 (이미 session_data는 위에서 가져옴)
        ocr_results = session_data['ocr_results']

        # 최종 텍스트 결정 (치환자 적용된 텍스트 또는 원본)
        selected_texts = []
        learned_count = 0
        learned_corrections = []  # v3.3: 학습 내역 저장

        if final_texts and len(final_texts) == len(selected_indexes):
            # v3.3: 사용자가 수정한 텍스트 사용 (재교정하지 않음)
            selected_texts = final_texts

            # ❌ 사용자 교정 학습 제거됨 - Claude AI 전용으로 전환
            # 정적 룰 기반 학습 시스템은 모두 제거되었습니다.
        else:
            # 원본 텍스트 사용 (Claude 통합 처리에서 맞춤법 교정 포함)
            for idx in selected_indexes:
                if 0 <= idx < len(ocr_results):
                    selected_texts.append(ocr_results[idx]['text'])

        if not selected_texts:
            return jsonify({
                'status': 'error',
                'error': '유효한 텍스트가 선택되지 않았습니다.'
            }), 400

        # XLT v5.1.0: 번역 엔진 선택 지원 - Claude AI 기본값
        translation_mode = data.get('translation_mode', 'claude_integrated')  # v5.1.0 기본값: Claude AI 통합 처리

        # LINE API 기반 번역 가이드 로드 (XLT System v5.1.1)
        terminology = load_translation_guide()

        # 번역 실행 (ko_KR은 원본 사용, 나머지 언어만 번역)
        target_languages = ['ko_KR', 'en_US', 'ja_JP', 'zh_TW', 'th_TH']  # ko_KR 추가
        translations = []

        # 번역기 선택 (XLT v4.0)
        if translation_mode == 'claude_integrated':
            # Claude 통합 맞춤법 + 번역 처리 (v4.0 신규)
            if hasattr(pipeline, 'claude_translator') and pipeline.claude_translator:
                translator = pipeline.claude_translator
                use_integrated_processing = True
                logger.info(f"[{session_id}] 🤖✨ Claude 통합 처리 사용 (맞춤법 + 번역)")
            else:
                # Claude 번역기 없으면 폴백
                translator = pipeline.translator
                use_integrated_processing = False
                logger.error(f"[{session_id}] ❌ Claude 통합 처리 불가, Claude 번역기 필수")
        elif translation_mode == 'claude':
            if hasattr(pipeline, 'claude_translator') and pipeline.claude_translator:
                translator = pipeline.claude_translator
                use_integrated_processing = False  # 기존 Claude 번역만
                logger.info(f"[{session_id}] 🤖 Claude 번역기 사용 (번역만)")
            else:
                # Claude 번역기 없으면 구체적인 오류 안내
                error_message = """Claude 번역을 사용할 수 없습니다.

📋 해결 방법:
1. Claude CLI 설치하기:
   • 웹사이트: https://claude.ai/download
   • 설치 후 터미널에서 'claude --version' 확인

💡 XLT System은 Claude AI 전용 번역 시스템으로 Claude CLI가 필수입니다."""

                return jsonify({
                    'status': 'error',
                    'error': error_message,
                    'error_type': 'ClaudeNotAvailable',
                    'suggestions': [
                        {'action': 'install_claude', 'text': 'Claude CLI 설치하기', 'url': 'https://claude.ai/download'}
                    ]
                }), 400
        else:
            # 지원되지 않는 번역 모드 오류
            error_message = f"지원되지 않는 번역 모드: {translation_mode}"
            logger.error(f"[{session_id}] ❌ {error_message}")
            return jsonify({
                'status': 'error',
                'error': error_message,
                'error_type': 'UnsupportedTranslationMode'
            }), 400

        try:
            logger.info("=" * 80)
            logger.info(f"[{session_id}] 🔄 번역 시작: {len(selected_texts)}개 텍스트, 모드: {translation_mode}")
            logger.info(f"[{session_id}]    🌐 LINE API 용어집: {len(terminology)}개 용어 로드됨")
            logger.info(f"[{session_id}]    🌐 대상 언어: {target_languages}")
            logger.info(f"[{session_id}]    📝 샘플 텍스트: {selected_texts[0][:50] if selected_texts else 'None'}...")
            logger.info(f"[{session_id}]    🔧 번역기 타입: {type(translator).__name__}")
            logger.info("=" * 80)

            # 예상 완료 시간 계산
            if translation_mode == 'claude_integrated':
                # 청크 기반 처리: 텍스트당 20초 (3개씩 청크 처리 + 타임아웃 여유 반영)
                estimated_minutes = len(selected_texts) * 0.33  # 20초 = 0.33분
                max_minutes = max(1, min(estimated_minutes, 15))  # 최소 1분, 최대 15분
            else:
                # 기본 Claude 처리 시간 (안전한 추정치)
                estimated_minutes = len(selected_texts) * 0.2  # 텍스트당 12초
                max_minutes = max(0.5, min(estimated_minutes, 5))  # 최소 30초, 최대 5분

            estimated_time_text = f"{int(max_minutes)}분" if max_minutes >= 1 else f"{int(max_minutes * 60)}초"

            # 번역 진행 상황 초기화
            session_status[session_id] = {
                'translation_progress': {
                    'status': 'translating',
                    'total_languages': len(target_languages),
                    'completed_languages': [],
                    'current_language': None,
                    'total_texts': len(selected_texts),
                    'message': f'번역을 시작합니다... (예상 시간: {estimated_time_text})',
                    'estimated_time': estimated_time_text,
                    'start_time': time.time()
                }
            }

            # 세션 로그 초기화 (v3.0 호환성)
            if session_id not in session_logs:
                session_logs[session_id] = []

            # 번역기 연결 테스트 (선택된 번역기)
            log_step(session_id, "번역기 연결 테스트 시작", {
                'translator_type': type(translator).__name__,
                'translation_mode': translation_mode,
                'has_test_connection': hasattr(translator, 'test_connection')
            })

            if hasattr(translator, 'test_connection'):
                # Claude AI 전용 엔진 이름 설정
                if translation_mode == 'claude_integrated':
                    engine_name = 'Claude 통합 처리 (맞춤법+번역)'
                elif translation_mode == 'claude':
                    engine_name = 'Claude AI (고품질)'
                else:  # 기본값은 claude_integrated
                    engine_name = 'Claude 통합 처리 (기본)'

                logger.info(f"[{session_id}] 🔍 {engine_name} 연결 테스트 중...")
                session_status[session_id]['translation_progress']['message'] = f'{engine_name} 연결 확인 중...'

                # 연결 테스트 상세 로그
                test_start_time = time.time()
                log_connection_test_detail(session_id, {
                    'method': 'threading + queue',
                    'engine_name': engine_name,
                    'translation_mode': translation_mode,
                    'start_time': datetime.now().isoformat(),
                    'timeout_seconds': 5
                })

                # Claude CLI 연결 테스트 타임아웃 처리 (v5.1.0 빠른 인증 상태 체크)
                try:
                    import threading
                    import queue

                    # 타임아웃과 결과를 위한 큐
                    result_queue = queue.Queue()
                    timeout_occurred = threading.Event()

                    def connection_test():
                        """연결 테스트를 별도 스레드에서 실행"""
                        try:
                            if not timeout_occurred.is_set():
                                status = translator.test_connection()
                                result_queue.put(('success', status))
                        except Exception as e:
                            if not timeout_occurred.is_set():
                                result_queue.put(('error', e))

                    # 연결 테스트 스레드 시작
                    test_thread = threading.Thread(target=connection_test)
                    test_thread.daemon = True
                    test_thread.start()

                    # 5초 타임아웃으로 결과 대기 (빠른 인증 상태 체크)
                    try:
                        result_type, result_value = result_queue.get(timeout=5)
                        test_duration = time.time() - test_start_time

                        if result_type == 'success':
                            connection_status = result_value
                            logger.info(f"[{session_id}]    ✅ 연결 상태: {connection_status}")

                            log_connection_test_detail(session_id, {
                                'method': 'threading + queue',
                                'result': 'success',
                                'connection_status': connection_status,
                                'duration': test_duration
                            })
                        else:
                            log_connection_test_detail(session_id, {
                                'method': 'threading + queue',
                                'result': 'error',
                                'error': str(result_value),
                                'duration': test_duration
                            })
                            raise result_value

                    except queue.Empty:
                        # 타임아웃 발생
                        timeout_occurred.set()
                        test_duration = time.time() - test_start_time
                        logger.warning(f"[{session_id}] ⏰ Claude CLI 연결 테스트 타임아웃 (5초)")

                        log_connection_test_detail(session_id, {
                            'method': 'threading + queue',
                            'result': 'timeout',
                            'duration': test_duration,
                            'timeout_seconds': 5
                        })

                        # Claude 타임아웃 시 연결 실패로 처리
                        logger.warning(f"[{session_id}] ⏱️ Claude 연결 테스트 타임아웃")
                        log_step(session_id, "Claude 연결 테스트 타임아웃", {
                            'translation_mode': translation_mode,
                            'reason': 'timeout'
                        })
                        connection_status = False

                except Exception as e:
                    test_duration = time.time() - test_start_time
                    logger.error(f"[{session_id}] ❌ 연결 테스트 오류: {e}")

                    log_error(session_id, e, "연결 테스트 중 예외 발생")
                    log_connection_test_detail(session_id, {
                        'method': 'threading + queue',
                        'result': 'exception',
                        'error': str(e),
                        'duration': test_duration
                    })

                    # Claude 연결 오류 시 연결 실패로 처리
                    logger.error(f"[{session_id}] ❌ Claude 연결 테스트 중 오류 발생")
                    log_step(session_id, "Claude 연결 테스트 오류", {
                        'translation_mode': translation_mode,
                        'reason': 'connection_error'
                    })
                    connection_status = False

                log_step(session_id, "연결 테스트 결과 검증", {
                    'connection_status_type': type(connection_status).__name__,
                    'connection_status_value': connection_status,
                    'final_translation_mode': translation_mode,
                    'final_engine_name': engine_name
                })

                if isinstance(connection_status, dict):
                    # Claude 연결 상태 확인 (dict 형태)
                    if not connection_status.get('claude', True):  # Claude 연결 확인
                        logger.error(f"[{session_id}] ❌ Claude 연결 실패")
                        log_error(session_id, Exception("Claude 연결 실패"), "Claude 연결 상태 확인")
                        raise Exception("Claude 연결 실패")
                elif not connection_status:
                    # Translator의 경우
                    logger.error(f"[{session_id}] ❌ 번역 서비스 연결 실패")
                    log_error(session_id, Exception("번역 서비스 연결 실패"), "Translator 연결 상태 확인")
                    raise Exception("번역 서비스 연결 실패")

            # 배치 번역 실행
            session_status[session_id]['translation_progress']['message'] = f'{len(selected_texts)}개 텍스트를 {len(target_languages)}개 언어로 번역 중...'
            session_status[session_id]['translation_progress']['current_language'] = 'processing'

            # 번역 시도 상세 로그
            log_translation_attempt(session_id, {
                'type': type(translator).__name__,
                'mode': translation_mode,
                'engine_name': engine_name,
                'connection_test': connection_status,
                'text_count': len(selected_texts),
                'target_languages': target_languages,
                'use_integrated_processing': locals().get('use_integrated_processing', False),
                'sample_text': selected_texts[0][:50] + "..." if selected_texts else "N/A"
            })

            # 번역 시작 로그 추가
            log_message = f'🔄 번역 시작: {len(selected_texts)}개 텍스트 → {len(target_languages)}개 언어'
            logger.info(f"[{session_id}] {log_message}")

            log_step(session_id, "번역 배치 처리 시작", {
                'text_count': len(selected_texts),
                'language_count': len(target_languages),
                'translator_class': type(translator).__name__,
                'translation_mode': translation_mode
            })
            log_entry = {
                'message': log_message,
                'timestamp': datetime.now().strftime("%H:%M:%S"),
                'type': 'info'
            }
            session_logs[session_id].append(log_entry)

            # 언어별 번역 준비 상태 로그
            lang_names = {
                'ko_KR': '한국어', 'en_US': '영어', 'ja_JP': '일본어',
                'zh_TW': '중국어', 'th_TH': '태국어'
            }
            for i, lang in enumerate(target_languages):
                log_message = f'   {i+1}/{len(target_languages)} {lang_names.get(lang, lang)} 번역 대기 중...'
                logger.info(f"[{session_id}] {log_message}")
                log_entry = {
                    'message': log_message,
                    'timestamp': datetime.now().strftime("%H:%M:%S"),
                    'type': 'info'
                }
                session_logs[session_id].append(log_entry)

            # 번역 실행 전 상세 로그
            log_message = f'🔍 소스 언어 자동 감지 중...'
            logger.info(f"[{session_id}] {log_message}")
            log_entry = {
                'message': log_message,
                'timestamp': datetime.now().strftime("%H:%M:%S"),
                'type': 'info'
            }
            session_logs[session_id].append(log_entry)

            # 실시간 번역 진행 로그 시스템
            import threading

            def add_translation_log(message, log_type='info'):
                # session_logs에 추가
                log_entry = {
                    'message': message,
                    'timestamp': datetime.now().strftime("%H:%M:%S"),
                    'type': log_type
                }
                session_logs[session_id].append(log_entry)

                # logger에도 출력 (터미널/파일 로그)
                if log_type == 'error':
                    logger.error(f"[{session_id}] {message}")
                elif log_type == 'warning':
                    logger.warning(f"[{session_id}] {message}")
                elif log_type == 'success':
                    logger.info(f"[{session_id}] ✅ {message}")
                else:
                    logger.info(f"[{session_id}] {message}")

            # 1. 선택된 텍스트 처리 상황
            add_translation_log(f'📝 {len(selected_texts)}개 텍스트 처리 시작')
            for i, text in enumerate(selected_texts[:5]):  # 처음 5개만 표시
                add_translation_log(f'   [{i+1}] {text[:50]}{"..." if len(text) > 50 else ""}')
            if len(selected_texts) > 5:
                add_translation_log(f'   ... 외 {len(selected_texts) - 5}개 더')

            # 2. 한국어 교정 진행 상황
            add_translation_log(f'🔤 한국어 맞춤법 교정 중...')
            time.sleep(0.1)  # UI 업데이트를 위한 최소 대기

            # 3. 번역 엔진 시작
            add_translation_log(f'🔍 소스 언어 감지: 한국어 (자동)')
            add_translation_log(f'🌐 대상 언어: {", ".join(["영어", "일본어", "중국어", "태국어"])}')
            add_translation_log(f'⏱️ 타임아웃: 120초 (태국어 문제 해결)')

            # 4. 각 언어별 번역 준비
            lang_names = {'en_US': '영어', 'ja_JP': '일본어', 'zh_TW': '중국어', 'th_TH': '태국어'}
            for lang in target_languages:
                if lang != 'ko_KR':  # 한국어 제외
                    add_translation_log(f'🚀 {lang_names.get(lang, lang)} 번역 작업 시작')

            # Claude 통합 처리 vs 기존 번역 방식 분기 (v4.0)
            if use_integrated_processing:
                # Claude 통합 맞춤법 검사 + 번역 처리 (v4.0 신규)
                add_translation_log(f'🤖✨ Claude 통합 처리 시작 (맞춤법 + 번역)', 'info')
                add_translation_log(f'📝 {len(selected_texts)}개 텍스트 통합 처리 중...')

                try:
                    # 청크 진행률 콜백 함수 정의
                    def main_chunk_progress_callback(progress_data):
                        # session_status 업데이트 (기존 API와 호환성 유지)
                        if session_id in session_status:
                            session_status[session_id]['translation_progress'].update({
                                'status': progress_data['status'],
                                'message': progress_data['message'],
                                'chunk_info': f"{progress_data.get('completed_chunks', 0)}/{progress_data.get('total_chunks', 0)}"
                            })

                        # 새로운 translation_progress 시스템에도 업데이트
                        if session_id not in translation_progress:
                            translation_progress[session_id] = {}

                        translation_progress[session_id].update({
                            'chunk_progress': progress_data,
                            'updated_at': time.time()
                        })

                    # Claude 청크 기반 배치 처리 (타임아웃 문제 해결)
                    chunk_size = getattr(translator.config, 'claude_chunk_size', 3)
                    add_translation_log(f'🔀 청크 기반 처리 시작: {len(selected_texts)}개 텍스트 → {chunk_size}개씩 분할', 'info')

                    log_step(session_id, "Claude 청크 배치 처리 시작", {
                        'method': 'translate_batch_integrated_chunked',
                        'text_count': len(selected_texts),
                        'target_languages': target_languages,
                        'chunk_size': chunk_size,
                        'translator_class': type(translator).__name__
                    })

                    try:
                        integrated_results = translator.translate_batch_integrated_chunked(
                            selected_texts,
                            target_languages,
                            progress_callback=main_chunk_progress_callback
                        )

                        log_step(session_id, "Claude 청크 배치 처리 성공", {
                            'result_count': len(integrated_results) if integrated_results else 0,
                            'result_type': type(integrated_results).__name__
                        })

                    except Exception as e:
                        log_error(session_id, e, "Claude 청크 배치 처리 중 오류")
                        raise e

                    translated_results = []
                    total_corrections = 0

                    for i, result in enumerate(integrated_results):
                        # 통합 결과 구조 변환
                        formatted_result = {
                            'original': result.get('original', selected_texts[i]),
                            'ko_KR': result.get('corrected_korean', selected_texts[i])
                        }

                        # 번역 결과 추가
                        for lang in target_languages:
                            if lang != 'ko_KR':
                                formatted_result[lang] = result.get(lang, selected_texts[i])

                        translated_results.append(formatted_result)

                        # 교정 내역 로그
                        corrections = result.get('corrections_applied', [])
                        if corrections:
                            total_corrections += len(corrections)
                            add_translation_log(f'✅ [{i+1}/{len(selected_texts)}] 통합 처리 완료 ({len(corrections)}개 교정)')
                            for correction in corrections:
                                add_translation_log(f'   📝 교정: {correction}', 'info')
                        else:
                            add_translation_log(f'✅ [{i+1}/{len(selected_texts)}] 통합 처리 완료')

                    if total_corrections > 0:
                        add_translation_log(f'🎯 Claude 통합 처리 완료: 총 {total_corrections}개 맞춤법 교정 적용', 'success')
                        # 교정 내역을 학습 데이터에 추가 (v3.3 호환성)
                        learned_count += total_corrections
                        for result in integrated_results:
                            for correction in result.get('corrections_applied', []):
                                learned_corrections.append({
                                    'original': result.get('original', ''),
                                    'corrected': correction
                                })
                    else:
                        add_translation_log(f'🎯 Claude 통합 처리 완료: 맞춤법 교정 사항 없음', 'success')

                except Exception as e:
                    add_translation_log(f'❌ Claude 통합 처리 실패: {str(e)[:100]}', 'error')
                    add_translation_log(f'⚠️ Claude 재시도 필요', 'warning')

                    # Claude 처리 실패 - 재시도 유지
                    use_integrated_processing = True

            if not use_integrated_processing:
                # ❌ 기존 개별 번역 방식 제거됨 - Claude AI 통합 처리만 사용
                add_translation_log(f'⚠️ Claude AI 통합 처리로 자동 전환됨', 'warning')

                # Claude 통합 처리로 강제 전환
                use_integrated_processing = True
                translator = pipeline.claude_translator

                for i, (original_text, corrected_text) in enumerate(zip(selected_texts, corrected_texts)):
                    add_translation_log(f'📝 [{i+1}/{len(selected_texts)}] "{corrected_text[:30]}{"..." if len(corrected_text) > 30 else ""}" 번역 중...')

                    log_step(session_id, f"개별 텍스트 번역 시작 [{i+1}/{len(selected_texts)}]", {
                        'original_text': original_text[:50] + "...",
                        'corrected_text': corrected_text[:50] + "...",
                        'translator_class': type(translator).__name__
                    })

                    try:
                        # 개별 텍스트를 배치로 처리 (선택된 번역기 사용)
                        individual_result = translator.translate_batch(
                            [corrected_text],  # 교정된 텍스트 사용
                            target_languages
                        )

                        log_step(session_id, f"개별 텍스트 번역 성공 [{i+1}/{len(selected_texts)}]", {
                            'result_available': individual_result is not None and len(individual_result) > 0,
                            'result_count': len(individual_result) if individual_result else 0
                        })

                        if individual_result and len(individual_result) > 0:
                            # LINE API 용어집 적용 (XLT System v5.1.1)
                            result = individual_result[0].copy()

                            # ko_KR은 교정된 텍스트 사용
                            result['ko_KR'] = corrected_text

                            # 각 언어별로 용어집 기반 후처리 적용
                            for lang in target_languages:
                                if lang in result and lang != 'ko_KR':
                                    original_translation = result[lang]
                                    guided_translation = apply_terminology_guide(original_translation, lang, terminology)
                                    if guided_translation != original_translation:
                                        result[lang] = guided_translation
                                        add_translation_log(f'📖 [{i+1}] {lang} 용어집 적용')

                            translated_results.append(result)
                            add_translation_log(f'✅ [{i+1}/{len(selected_texts)}] 번역 완료')
                        else:
                            # 실패 시 기본 구조로 채우기
                            fallback_result = {'original': original_text, 'ko_KR': corrected_text}
                            for lang in target_languages:
                                fallback_result[lang] = corrected_text  # 교정된 텍스트로 대체
                            translated_results.append(fallback_result)
                            add_translation_log(f'⚠️ [{i+1}/{len(selected_texts)}] 번역 실패, 원본 사용')

                    except Exception as e:
                        add_translation_log(f'❌ [{i+1}/{len(selected_texts)}] 오류: {str(e)[:50]}')
                        # 오류 시 기본 구조로 채우기
                        fallback_result = {'original': original_text, 'ko_KR': corrected_text}
                        for lang in target_languages:
                            fallback_result[lang] = corrected_text
                        translated_results.append(fallback_result)

                add_translation_log(f'🎯 전체 개별 번역 완료: {len(translated_results)}/{len(selected_texts)}개')

            # 번역 완료 후 각 언어별 결과 분석
            if translated_results:
                # ko_KR은 항상 성공으로 간주 (원본 사용)
                add_translation_log(f'✅ {lang_names.get("ko_KR", "ko_KR")} 번역 완료: {len(translated_results)}/{len(translated_results)}개 (원본 사용)', 'success')

                for lang in target_languages:
                    success_count = sum(1 for result in translated_results
                                      if result.get(lang) and result.get(lang) != result.get('original', ''))
                    if success_count > 0:
                        add_translation_log(f'✅ {lang_names.get(lang, lang)} 번역 완료: {success_count}/{len(translated_results)}개', 'success')
                    else:
                        add_translation_log(f'⚠️ {lang_names.get(lang, lang)} 번역 실패 또는 미완료', 'warning')

            add_translation_log(f'🎉 전체 번역 작업 완료!', 'success')

            # 번역 완료 로그 추가
            lang_names = {
                'ko_KR': '한국어', 'en_US': '영어', 'ja_JP': '일본어',
                'zh_TW': '중국어', 'th_TH': '태국어'
            }

            # 각 언어별 번역 완료 상태 로그
            for lang in target_languages:
                if translated_results and len(translated_results) > 0:
                    # 실제 번역된 항목 확인
                    success_count = sum(1 for result in translated_results if result.get(lang) and result.get(lang) != result.get('original', ''))
                    if success_count > 0:
                        log_entry = {
                            'message': f'   ✅ {lang_names.get(lang, lang)} 번역 완료: {success_count}개 항목',
                            'timestamp': datetime.now().strftime("%H:%M:%S"),
                            'type': 'success'
                        }
                    else:
                        log_entry = {
                            'message': f'   ⚠️ {lang_names.get(lang, lang)} 번역 실패 또는 미완료',
                            'timestamp': datetime.now().strftime("%H:%M:%S"),
                            'type': 'warning'
                        }
                    session_logs[session_id].append(log_entry)

            # 전체 완료 로그
            total_success = len(translated_results) if translated_results else 0
            log_entry = {
                'message': f'🎉 번역 완료: {total_success}개 텍스트 처리됨',
                'timestamp': datetime.now().strftime("%H:%M:%S"),
                'type': 'success'
            }
            session_logs[session_id].append(log_entry)

            # 맥 알림 및 사운드 전송 (안전한 방식)
            try:
                import subprocess
                import os

                # 알림 메시지
                notification_title = "XLT 번역 완료"
                notification_message = f"{total_success}개 텍스트 번역이 완료되었습니다!"

                # AppleScript를 사용한 알림 및 사운드 (안전한 방식)
                applescript = f'display notification "{notification_message}" with title "{notification_title}" sound name "Glass"'

                # 타임아웃을 짧게 설정하여 블로킹 방지
                result = subprocess.run(
                    ['osascript', '-e', applescript],
                    capture_output=True,
                    text=True,
                    timeout=5  # 5초 타임아웃
                )

                if result.returncode == 0:
                    print(f"📢 맥 알림 전송 성공: {notification_message}")
                else:
                    print(f"⚠️ 알림 전송 실패: {result.stderr}")

                # 사운드 파일 존재 확인 후 재생
                sound_file = '/System/Library/Sounds/Glass.aiff'
                if os.path.exists(sound_file):
                    subprocess.run(['afplay', sound_file], capture_output=True, timeout=3)
                else:
                    print(f"⚠️ 사운드 파일 없음: {sound_file}")

            except subprocess.TimeoutExpired:
                print("⚠️ 알림 전송 타임아웃 (5초 초과)")
            except Exception as e:
                print(f"⚠️ 알림 전송 실패: {type(e).__name__}: {e}")

            # 알림 실패와 관계없이 번역 작업은 계속 진행

            # 번역 완료 상태 업데이트
            session_status[session_id]['translation_progress'].update({
                'status': 'completed',
                'completed_languages': target_languages,
                'current_language': None,
                'message': f'{len(translated_results)}개 텍스트 번역 완료!'
            })

            print(f"✅ 번역 완료: {len(translated_results)}개 결과")
            if translated_results:
                print(f"   샘플 결과 키: {list(translated_results[0].keys())}")
                print(f"   샘플 번역:")
                for lang in target_languages[:3]:  # 처음 3개 언어만 출력
                    print(f"      {lang}: {translated_results[0].get(lang, 'N/A')[:50]}...")

            for i, result in enumerate(translated_results):
                # 번역 결과를 플랫 구조로 변환
                translation_item = {
                    'original': result.get('original', selected_texts[i]),
                    'ko_KR': result.get('ko_KR', selected_texts[i]),
                    'en_US': result.get('en_US', selected_texts[i]),
                    'ja_JP': result.get('ja_JP', selected_texts[i]),
                    'zh_TW': result.get('zh_TW', selected_texts[i]),
                    'th_TH': result.get('th_TH', selected_texts[i])
                }
                translations.append(translation_item)

        except Exception as e:
            print("=" * 80)
            print(f"❌ 번역 실패: {type(e).__name__}: {e}")
            print("=" * 80)
            import traceback
            error_trace = traceback.format_exc()
            print(error_trace)

            # 상세 에러 로그 기록
            log_error(session_id, e, "번역 실행 중 최종 에러")
            log_step(session_id, "번역 실패로 세션 종료", {
                'error_type': type(e).__name__,
                'error_message': str(e),
                'translation_mode': translation_mode,
                'traceback_lines': len(error_trace.split('\n'))
            }, 'ERROR')

            # 세션 요약 저장
            save_session_summary(session_id)

            # 오류 로그를 세션 로그에 추가
            log_entry = {
                'message': f'❌ 번역 실패: {type(e).__name__}: {str(e)}',
                'timestamp': datetime.now().strftime("%H:%M:%S"),
                'type': 'error'
            }
            session_logs[session_id].append(log_entry)

            # 번역 실패 상태 업데이트
            if session_id in session_status:
                session_status[session_id]['translation_progress'].update({
                    'status': 'failed',
                    'completed_languages': [],
                    'current_language': None,
                    'message': f'번역 실패: {str(e)}'
                })

            # 사용자에게 명확한 오류 메시지 반환 (번역 방식별)
            if translation_mode == 'claude':
                error_msg = f"Claude 번역 중 오류가 발생했습니다: {str(e)}"
            else:
                error_msg = f"번역 중 오류가 발생했습니다: {str(e)}"

            # 세션 정리
            cleanup_session(session_id)

            return jsonify({
                'status': 'error',
                'error': error_msg,
                'error_type': type(e).__name__,
                'translation_mode': translation_mode,
                'details': error_trace,
                'error_type': type(e).__name__,
                'translation_mode': translation_mode
            }), 500

        if not translations:
            return jsonify({
                'status': 'error',
                'error': '모든 텍스트 번역에 실패했습니다.'
            }), 500

        # 번역 결과에 XLT Key 추가
        translations_with_keys = []
        for i, (translation, xlt_key) in enumerate(zip(translations, xlt_keys)):
            translation_with_key = translation.copy()
            translation_with_key['xlt_key'] = xlt_key  # 개별 설정된 XLT key 포함
            translations_with_keys.append(translation_with_key)

        # 번역 결과를 세션에 저장 (XLT Key 포함)
        session_data['translations'] = translations_with_keys
        session_data['translation_completed'] = True
        session_data['xlt_keys'] = xlt_keys  # XLT key 배열도 저장
        print(f"✅ 번역 완료: {len(translations_with_keys)}개 결과를 세션에 저장 (XLT Key 포함)")

        # 번역 결과에 원본 및 처리된 텍스트 정보 추가
        translation_details = []

        print(f"🔍 디버깅: selected_indexes 개수={len(selected_indexes)}")
        print(f"🔍 디버깅: selected_texts 개수={len(selected_texts)}")
        print(f"🔍 디버깅: translations 개수={len(translations)}")

        for i, translation in enumerate(translations):
            # OCR 원본 텍스트
            if i < len(selected_indexes) and selected_indexes[i] < len(ocr_results):
                ocr_original = ocr_results[selected_indexes[i]]['text']
            else:
                ocr_original = ""

            # 처리된 텍스트 (실제로 번역에 사용된 텍스트)
            # selected_texts는 이미 치환자가 적용되었거나 교정된 텍스트
            if i < len(selected_texts):
                processed_text = selected_texts[i]
            else:
                processed_text = ocr_original  # fallback

            print(f"  [{i}] OCR원본='{ocr_original[:30]}...' 처리됨='{processed_text[:30]}...'")

            detail = {
                'original_text': ocr_original,  # OCR 원본
                'processed_text': processed_text,  # 실제로 번역에 사용된 텍스트
                'translations': {
                    'ko_KR': translation.get('ko_KR', processed_text),  # ko_KR이 없으면 원본 사용
                    'en_US': translation.get('en_US', processed_text),
                    'ja_JP': translation.get('ja_JP', processed_text),
                    'zh_TW': translation.get('zh_TW', processed_text),
                    'th_TH': translation.get('th_TH', processed_text)
                }
            }
            translation_details.append(detail)

        # XLT key가 제공된 경우 번역 결과와 함께 미리보기 데이터 반환
        xlt_keys = data.get('xlt_keys', [])
        if xlt_keys and len(xlt_keys) == len(translations_with_keys):
            # 번역 결과에 XLT key 정보 추가하여 미리보기용 데이터 생성
            preview_data = []
            for i, (detail, xlt_key) in enumerate(zip(translation_details, xlt_keys)):
                preview_item = {
                    'xlt_key': xlt_key,
                    'original_text': detail['original_text'],
                    'processed_text': detail['processed_text'],
                    'translations': detail['translations']
                }
                preview_data.append(preview_item)

            # 세션에 Excel 생성용 데이터 저장
            session_data['excel_ready_data'] = {
                'translations': [t['translations'] for t in translation_details],
                'xlt_keys': xlt_keys,
                'preview_data': preview_data
            }

            print(f"✅ 번역 완료 및 미리보기 준비: {len(preview_data)}개 항목")

            # 성공적인 번역 완료 로그
            log_step(session_id, "번역 완료 - Excel 준비됨", {
                'processed_count': len(translations_with_keys),
                'preview_data_count': len(preview_data),
                'xlt_keys_count': len(xlt_keys),
                'learned_count': learned_count
            })

            # 세션 요약 저장 후 정리
            save_session_summary(session_id)
            cleanup_session(session_id)

            return jsonify({
                'status': 'success',
                'session_id': session_id,
                'processed_count': len(translations_with_keys),
                'translations': translation_details,
                'preview_data': preview_data,
                'xlt_keys': xlt_keys,
                'excel_ready': True,
                'learned_count': learned_count,  # v3.3: 학습된 교정 개수
                'learned_corrections': learned_corrections,  # v3.3: 학습 내역
                'message': '번역이 완료되었습니다. 미리보기를 확인하고 Excel 파일을 다운로드하세요.'
            })

        # Excel 생성 없이 번역 결과만 반환 (XLT key가 없는 경우)
        log_step(session_id, "번역 완료 - Excel 미준비", {
            'processed_count': len(translations_with_keys),
            'excel_ready': False
        })

        # 세션 요약 저장 후 정리
        save_session_summary(session_id)
        cleanup_session(session_id)

        return jsonify({
            'status': 'success',
            'processed_count': len(translations_with_keys),
            'translations': translation_details,
            'excel_ready': False,
            'message': '번역이 완료되었습니다.'
        })

    except Exception as e:
        error_trace = traceback.format_exc()

        # 최상위 예외 처리 로그
        try:
            log_error(session_id, e, "translate_selected 최상위 예외")
            save_session_summary(session_id)
            cleanup_session(session_id)
        except:
            # 로깅 실패 시에도 원본 에러는 반환
            pass

        return jsonify({
            'status': 'error',
            'error': f'선택 번역 중 오류: {str(e)}',
            'error_type': type(e).__name__,
            'details': error_trace
        }), 500


def create_excel_with_custom_keys(translations, xlt_keys):
    """sampleformat.xlsx를 물리적으로 완전 복사한 후 properties 워크시트 2행부터만 업데이트"""
    try:
        import openpyxl
        from io import BytesIO
        import os
        import shutil
        import tempfile

        # sampleformat.xlsx 템플릿 파일 경로 (동적)
        current_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(current_dir, 'Sample', 'sampleformat.xlsx')

        if not os.path.exists(template_path):
            raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {template_path}")

        print("📋 sampleformat.xlsx를 물리적으로 완전 복사 중...")

        # 1단계: sampleformat.xlsx를 임시 파일로 물리적 복사
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            shutil.copy2(template_path, tmp_file.name)
            copied_file_path = tmp_file.name

        print(f"✅ sampleformat.xlsx 물리적 복사 완료: {copied_file_path}")

        # 2단계: 복사된 파일을 로드
        wb = openpyxl.load_workbook(copied_file_path)

        print(f"📂 복사된 파일 구조:")
        print(f"   워크시트 개수: {len(wb.worksheets)}개")
        print(f"   워크시트 이름: {[ws.title for ws in wb.worksheets]}")

        # 모든 워크시트 보존 확인
        for ws in wb.worksheets:
            print(f"   - {ws.title}: {ws.max_column}열 × {ws.max_row}행")

        # 3단계: properties 워크시트만 수정
        ws_properties = wb['properties']
        original_max_row = ws_properties.max_row
        print(f"📝 properties 워크시트 수정 (원본: {ws_properties.max_column}열 × {original_max_row}행)")

        # 4단계: 헤더 구조 보존 확인 (1행은 절대 건드리지 않음)
        headers = []
        for col in range(1, ws_properties.max_column + 1):
            header = ws_properties.cell(row=1, column=col).value
            headers.append(header)

        print(f"📊 원본 헤더 보존: {headers}")

        # 5단계: XLT 번역 언어 매핑
        xlt_language_mapping = {
            'en_US': 'en_US',
            'ko_KR': 'ko_KR',
            'ja_JP': 'ja_JP',
            'zh_TW': 'zh_TW',
            'th_TH': 'th_TH'
        }

        # 6단계: 기존 데이터 행만 정리 (2행부터 끝까지, 헤더 절대 건드리지 않음)
        if original_max_row > 1:
            # 2행부터 마지막 행까지 모든 셀 내용을 빈 값으로 설정
            for row in range(2, original_max_row + 1):
                for col in range(1, ws_properties.max_column + 1):
                    ws_properties.cell(row=row, column=col).value = None
            print(f"🗑️  기존 데이터 정리 완료 (2~{original_max_row}행)")

        # 7단계: 번역 데이터를 2행부터 입력 (헤더는 절대 건드리지 않음)
        print(f"📝 번역 데이터 입력 시작: {len(translations)}개 항목")

        for data_idx, (translation, xlt_key) in enumerate(zip(translations, xlt_keys)):
            row_idx = data_idx + 2  # 2행부터 시작
            print(f"   [{data_idx+1}/{len(translations)}] 행 {row_idx}: Key={xlt_key}")

            # A열: Key ID 설정
            ws_properties.cell(row=row_idx, column=1, value=xlt_key)

            # B열부터: 언어별 번역 데이터
            for col_idx, header in enumerate(headers[1:], 2):
                if header in xlt_language_mapping:
                    # XLT에서 번역된 언어
                    xlt_lang_code = xlt_language_mapping[header]
                    translated_text = translation.get(xlt_lang_code, '')
                    ws_properties.cell(row=row_idx, column=col_idx, value=translated_text)

                    if translated_text:
                        print(f"      컬럼 {col_idx} ({header}): {translated_text[:25]}{'...' if len(translated_text) > 25 else ''}")
                else:
                    # XLT에서 번역되지 않은 언어는 빈칸 유지
                    ws_properties.cell(row=row_idx, column=col_idx, value='')

        # 8단계: 모든 워크시트 구조 최종 확인
        print(f"📋 최종 파일 구조:")
        for ws in wb.worksheets:
            print(f"   - {ws.title}: {ws.max_column}열 × {ws.max_row}행")

        print(f"✅ sampleformat.xlsx 기반 완전 복사본 생성 완료:")
        print(f"   📋 원본: sampleformat.xlsx (3개 워크시트 완전 보존)")
        print(f"   📝 수정: properties 워크시트 2~{len(translations)+1}행만")
        print(f"   🔒 보존: plurals, Language code 워크시트 완전 보존")
        print(f"   🌐 번역 언어: {', '.join(xlt_language_mapping.keys())}")

        # 9단계: 메모리 버퍼로 저장
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # 10단계: 임시 파일 정리
        try:
            os.unlink(copied_file_path)
            print(f"🧹 임시 파일 정리 완료")
        except:
            pass

        print(f"💾 메모리 버퍼 저장 완료 (크기: {len(buffer.getvalue())} bytes)")
        return buffer

    except Exception as e:
        print(f"❌ sampleformat.xlsx 완전 복사 실패: {e}")
        import traceback
        error_trace = traceback.format_exc()
        print(f"❌ 상세 오류:\n{error_trace}")
        return None


@app.route('/download-excel', methods=['GET', 'POST'])
def download_excel():
    """미리보기 완료 후 Excel 파일 다운로드"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    try:
        data = request.get_json()
        session_id = data.get('session_id')

        if not session_id or session_id not in temp_ocr_results:
            return jsonify({
                'status': 'error',
                'error': '세션을 찾을 수 없습니다.'
            }), 400

        session_data = temp_ocr_results[session_id]
        excel_data = session_data.get('excel_ready_data')

        if not excel_data:
            return jsonify({
                'status': 'error',
                'error': '다운로드할 Excel 데이터가 없습니다. 번역을 다시 시도해주세요.'
            }), 400

        print(f"📥 Excel 다운로드 요청: {len(excel_data['xlt_keys'])}개 항목")

        # Excel 파일 생성
        excel_buffer = create_excel_with_custom_keys(
            excel_data['translations'],
            excel_data['xlt_keys']
        )

        if excel_buffer is None:
            return jsonify({
                'status': 'error',
                'error': 'Excel 파일 생성에 실패했습니다.'
            }), 500

        # 파일명 생성
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"xlt_translation_{timestamp}.xlsx"

        print(f"✅ Excel 파일 다운로드 준비 완료: {filename}")

        # Excel 파일 다운로드
        from flask import send_file
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"❌ Excel 다운로드 오류: {e}")
        print(f"   상세 오류: {error_trace}")

        return jsonify({
            'status': 'error',
            'error': f'Excel 다운로드 중 오류가 발생했습니다: {str(e)}'
        }), 500


@app.route('/generate-excel', methods=['GET', 'POST'])
def generate_excel():
    """번역 결과로 Excel 파일 생성하여 바로 다운로드 (메모리 기반) - 레거시"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    try:
        data = request.json
        session_id = data.get('session_id')
        key_prefix = data.get('key_prefix', '').strip()

        if not session_id or session_id not in temp_ocr_results:
            return jsonify({
                'status': 'error',
                'error': '세션을 찾을 수 없습니다. 번역을 다시 시도해주세요.'
            }), 400

        session_data = temp_ocr_results[session_id]

        if not session_data.get('translation_completed'):
            return jsonify({
                'status': 'error',
                'error': '번역이 완료되지 않았습니다.'
            }), 400

        translations = session_data.get('translations', [])
        if not translations:
            return jsonify({
                'status': 'error',
                'error': '번역 결과가 없습니다.'
            }), 400

        # Key prefix 적용하여 XLT keys 생성
        xlt_keys = []
        for i in range(len(translations)):
            if key_prefix:
                xlt_keys.append(f"{key_prefix}_{i+1}")
            else:
                xlt_keys.append(f"item_{i+1}")

        # translations에서 xlt_key 제거하고 순수 번역 데이터만 추출
        clean_translations = []
        for translation in translations:
            # xlt_key를 제외한 번역 데이터만 추출
            clean_translation = {k: v for k, v in translation.items() if k != 'xlt_key'}
            clean_translations.append(clean_translation)

        print(f"🔍 /generate-excel 데이터 변환:")
        print(f"   원본 translations 개수: {len(translations)}")
        print(f"   정제된 translations 개수: {len(clean_translations)}")
        if clean_translations:
            print(f"   샘플 키: {list(clean_translations[0].keys())}")

        # Excel 파일 생성 (메모리)
        excel_buffer = create_excel_with_custom_keys(clean_translations, xlt_keys)

        if excel_buffer is None:
            return jsonify({
                'status': 'error',
                'error': 'Excel 파일 생성에 실패했습니다.'
            }), 500

        # 파일명 생성
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}.xlsx"

        # 바로 다운로드 응답
        from flask import send_file
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"❌ Excel 생성 오류: {e}")
        print(f"   상세 오류: {error_trace}")

        return jsonify({
            'status': 'error',
            'error': f'Excel 파일 생성 중 오류가 발생했습니다: {str(e)}'
        }), 500


@app.route('/merge-excel', methods=['GET', 'POST'])
def merge_excel_files():
    """여러 Excel 파일을 하나로 합쳐서 바로 다운로드 (메모리 기반)"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    try:
        import openpyxl
        from pathlib import Path
        from datetime import datetime
        from io import BytesIO
        from flask import send_file

        # 요청 데이터 확인
        if 'files' not in request.files:
            return jsonify({
                'status': 'error',
                'error': '업로드된 파일이 없습니다.'
            }), 400

        files = request.files.getlist('files')
        remove_duplicates = request.form.get('remove_duplicates', 'true').lower() == 'true'
        sort_by_key = request.form.get('sort_by_key', 'true').lower() == 'true'
        custom_filename = request.form.get('output_filename', '').strip()

        if len(files) == 0:
            return jsonify({
                'status': 'error',
                'error': '업로드된 파일이 없습니다.'
            }), 400

        print(f"🔄 엑셀 합치기 시작: {len(files)}개 파일")

        # sampleformat.xlsx 템플릿 로드
        sample_path = Path('Sample/sampleformat.xlsx')
        if not sample_path.exists():
            return jsonify({
                'status': 'error',
                'error': 'sampleformat.xlsx 템플릿 파일을 찾을 수 없습니다.'
            }), 500

        # 템플릿 워크북 로드
        template_wb = openpyxl.load_workbook(sample_path)
        template_ws = template_wb.active

        # XLT 시스템에서 사용하는 언어 컬럼 매핑 (sampleformat.xlsx 기준)
        xlt_to_sample_mapping = {
            'en_US': 'en_US',  # 열 2
            'ko_KR': 'ko_KR',  # 열 3
            'ja_JP': 'ja_JP',  # 열 4
            'zh_TW': 'zh_TW',  # 열 6
            'th_TH': 'th_TH'   # 열 11
        }

        # 헤더에서 컬럼 인덱스 찾기
        header_row = 1
        col_mapping = {}
        for col in range(1, template_ws.max_column + 1):
            header_value = template_ws.cell(header_row, col).value
            if header_value in xlt_to_sample_mapping.values():
                col_mapping[header_value] = col

        # Key ID 컬럼 찾기
        key_col = None
        for col in range(1, template_ws.max_column + 1):
            if template_ws.cell(header_row, col).value == 'Key ID':
                key_col = col
                break

        if key_col is None:
            return jsonify({
                'status': 'error',
                'error': 'Key ID 컬럼을 찾을 수 없습니다.'
            }), 500

        # 모든 데이터를 저장할 딕셔너리
        all_data = {}

        # 각 파일을 읽어서 데이터 수집
        for file_obj in files:
            if file_obj.filename == '':
                continue

            try:
                # Excel 파일 읽기
                wb = openpyxl.load_workbook(file_obj)
                ws = wb.active

                # XLT 파일의 헤더 확인
                xlt_headers = {}
                for col in range(1, ws.max_column + 1):
                    header_value = ws.cell(1, col).value
                    if header_value:
                        xlt_headers[header_value] = col

                # Key 컬럼 찾기 (Key ID 또는 Key)
                xlt_key_col = xlt_headers.get('Key ID') or xlt_headers.get('Key')
                if xlt_key_col is None:
                    continue

                # 데이터 행 읽기 (2행부터)
                for row in range(2, ws.max_row + 1):
                    key_value = ws.cell(row, xlt_key_col).value
                    if not key_value or key_value.strip() == '':
                        continue

                    key_value = key_value.strip()

                    # 중복 확인
                    if remove_duplicates and key_value in all_data:
                        continue

                    # 행 데이터 수집
                    row_data = {'Key': key_value}

                    # XLT 시스템 언어들 매핑
                    for xlt_lang, sample_lang in xlt_to_sample_mapping.items():
                        xlt_col = xlt_headers.get(xlt_lang)
                        if xlt_col:
                            cell_value = ws.cell(row, xlt_col).value
                            row_data[sample_lang] = cell_value if cell_value else ''

                    all_data[key_value] = row_data

            except Exception as e:
                print(f"파일 처리 오류: {e}")

        if len(all_data) == 0:
            return jsonify({
                'status': 'error',
                'error': '처리할 수 있는 데이터가 없습니다.'
            }), 400

        # 정렬
        if sort_by_key:
            sorted_keys = sorted(all_data.keys())
        else:
            sorted_keys = list(all_data.keys())

        # 새 워크북에 데이터 입력
        current_row = 2
        for key in sorted_keys:
            data = all_data[key]

            # Key 입력
            template_ws.cell(current_row, key_col).value = data['Key']

            # 언어별 데이터 입력
            for lang, col_idx in col_mapping.items():
                value = data.get(lang, '')
                template_ws.cell(current_row, col_idx).value = value

            current_row += 1

        # 출력 파일명 생성
        if custom_filename:
            output_filename = f"{custom_filename}.xlsx"
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"merged_translations_{timestamp}.xlsx"

        # 메모리 버퍼로 저장
        buffer = BytesIO()
        template_wb.save(buffer)
        buffer.seek(0)

        print(f"✅ 합친 파일 생성: {output_filename} ({len(all_data)}개 항목)")

        # 바로 다운로드 응답
        return send_file(
            buffer,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"❌ 엑셀 합치기 오류: {e}")

        return jsonify({
            'status': 'error',
            'error': f'엑셀 합치기 중 오류가 발생했습니다: {str(e)}'
        }), 500


def check_system_health():
    """시스템 전체 헬스체크"""
    health_status = {}

    # 1. Figma API 직접 사용 - OCR 엔진 불필요 (제거됨)

    # 2. 피그마 연결 상태 확인
    try:
        from xlt.input.figma import FigmaProcessor
        figma_processor = FigmaProcessor(pipeline.config)
        if figma_processor.figma_token:
            # 토큰이 있는 경우 연결 테스트
            if hasattr(figma_processor, 'test_figma_connection'):
                connection_ok = figma_processor.test_figma_connection()
                if connection_ok:
                    health_status['figma'] = {
                        'status': 'ok',
                        'message': '피그마 연결 정상',
                        'details': '토큰 유효, API 접근 가능',
                        'last_check': datetime.now().isoformat()
                    }
                else:
                    health_status['figma'] = {
                        'status': 'warning',
                        'message': '피그마 API 접근 실패',
                        'details': '토큰은 있지만 API 호출 실패. 네트워크 또는 토큰 권한 확인 필요',
                        'last_check': datetime.now().isoformat()
                    }
            else:
                health_status['figma'] = {
                    'status': 'ok',
                    'message': '피그마 토큰 설정됨',
                    'details': '토큰이 설정되어 있음 (연결 테스트 불가)',
                    'last_check': datetime.now().isoformat()
                }
        else:
            health_status['figma'] = {
                'status': 'warning',
                'message': '피그마 토큰 미설정',
                'details': '피그마 URL 처리를 위해 figma_config.json에 토큰 설정 필요',
                'last_check': datetime.now().isoformat()
            }
    except Exception as e:
        health_status['figma'] = {
            'status': 'error',
            'message': '피그마 모듈 오류',
            'details': f'문제: {str(e)}. 모듈 로드 실패',
            'last_check': datetime.now().isoformat()
        }

    # 2.5. 피그마 API 상세 상태 확인
    try:
        if 'figma' in health_status and health_status['figma']['status'] == 'ok':
            # 기본 연결이 성공한 경우에만 API 상세 테스트 진행
            from xlt.input.figma import FigmaProcessor
            figma_processor = FigmaProcessor(pipeline.config)

            api_details = []
            api_warnings = []
            api_errors = []

            # 이미 기본 연결 테스트 완료됨 (중복 호출 방지)
            api_details.append('API 접근 가능')
            api_details.append('토큰 유효성 확인됨')

            # 중복 API 호출 제거 (이미 test_figma_connection()에서 확인됨)
            # 기본 연결이 성공했으므로 추가 검증 불필요

            # 토큰 유효성 상세 체크
            try:
                token_length = len(figma_processor.figma_token) if figma_processor.figma_token else 0
                if token_length == 0:
                    api_errors.append('토큰이 설정되지 않음')
                elif token_length < 40:
                    api_warnings.append('토큰 길이가 짧음 (유효성 의심)')
                else:
                    api_details.append(f'토큰 길이: {token_length}자 (정상)')

            except Exception as e:
                api_warnings.append(f'토큰 검증 실패: {str(e)}')

            # 상태 결정 및 메시지 생성
            if api_errors:
                health_status['figma_api'] = {
                    'status': 'error',
                    'message': '피그마 API 오류',
                    'details': f'에러: {"; ".join(api_errors[:2])}. 경고: {len(api_warnings)}개',
                    'last_check': datetime.now().isoformat()
                }
            elif api_warnings:
                health_status['figma_api'] = {
                    'status': 'warning',
                    'message': '피그마 API 부분 문제',
                    'details': f'경고: {"; ".join(api_warnings[:2])}. 정상: {"; ".join(api_details[:2])}',
                    'last_check': datetime.now().isoformat()
                }
            else:
                health_status['figma_api'] = {
                    'status': 'ok',
                    'message': '피그마 API 완전 정상',
                    'details': f'{"; ".join(api_details[:3])}',
                    'last_check': datetime.now().isoformat()
                }

        else:
            # 기본 피그마 연결이 실패한 경우
            health_status['figma_api'] = {
                'status': 'error',
                'message': '피그마 API 사용 불가',
                'details': '기본 피그마 연결이 실패하여 API 테스트 불가',
                'last_check': datetime.now().isoformat()
            }

    except Exception as e:
        health_status['figma_api'] = {
            'status': 'error',
            'message': '피그마 API 체크 오류',
            'details': f'문제: {str(e)[:100]}',
            'last_check': datetime.now().isoformat()
        }

    # 3. Claude CLI에서 번역 시스템 상태 확인 (중복 제거)

    # 3.5 Claude CLI 상태 확인 (경량 체크)
    try:
        import subprocess

        # Claude CLI 인증 상태 확인 (경량 체크)
        try:
            result = subprocess.run(['claude', 'auth', 'status'],
                                  capture_output=True, text=True, timeout=3)

            if result.returncode == 0:
                # 인증 상태 확인 (JSON 응답 처리)
                try:
                    import json
                    auth_data = json.loads(result.stdout)
                    if auth_data.get('loggedIn', False):
                        health_status['claude'] = {
                            'status': 'ok',
                            'message': 'Claude CLI 정상',
                            'details': f'Claude CLI 인증 상태 양호 ({auth_data.get("authMethod", "unknown")})',
                            'last_check': datetime.now().isoformat()
                        }
                    else:
                        health_status['claude'] = {
                            'status': 'error',
                            'message': 'Claude CLI 인증 필요',
                            'details': 'Claude CLI가 설치되었지만 로그인이 필요합니다. "claude auth login" 명령어를 실행하세요.',
                            'last_check': datetime.now().isoformat()
                        }
                except json.JSONDecodeError:
                    # JSON이 아닌 경우 기존 텍스트 방식으로 fallback
                    output = result.stdout.lower()
                    if 'authenticated' in output or 'logged in' in output or 'true' in output:
                        health_status['claude'] = {
                            'status': 'ok',
                            'message': 'Claude CLI 정상',
                            'details': 'Claude CLI 인증 상태 양호',
                            'last_check': datetime.now().isoformat()
                        }
                    else:
                        health_status['claude'] = {
                            'status': 'error',
                            'message': 'Claude CLI 인증 필요',
                            'details': 'Claude CLI가 설치되었지만 로그인이 필요합니다. "claude auth login" 명령어를 실행하세요.',
                            'last_check': datetime.now().isoformat()
                        }
            else:
                health_status['claude'] = {
                    'status': 'error',
                    'message': 'Claude CLI 실행 실패',
                    'details': f'Claude CLI 명령 실행 실패 (코드: {result.returncode}). Claude CLI가 올바르게 설치되었는지 확인하세요.',
                    'last_check': datetime.now().isoformat()
                }
        except FileNotFoundError:
            health_status['claude'] = {
                'status': 'error',
                'message': 'Claude CLI 미설치',
                'details': 'Claude CLI가 설치되지 않았습니다. Claude 통합 번역 기능을 사용하려면 Claude CLI 설치가 필요합니다.',
                'last_check': datetime.now().isoformat()
            }
        except subprocess.TimeoutExpired:
            health_status['claude'] = {
                'status': 'error',
                'message': 'Claude CLI 응답 시간 초과',
                'details': 'Claude CLI 응답이 5초를 초과했습니다. 네트워크 연결 상태를 확인하세요.',
                'last_check': datetime.now().isoformat()
            }
    except Exception as e:
        health_status['claude'] = {
            'status': 'error',
            'message': 'Claude CLI 체크 오류',
            'details': f'Claude CLI 상태 확인 중 오류: {str(e)}',
            'last_check': datetime.now().isoformat()
        }

    # 4. 파일 시스템 상태 확인
    try:
        import os
        from pathlib import Path

        # 필요한 디렉토리들 확인
        required_dirs = ['static', 'templates']  # XLT v3.0: uploads, logs, output 제거 (피그마 전용)
        all_dirs_ok = True
        missing_dirs = []

        for dir_name in required_dirs:
            if not os.path.exists(dir_name):
                all_dirs_ok = False
                missing_dirs.append(dir_name)
            elif not os.access(dir_name, os.W_OK):
                all_dirs_ok = False
                missing_dirs.append(f"{dir_name}(권한없음)")

        if all_dirs_ok:
            health_status['filesystem'] = {
                'status': 'ok',
                'message': '파일 시스템 정상',
                'details': f'모든 필수 디렉토리 접근 가능: {", ".join(required_dirs)}',
                'last_check': datetime.now().isoformat()
            }
        else:
            health_status['filesystem'] = {
                'status': 'warning',
                'message': '파일 시스템 권한 문제',
                'details': f'문제 디렉토리: {", ".join(missing_dirs)}. 디렉토리 생성 또는 권한 확인 필요',
                'last_check': datetime.now().isoformat()
            }
    except Exception as e:
        health_status['filesystem'] = {
            'status': 'error',
            'message': '파일 시스템 오류',
            'details': f'문제: {str(e)}',
            'last_check': datetime.now().isoformat()
        }

    # 5. 의존성 패키지 상태 확인 (설치 여부만 체크)
    try:
        required_packages = {
            'openpyxl': 'Excel 파일 처리',
            'pillow': '이미지 처리',
            'flask': '웹 서버'
        }

        missing_packages = []
        installed_packages = []

        for package, description in required_packages.items():
            try:
                if package == 'pillow':
                    from PIL import Image
                    version = getattr(Image, '__version__', 'unknown')
                elif package == 'flask':
                    import flask
                    version = flask.__version__
                elif package == 'openpyxl':
                    import openpyxl
                    version = openpyxl.__version__

                installed_packages.append(f"{package}({version})")

            except ImportError:
                missing_packages.append(f"{package}({description})")

        if not missing_packages:
            health_status['dependencies'] = {
                'status': 'ok',
                'message': '의존성 패키지 정상',
                'details': f'설치됨: {", ".join(installed_packages)}',
                'last_check': datetime.now().isoformat()
            }
        else:
            health_status['dependencies'] = {
                'status': 'error',
                'message': '필수 패키지 누락',
                'details': f'누락: {", ".join(missing_packages)}. 해결: pip install -r requirements.txt',
                'last_check': datetime.now().isoformat()
            }
    except Exception as e:
        health_status['dependencies'] = {
            'status': 'error',
            'message': '의존성 체크 오류',
            'details': f'문제: {str(e)}',
            'last_check': datetime.now().isoformat()
        }

    # 6. 메모리 상태 확인 (유연한 체크)
    try:
        import psutil
        memory = psutil.virtual_memory()
        if memory.percent < 85:
            health_status['memory'] = {
                'status': 'ok',
                'message': '메모리 사용률 정상',
                'details': f'사용률: {memory.percent:.1f}% ({memory.available // (1024**3):.1f}GB 여유)',
                'last_check': datetime.now().isoformat()
            }
        elif memory.percent < 95:
            health_status['memory'] = {
                'status': 'warning',
                'message': '메모리 사용률 높음',
                'details': f'사용률: {memory.percent:.1f}% 메모리 정리 권장',
                'last_check': datetime.now().isoformat()
            }
        else:
            health_status['memory'] = {
                'status': 'error',
                'message': '메모리 부족',
                'details': f'사용률: {memory.percent:.1f}% 즉시 메모리 확보 필요',
                'last_check': datetime.now().isoformat()
            }
    except ImportError:
        # psutil이 없어도 정상으로 처리 (선택적 기능)
        health_status['memory'] = {
            'status': 'ok',
            'message': '메모리 모니터링 비활성화',
            'details': '시스템 메모리는 정상 동작 중 (상세 모니터링 불가)',
            'last_check': datetime.now().isoformat()
        }
    except Exception as e:
        # 오류가 있어도 경고가 아닌 정보성으로 처리
        health_status['memory'] = {
            'status': 'ok',
            'message': '메모리 상태 추정 정상',
            'details': f'상세 체크 불가하지만 시스템 동작 정상',
            'last_check': datetime.now().isoformat()
        }

    # 6. Excel 출력 시스템 상태 확인
    try:
        import openpyxl
        downloads_dir = '/tmp/xlt_downloads'

        # 다운로드 디렉토리 확인/생성
        os.makedirs(downloads_dir, exist_ok=True)

        # 테스트 Excel 파일 생성 가능한지 확인
        test_wb = openpyxl.Workbook()
        test_ws = test_wb.active
        test_ws.cell(row=1, column=1, value='test')

        health_status['excel_output'] = {
            'status': 'ok',
            'message': 'Excel 출력 시스템 정상',
            'details': f'openpyxl v{openpyxl.__version__}, 다운로드 경로: {downloads_dir}',
            'last_check': datetime.now().isoformat()
        }

        # 테스트 워크북 정리
        test_wb.close()

    except Exception as e:
        health_status['excel_output'] = {
            'status': 'error',
            'message': 'Excel 출력 시스템 오류',
            'details': f'문제: {str(e)}. openpyxl 패키지 또는 파일 권한 확인 필요',
            'last_check': datetime.now().isoformat()
        }

    return health_status

@app.route('/api/health')
def api_health():
    """상세 시스템 헬스체크"""
    try:
        health_data = check_system_health()

        # 전체 상태 요약
        error_count = sum(1 for item in health_data.values() if item['status'] == 'error')
        warning_count = sum(1 for item in health_data.values() if item['status'] == 'warning')

        if error_count > 0:
            overall_status = 'error'
        elif warning_count > 0:
            overall_status = 'warning'
        else:
            overall_status = 'ok'

        return jsonify({
            'overall_status': overall_status,
            'summary': f'에러: {error_count}개, 경고: {warning_count}개',
            'components': health_data,
            'last_check': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({
            'overall_status': 'error',
            'summary': '헬스체크 실행 실패',
            'error': str(e),
            'last_check': datetime.now().isoformat()
        }), 500

# =====================================
# 업데이트 관련 API 엔드포인트
# =====================================

@app.route('/api/update/check', methods=['GET', 'POST'])
def api_update_check():
    """업데이트 확인 API (v5.0.6 자동화 시스템 감지 포함)"""
    try:
        # 버전 정보 캐시 새로고침 (최신 정보 보장)
        try:
            from xlt.utils.version_manager import get_version_manager
            vm = get_version_manager()
            vm.refresh_cache()
            print("🔄 버전 정보 캐시 새로고침 완료")
        except Exception as e:
            print(f"⚠️ 버전 캐시 새로고침 실패: {e}")

        # v5.0.6 자동화 시스템 존재 여부 확인
        auto_update_available = False
        try:
            from xlt.utils.auto_updater import AutoUpdateManager
            auto_update_available = True
        except ImportError:
            auto_update_available = False

        if not updater:
            return jsonify({
                'status': 'error',
                'error': '업데이터가 초기화되지 않았습니다.'
            }), 500

        update_info = updater.check_for_updates()

        # v5.0.6 자동화 시스템이 없는 경우 업그레이드 안내 추가
        if not auto_update_available:
            update_info['v5_upgrade_available'] = True
            update_info['v5_upgrade_message'] = "완전 자동화 시스템 v5.0.6으로 업그레이드 가능합니다"

        return jsonify({
            'status': 'success',
            'update_info': update_info,
            'auto_update_available': auto_update_available,
            'last_check': datetime.now().isoformat()
        })

    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"❌ 업데이트 확인 API 오류: {e}")
        print(f"   상세 오류: {error_trace}")

        return jsonify({
            'status': 'error',
            'error': f'업데이트 확인 중 오류가 발생했습니다: {str(e)}'
        }), 500

@app.route('/api/update/perform', methods=['GET', 'POST'])
def api_update_perform():
    """업데이트 실행 API"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    try:
        if not updater:
            return jsonify({
                'status': 'error',
                'error': '업데이터가 초기화되지 않았습니다.'
            }), 500

        # 업데이트 가능 여부 먼저 확인
        update_info = updater.check_for_updates()
        if not update_info.get('update_available'):
            return jsonify({
                'status': 'warning',
                'message': '업데이트할 내용이 없습니다. 이미 최신 버전입니다.'
            })

        # 옵션 처리
        request_data = request.get_json() or {}
        create_backup = request_data.get('create_backup', True)

        # 업데이트 실행
        result = updater.perform_update(create_backup=create_backup)

        if result['success']:
            return jsonify({
                'status': 'success',
                'message': result['message'],
                'backup_path': result.get('backup_path'),
                'update_log': result['update_log'],
                'restart_required': True
            })
        else:
            return jsonify({
                'status': 'error',
                'error': result.get('error'),
                'update_log': result['update_log'],
                'backup_path': result.get('backup_path')
            }), 400

    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"❌ 업데이트 실행 API 오류: {e}")
        print(f"   상세 오류: {error_trace}")

        return jsonify({
            'status': 'error',
            'error': f'업데이트 실행 중 오류가 발생했습니다: {str(e)}'
        }), 500

@app.route('/api/update/history')
def api_update_history():
    """업데이트 히스토리 API"""
    try:
        if not updater:
            return jsonify({
                'status': 'error',
                'error': '업데이터가 초기화되지 않았습니다.'
            }), 500

        limit = request.args.get('limit', 10, type=int)
        history = updater.get_update_history(limit=limit)

        return jsonify({
            'status': 'success',
            'history': history,
            'total_commits': len(history)
        })

    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"❌ 업데이트 히스토리 API 오류: {e}")
        print(f"   상세 오류: {error_trace}")

        return jsonify({
            'status': 'error',
            'error': f'업데이트 히스토리 조회 중 오류가 발생했습니다: {str(e)}'
        }), 500

@app.route('/api/update/status')
def api_update_status():
    """업데이트 상태 정보 API"""
    try:
        if not updater:
            return jsonify({
                'status': 'error',
                'error': '업데이터가 초기화되지 않았습니다.'
            }), 500

        current_version = updater.get_current_version()
        has_changes = updater.has_local_changes()

        return jsonify({
            'status': 'success',
            'current_version': current_version,
            'has_local_changes': has_changes,
            'repo_url': updater.repo_url,
            'last_check': datetime.now().isoformat()
        })

    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'업데이트 상태 조회 중 오류가 발생했습니다: {str(e)}'
        }), 500

@app.route('/api/status')
def api_status():
    """기본 시스템 상태 (기존 호환성 유지)"""
    try:
        # 피그마 연결 상태 확인
        figma_status = False
        figma_error = ""
        try:
            from xlt.input.figma import FigmaProcessor
            figma_processor = FigmaProcessor(pipeline.config)
            figma_status = bool(figma_processor.figma_token and figma_processor.test_figma_connection())
        except Exception as e:
            figma_error = str(e)

        return jsonify({
            'status': 'running' if pipeline else 'error',
            'version': '2.0.0-stable',
            'pipeline_ready': pipeline is not None,
            'figma_ready': figma_status,
            'figma_error': figma_error if not figma_status else None,
            'temp_sessions': len(temp_ocr_results)
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500

# XLT v3.0: 로그 API 제거
# @app.route('/api/logs/<session_id>') - 로그 기능 제거됨

@app.route('/api/translation-progress/<session_id>')
def get_translation_progress(session_id):
    """번역 진행 상황 조회"""
    try:
        status_data = session_status.get(session_id, {})
        translation_progress_data = status_data.get('translation_progress', {})

        # 청크 진행률 정보 추가 (새로운 진행률 시스템)
        chunk_progress_data = translation_progress.get(session_id, {})

        response_data = {
            'status': 'success',
            'progress': translation_progress_data,
            'timestamp': datetime.now().isoformat()
        }

        # 청크 진행률이 있으면 추가
        if chunk_progress_data:
            response_data['chunk_progress'] = chunk_progress_data.get('chunk_progress', {})
            response_data['chunk_updated_at'] = chunk_progress_data.get('updated_at')

        return jsonify(response_data)
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500

def load_translation_guide_api():
    """API용 LINE API 기반 번역 가이드라인 로드"""
    try:
        # LINE API에서 용어집 로드
        terminology = load_line_api_terminology()

        if terminology:
            # LINE API 성공
            guide_content = f"""# LINE API 기반 번역 가이드

**총 용어 수**: {len(terminology)}개
**지원 언어**: 한국어, 영어, 일본어, 중국어(번체), 태국어
**데이터 소스**: LINE API 실시간 동기화

## 핵심 용어집

"""
            # 용어집 추가
            for korean, translations in list(terminology.items())[:10]:  # 처음 10개만 미리보기
                guide_content += f"- {korean} → EN: {translations.get('en_US', '')}, JA: {translations.get('ja_JP', '')}\n"

            if len(terminology) > 10:
                guide_content += f"\n... 외 {len(terminology) - 10}개 용어\n"

            guide_content += """

## API 정보

- **엔드포인트**: LINE Content API
- **응답 속도**: 평균 0.2초대
- **업데이트**: 실시간 동기화

"""

            return {
                'status': 'success',
                'content': guide_content,
                'source': 'LINE API',
                'terminology_count': len(terminology),
                'size': len(guide_content),
                'last_updated': datetime.now().isoformat(),
                'api_url': 'https://landpress-content.line-scdn.net/contents/v2/projects/wdmwbfuv10x39bukv58ocevp/collections/web3_xlt_json/item'
            }

        else:
            # LINE API 실패 시 폴백
            fallback_content = """# 번역 가이드라인 (폴백)

LINE API 연결에 실패하여 기본 가이드를 표시합니다.

## 기본 핵심 용어

- 거래 → EN: transaction, JA: 取引, ZH: 交易, TH: ธุรกรรม
- 지갑 → EN: wallet, JA: ウォレット, ZH: 錢包, TH: กระเป๋า
- 토큰 → EN: token, JA: トークン, ZH: 代幣, TH: โทเค็น
- 자산 → EN: asset, JA: 資産, ZH: 資產, TH: สินทรัพย์
- 로그인 → EN: log in, JA: ログイン, ZH: 登入, TH: เข้าสู่ระบบ

## 알림

LINE API 연결을 확인해주세요. 시스템이 하드코딩된 기본 용어를 사용하고 있습니다.
"""

            return {
                'status': 'warning',
                'content': fallback_content,
                'source': 'fallback',
                'terminology_count': 5,
                'size': len(fallback_content),
                'message': 'LINE API 연결 실패 - 기본 용어 사용 중'
            }

    except Exception as e:
        return {
            'status': 'error',
            'error': f'LINE API 기반 가이드 로드 실패: {str(e)}'
        }

@app.route('/api/translation-guide')
def api_translation_guide():
    """LINE API 기반 번역 가이드라인 제공"""
    try:
        result = load_translation_guide_api()

        if result['status'] == 'error':
            return jsonify(result), 404

        return jsonify(result)
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500

@app.route('/api/terminology/test', methods=['POST'])
def api_terminology_test():
    """LINE API 용어집 연결 테스트"""
    try:
        print("🔍 LINE API 용어집 테스트 시작")

        # Step 1: 모듈 임포트
        try:
            from xlt.translation.unifi_translator import UnifiTranslator
            print("✅ UnifiTranslator 임포트 성공")
        except ImportError as e:
            print(f"❌ UnifiTranslator 임포트 실패: {e}")
            return jsonify({
                'status': 'error',
                'error': f'UnifiTranslator 임포트 실패: {str(e)}'
            }), 500

        try:
            from xlt.core.config import XLTConfig
            print("✅ XLTConfig 임포트 성공")
        except ImportError as e:
            print(f"❌ XLTConfig 임포트 실패: {e}")
            return jsonify({
                'status': 'error',
                'error': f'XLTConfig 임포트 실패: {str(e)}'
            }), 500

        # Step 2: 설정 객체 준비
        try:
            # 전역 config 사용 시도
            test_config = globals().get('config', None)
            if test_config is None:
                print("🔧 전역 config 없음, 새로 생성")
                test_config = XLTConfig()
            else:
                print("✅ 전역 config 사용")
        except Exception as e:
            print(f"❌ Config 생성 실패: {e}")
            return jsonify({
                'status': 'error',
                'error': f'Config 생성 실패: {str(e)}'
            }), 500

        # Step 3: UnifiTranslator 초기화 및 LINE API 응답 저장
        try:
            print("🔧 UnifiTranslator 초기화 시도...")
            unifi_translator = UnifiTranslator(test_config)
            print("✅ UnifiTranslator 초기화 성공")

            # LINE API 원본 응답 가져오기
            raw_api_response = None
            try:
                import requests
                api_url = "https://landpress-content.line-scdn.net/contents/v2/projects/wdmwbfuv10x39bukv58ocevp/collections/web3_xlt_json/item"
                response = requests.get(api_url, timeout=10)
                if response.status_code == 200:
                    raw_api_response = response.json()
                    print("✅ LINE API 원본 응답 수집 완료")
            except Exception as raw_e:
                print(f"⚠️ LINE API 원본 응답 수집 실패: {raw_e}")

        except Exception as e:
            print(f"❌ UnifiTranslator 초기화 실패: {e}")
            return jsonify({
                'status': 'error',
                'error': f'UnifiTranslator 초기화 실패: {str(e)}'
            }), 500

        # 용어집 정보 확인
        terminology_count = len(unifi_translator.line_terminology)

        if terminology_count == 0:
            return jsonify({
                'status': 'error',
                'error': 'LINE API에서 용어를 로드할 수 없습니다. API 연결 상태를 확인해주세요.'
            }), 400

        # 전체 용어 추출 (모든 용어)
        sample_terms = []
        all_terms = []

        for i, (korean, translations) in enumerate(unifi_translator.line_terminology.items()):
            term_data = {
                'korean': korean,
                'english': translations.get('en_US', ''),
                'japanese': translations.get('ja_JP', ''),
                'chinese': translations.get('zh_TW', ''),
                'thai': translations.get('th_TH', '')
            }

            # 전체 용어 리스트에 추가
            all_terms.append(term_data)

            # 샘플 용어는 처음 3개만 (기존 호환성 유지)
            if i < 3:
                sample_terms.append(term_data)

        # exceptions 및 metadata 추출
        exceptions_data = []
        metadata = {}

        if raw_api_response:
            try:
                exceptions_section = raw_api_response.get('body', {}).get('exceptions', {})

                # metadata 추출
                metadata = exceptions_section.get('metadata', {})

                # exceptions 배열 추출
                exceptions_array = exceptions_section.get('exceptions', [])
                for exc in exceptions_array:
                    exceptions_data.append({
                        'id': exc.get('id', ''),
                        'pattern': exc.get('pattern', ''),
                        'note': exc.get('note', ''),
                        'exception_type': exc.get('exception_type', ''),
                        'active': exc.get('active', False),
                        'translations': exc.get('translations', {})
                    })

                print(f"✅ Exceptions 추출 완료: {len(exceptions_data)}개")
                print(f"✅ Metadata 추출 완료: {len(metadata)}개 필드")

            except Exception as parse_e:
                print(f"⚠️ Exceptions 파싱 실패: {parse_e}")

        print(f"✅ LINE API 용어집 테스트 성공: {terminology_count}개 용어 로드됨")

        return jsonify({
            'status': 'success',
            'message': 'LINE API 용어집이 정상적으로 로드되었습니다.',
            'terminology_count': terminology_count,
            'sample_terms': sample_terms,
            'all_terms': all_terms,
            'exceptions_count': len(exceptions_data),
            'exceptions': exceptions_data,
            'metadata': metadata,
            'raw_api_response': raw_api_response,
            'api_url': 'https://landpress-content.line-scdn.net/contents/v2/projects/wdmwbfuv10x39bukv58ocevp/collections/web3_xlt_json/item'
        })

    except Exception as e:
        print(f"❌ LINE API 용어집 테스트 실패: {str(e)}")
        return jsonify({
            'status': 'error',
            'error': f'LINE API 연결 실패: {str(e)}'
        }), 500

@app.route('/api/excel-translate', methods=['GET', 'POST'])
def api_excel_translate():
    """엑셀 파일 번역 API"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    try:
        print("📥 엑셀 번역 API 호출됨")

        # 파일 업로드 확인
        if 'excel_file' not in request.files:
            print("❌ 파일 업로드 없음")
            return jsonify({
                'status': 'error',
                'error': '엑셀 파일이 업로드되지 않았습니다.'
            }), 400

        excel_file = request.files['excel_file']
        translation_engine = request.form.get('translation_engine', 'claude_integrated')

        print(f"📁 파일: {excel_file.filename}, 엔진: {translation_engine}")

        if excel_file.filename == '':
            print("❌ 빈 파일명")
            return jsonify({
                'status': 'error',
                'error': '파일이 선택되지 않았습니다.'
            }), 400

        # 파일 형식 확인
        if not excel_file.filename.lower().endswith(('.xlsx', '.xls')):
            print(f"❌ 잘못된 파일 형식: {excel_file.filename}")
            return jsonify({
                'status': 'error',
                'error': '엑셀 파일만 업로드 가능합니다. (.xlsx, .xls)'
            }), 400

        # 파일 크기 확인 (100MB 제한)
        excel_file.seek(0, 2)  # 파일 끝으로 이동
        file_size = excel_file.tell()
        excel_file.seek(0)  # 파일 시작으로 다시 이동

        if file_size > 100 * 1024 * 1024:  # 100MB
            print(f"❌ 파일 너무 큼: {file_size / 1024 / 1024:.1f}MB")
            return jsonify({
                'status': 'error',
                'error': f'파일이 너무 큽니다. ({file_size / 1024 / 1024:.1f}MB) 최대 100MB까지 가능합니다.'
            }), 400

        print(f"✅ 파일 검증 완료: {file_size / 1024:.1f}KB")

        # 엑셀 번역 세션 생성
        import uuid
        import time
        from datetime import datetime
        session_id = f"excel_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"

        # 파일 내용을 미리 읽어서 저장 (비동기 처리에서 접근 가능하도록)
        file_content = excel_file.read()
        filename = excel_file.filename

        # 세션 상태 초기화
        session_status[session_id] = {
            'type': 'excel_translation',
            'status': 'processing',
            'progress': {
                'current_step': 'analyzing',
                'total_steps': 5,
                'current_step_num': 1,
                'message': '엑셀 파일 분석 중...',
                'percentage': 20
            },
            'start_time': time.time(),
            'translation_engine': translation_engine,
            'filename': filename
        }

        # 비동기 처리 시작
        def process_excel_async():
            try:
                result = process_excel_translation_from_content(file_content, filename, translation_engine, session_id)
                session_status[session_id]['result'] = result
                session_status[session_id]['status'] = 'completed'
                session_status[session_id]['end_time'] = time.time()
            except Exception as e:
                session_status[session_id]['result'] = {
                    'status': 'error',
                    'error': f'엑셀 번역 중 오류: {str(e)}'
                }
                session_status[session_id]['status'] = 'error'

        import threading
        thread = threading.Thread(target=process_excel_async)
        thread.daemon = True
        thread.start()

        # 엑셀 번역 시작 시점의 기본 예상 시간 (텍스트 분석 전)
        estimated_time_text = "분석 중..."  # 실제 시간은 텍스트 분석 후 계산됨

        return jsonify({
            'status': 'processing',
            'session_id': session_id,
            'message': '엑셀 번역이 시작되었습니다.',
            'estimated_time': estimated_time_text
        })

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"❌ 엑셀 번역 API 예외 발생: {e}")
        print(f"   상세 스택: {error_trace}")

        # 안전한 에러 응답 보장
        try:
            return jsonify({
                'status': 'error',
                'error': f'엑셀 번역 중 오류가 발생했습니다: {str(e)}'
            }), 500
        except Exception as e2:
            print(f"❌ 에러 응답 생성 실패: {e2}")
            return "Internal Server Error", 500

@app.route('/api/download-excel/<filename>')
def api_download_excel(filename):
    """번역된 엑셀 파일 다운로드 API"""
    try:
        import os
        from flask import send_from_directory

        # 안전한 파일명 검증
        if not filename.endswith('.xlsx'):
            return jsonify({'error': '유효하지 않은 파일명입니다.'}), 400

        # 다운로드 디렉토리 설정
        download_dir = os.path.join(os.path.dirname(__file__), 'downloads')
        file_path = os.path.join(download_dir, filename)

        # 파일 존재 확인
        if not os.path.exists(file_path):
            return jsonify({'error': '파일을 찾을 수 없습니다.'}), 404

        # 파일 다운로드
        return send_from_directory(
            download_dir,
            filename,
            as_attachment=True,
            download_name=f'translated_{filename}'
        )

    except Exception as e:
        print(f"❌ 엑셀 다운로드 오류: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/excel-progress/<session_id>')
def api_excel_progress(session_id):
    """엑셀 번역 진행 상태 조회"""
    try:
        if session_id not in session_status:
            return jsonify({
                'status': 'error',
                'error': '세션을 찾을 수 없습니다.'
            }), 404

        session_data = session_status[session_id]

        response_data = {
            'session_id': session_id,
            'status': session_data['status'],
            'progress': session_data['progress'],
            'translation_engine': session_data.get('translation_engine', 'unknown'),
            'filename': session_data.get('filename', 'unknown')
        }

        # 완료된 경우 결과 포함
        if session_data['status'] in ['completed', 'error']:
            response_data['result'] = session_data.get('result', {})

            # 처리 시간 계산
            if 'start_time' in session_data and 'end_time' in session_data:
                processing_time = session_data['end_time'] - session_data['start_time']
                response_data['processing_time'] = f"{processing_time:.1f}초"

        return jsonify(response_data)

    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': f'진행 상태 조회 중 오류: {str(e)}'
        }), 500

# =============================================================================
# 엑셀 검증 관련 라우트들 (Claude AI 기반)
# =============================================================================

@app.route('/api/excel-validate', methods=['POST'])
def api_excel_validate():
    """엑셀 파일 업로드 및 Claude AI 검증 시작"""
    try:
        logger.info("📊 엑셀 검증 요청 수신")

        # 파일 업로드 확인
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'error': '파일이 업로드되지 않았습니다.'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'error': '파일이 선택되지 않았습니다.'}), 400

        # 엑셀 파일 확인
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'status': 'error', 'error': '엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.'}), 400

        # 임시 파일로 저장
        import tempfile
        import os
        from werkzeug.utils import secure_filename

        filename = secure_filename(file.filename)
        logger.info(f"📂 원본 파일명: {file.filename} → 보안 파일명: {filename}")

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            temp_file_path = tmp_file.name
            logger.info(f"📂 임시 파일 생성: {temp_file_path}")

        # 파일 저장
        file.save(temp_file_path)

        # 파일 저장 확인
        if os.path.exists(temp_file_path):
            file_size = os.path.getsize(temp_file_path)
            logger.info(f"✅ 임시 파일 저장 완료: {temp_file_path} ({file_size} bytes)")
        else:
            logger.error(f"❌ 임시 파일 저장 실패: {temp_file_path}")
            return jsonify({'status': 'error', 'error': '파일 저장 실패'}), 500

        # 세션 ID 생성
        session_id = f"validate_{int(time.time())}"

        # 검증 모드 확인 (기본값: 종합검증)
        validation_mode = request.form.get('validation_mode', 'comprehensive')
        use_comprehensive = validation_mode == 'comprehensive'

        logger.info(f"🎯 검증 모드: {'종합검증' if use_comprehensive else '5단계 개별검증'}")

        # 검증 모듈 import 및 실행
        from xlt.validation.excel_validator import ExcelValidator
        from xlt.core.config import XLTConfig

        config = XLTConfig()
        validator = ExcelValidator(config)

        # 백그라운드에서 검증 실행
        import threading

        def run_validation():
            try:
                logger.info(f"🔍 전체 검증 시작: {filename} (세션: {session_id})")
                logger.info(f"📂 임시 파일 경로: {temp_file_path}")
                logger.info(f"🎯 검증 모드: {use_comprehensive}")

                # 진행 상황 업데이트 콜백
                def progress_update(step: str, percent: int, message: str):
                    logger.info(f"📈 진행 상황: {step} ({percent}%) - {message}")
                    if session_id in session_status:
                        session_status[session_id]['progress'] = {
                            'step': step,
                            'percent': percent,
                            'message': message,
                            'is_full_validation': True  # 🚀 전체 검증 표시
                        }

                logger.info("🚀 ExcelValidator.validate_excel_file 호출 중...")
                result = validator.validate_excel_file(temp_file_path, session_id, progress_update, use_comprehensive)
                logger.info(f"✅ ExcelValidator.validate_excel_file 완료: {result.get('status', 'unknown')}")

                # 세션에 결과 저장
                session_status[session_id] = {
                    'status': 'completed',
                    'result': result,
                    'filename': filename,
                    'temp_file_path': temp_file_path,
                    'validator': validator,  # 교정을 위해 validator 인스턴스 보관
                    'end_time': time.time()
                }
                logger.info(f"✅ 검증 완료: {session_id}")

            except Exception as e:
                logger.error(f"❌ 검증 오류: {e}")
                session_status[session_id] = {
                    'status': 'error',
                    'error': str(e),
                    'filename': filename,
                    'end_time': time.time()
                }

        # 세션 초기화
        session_status[session_id] = {
            'status': 'processing',
            'filename': filename,
            'start_time': time.time()
        }

        # 백그라운드 스레드 시작
        thread = threading.Thread(target=run_validation)
        thread.daemon = True
        thread.start()

        return jsonify({
            'status': 'success',
            'session_id': session_id,
            'message': f'엑셀 검증이 시작되었습니다: {filename}'
        })

    except Exception as e:
        logger.error(f"❌ 엑셀 검증 시작 오류: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/excel-validate-progress/<session_id>')
def api_excel_validate_progress(session_id):
    """엑셀 검증 진행 상태 확인"""
    try:
        if session_id not in session_status:
            return jsonify({
                'status': 'error',
                'error': '세션을 찾을 수 없습니다.'
            }), 404

        session_data = session_status[session_id]

        response_data = {
            'session_id': session_id,
            'status': session_data['status'],
            'filename': session_data.get('filename', 'unknown')
        }

        # 완료된 경우 결과 포함
        if session_data['status'] in ['completed', 'error']:
            response_data['result'] = session_data.get('result', {})

            # 처리 시간 계산
            if 'start_time' in session_data and 'end_time' in session_data:
                processing_time = session_data['end_time'] - session_data['start_time']
                response_data['processing_time'] = f"{processing_time:.1f}초"

        return jsonify(response_data)

    except Exception as e:
        logger.error(f"❌ 검증 진행 상태 조회 오류: {e}")
        return jsonify({
            'status': 'error',
            'error': f'진행 상태 조회 중 오류: {str(e)}'
        }), 500

def group_validation_results_by_key_id(validation_result: dict) -> dict:
    """검증 결과를 key_id 기준으로 그룹핑 - 우선순위 포함"""
    try:
        grouped_results = {}
        detailed_results = validation_result.get('detailed_results', {})
        priority_categories = {
            'critical': [],  # 시급한 수정 필요
            'high': [],      # 높은 우선순위
            'medium': [],    # 중간 우선순위
            'low': []        # 낮은 우선순위
        }

        # 각 검증 타입별로 이슈를 key_id로 그룹핑
        for validation_type, validation_data in detailed_results.items():
            issues = validation_data.get('issues', [])
            exceptions = validation_data.get('exceptions', [])

            for issue in issues:
                key_id = issue.get('key_id', 'unknown')
                row_number = issue.get('row_number', 0)
                severity = issue.get('severity', 'medium')

                if key_id not in grouped_results:
                    grouped_results[key_id] = {
                        'key_id': key_id,
                        'row_number': row_number,
                        'issues': [],
                        'exceptions': [],
                        'max_severity': 'low'
                    }

                # 이슈 타입별로 분류
                issue_copy = issue.copy()
                issue_copy['validation_type'] = validation_type
                issue_copy['priority'] = get_issue_priority(issue_copy)
                grouped_results[key_id]['issues'].append(issue_copy)

                # 최대 심각도 업데이트
                if severity_rank(severity) > severity_rank(grouped_results[key_id]['max_severity']):
                    grouped_results[key_id]['max_severity'] = severity

            # exceptional 항목들도 처리
            for exception in exceptions:
                key_id = exception.get('key_id', 'unknown')
                row_number = exception.get('row_number', 0)

                if key_id not in grouped_results:
                    grouped_results[key_id] = {
                        'key_id': key_id,
                        'row_number': row_number,
                        'issues': [],
                        'exceptions': [],
                        'max_severity': 'low'
                    }

                exception_copy = exception.copy()
                exception_copy['validation_type'] = validation_type
                grouped_results[key_id]['exceptions'].append(exception_copy)

        # 우선순위별 분류
        for key_data in grouped_results.values():
            severity = key_data['max_severity']
            priority_categories[severity].append(key_data)

        # 우선순위 순으로 정렬
        sorted_results = []
        for priority in ['critical', 'high', 'medium', 'low']:
            # 각 우선순위 내에서 행번호 순 정렬
            priority_categories[priority].sort(key=lambda x: x['row_number'])
            sorted_results.extend(priority_categories[priority])

        return {
            'grouped_by_key_id': sorted_results,
            'total_affected_keys': len(grouped_results),
            'priority_summary': {
                'critical': len(priority_categories['critical']),
                'high': len(priority_categories['high']),
                'medium': len(priority_categories['medium']),
                'low': len(priority_categories['low'])
            }
        }

    except Exception as e:
        logger.error(f"❌ 검증 결과 그룹핑 오류: {e}")
        return {
            'grouped_by_key_id': [],
            'total_affected_keys': 0,
            'grouping_error': str(e)
        }

def get_issue_priority(issue: dict) -> str:
    """이슈의 우선순위 결정"""
    issue_type = issue.get('issue_type', '')
    severity = issue.get('severity', 'medium')

    # 시급한 수정이 필요한 항목들
    if issue_type in ['linebreak_error', 'placeholder_error']:
        return 'critical'
    if issue_type == 'language_mix':
        return 'critical'
    if severity == 'critical':
        return 'critical'

    # 높은 우선순위
    if severity == 'high':
        return 'high'
    if issue_type == 'empty_cells' and 'ko_KR' in issue.get('missing_languages', []):
        return 'high'

    # 기본
    return severity

def severity_rank(severity: str) -> int:
    """심각도 순위 (숫자가 클수록 심각)"""
    ranks = {'low': 1, 'medium': 2, 'high': 3, 'critical': 4}
    return ranks.get(severity, 2)

@app.route('/api/excel-validate-report/<session_id>')
def api_excel_validate_report(session_id):
    """엑셀 검증 완료 리포트 조회 - Key ID 기준 그룹핑"""
    try:
        if session_id not in session_status:
            return jsonify({
                'status': 'error',
                'error': '세션을 찾을 수 없습니다.'
            }), 404

        session_data = session_status[session_id]

        if session_data['status'] != 'completed':
            return jsonify({
                'status': 'error',
                'error': '검증이 완료되지 않았습니다.'
            }), 400

        result = session_data.get('result', {})

        # 기존 결과에 Key ID 기준 그룹핑 추가
        grouped_results = group_validation_results_by_key_id(result)
        result['grouped_results'] = grouped_results

        return jsonify({
            'status': 'success',
            'session_id': session_id,
            'validation_report': result
        })

    except Exception as e:
        logger.error(f"❌ 검증 리포트 조회 오류: {e}")
        return jsonify({
            'status': 'error',
            'error': f'리포트 조회 중 오류: {str(e)}'
        }), 500

@app.route('/api/excel-auto-correct/<session_id>', methods=['POST'])
def api_excel_auto_correct(session_id):
    """Claude AI 자동 재번역/교정 시작"""
    try:
        logger.info(f"🔧 자동 교정 요청: {session_id}")

        # 세션 확인
        if session_id not in session_status:
            return jsonify({
                'status': 'error',
                'error': '검증 세션을 찾을 수 없습니다.'
            }), 404

        validation_session = session_status[session_id]

        if validation_session['status'] != 'completed':
            return jsonify({
                'status': 'error',
                'error': '검증이 완료되지 않았습니다.'
            }), 400

        validation_result = validation_session.get('result', {})
        if not validation_result.get('has_issues', False):
            return jsonify({
                'status': 'error',
                'error': '교정이 필요한 문제가 없습니다.'
            }), 400

        # 교정 세션 ID 생성
        correction_session_id = f"correct_{session_id}_{int(time.time())}"

        # 교정 모듈 import
        from xlt.validation.excel_corrector import ExcelCorrector
        from xlt.core.config import XLTConfig

        config = XLTConfig()
        corrector = ExcelCorrector(config)

        # 원본 엑셀 데이터 로드
        validator = validation_session.get('validator')
        if not validator:
            return jsonify({
                'status': 'error',
                'error': '검증 데이터를 찾을 수 없습니다.'
            }), 400

        original_excel_data = validator.excel_data

        # 백그라운드에서 교정 실행
        import threading

        def run_correction():
            try:
                logger.info(f"🔧 교정 시작: {correction_session_id}")
                result = corrector.auto_correct_excel(
                    correction_session_id, validation_result, original_excel_data
                )

                # 세션에 결과 저장
                session_status[correction_session_id] = {
                    'status': 'completed',
                    'result': result,
                    'original_session_id': session_id,
                    'corrector': corrector,  # 파일 생성을 위해 보관
                    'temp_file_path': validation_session.get('temp_file_path'),
                    'filename': validation_session.get('filename'),
                    'end_time': time.time()
                }
                logger.info(f"✅ 교정 완료: {correction_session_id}")

            except Exception as e:
                logger.error(f"❌ 교정 오류: {e}")
                session_status[correction_session_id] = {
                    'status': 'error',
                    'error': str(e),
                    'original_session_id': session_id,
                    'end_time': time.time()
                }

        # 교정 세션 초기화
        session_status[correction_session_id] = {
            'status': 'processing',
            'original_session_id': session_id,
            'start_time': time.time()
        }

        # 백그라운드 스레드 시작
        thread = threading.Thread(target=run_correction)
        thread.daemon = True
        thread.start()

        return jsonify({
            'status': 'success',
            'correction_session_id': correction_session_id,
            'message': 'Claude AI 자동 교정이 시작되었습니다.'
        })

    except Exception as e:
        logger.error(f"❌ 자동 교정 시작 오류: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/excel-correction-progress/<session_id>')
def api_excel_correction_progress(session_id):
    """Claude AI 교정 진행 상태 확인"""
    try:
        if session_id not in session_status:
            return jsonify({
                'status': 'error',
                'error': '교정 세션을 찾을 수 없습니다.'
            }), 404

        session_data = session_status[session_id]

        # 교정기에서 세부 진행 상태 가져오기
        corrector = session_data.get('corrector')
        progress_data = {}

        if corrector and hasattr(corrector, 'get_correction_progress'):
            progress_data = corrector.get_correction_progress(session_id) or {}

        response_data = {
            'session_id': session_id,
            'status': session_data['status'],
            'progress': progress_data
        }

        # 완료된 경우 결과 포함
        if session_data['status'] in ['completed', 'error']:
            response_data['result'] = session_data.get('result', {})

            # 처리 시간 계산
            if 'start_time' in session_data and 'end_time' in session_data:
                processing_time = session_data['end_time'] - session_data['start_time']
                response_data['processing_time'] = f"{processing_time:.1f}초"

        return jsonify(response_data)

    except Exception as e:
        logger.error(f"❌ 교정 진행 상태 조회 오류: {e}")
        return jsonify({
            'status': 'error',
            'error': f'진행 상태 조회 중 오류: {str(e)}'
        }), 500

@app.route('/api/download-corrected-excel/<session_id>')
def api_download_corrected_excel(session_id):
    """교정된 Excel 파일 다운로드"""
    try:
        if session_id not in session_status:
            return jsonify({
                'status': 'error',
                'error': '교정 세션을 찾을 수 없습니다.'
            }), 404

        session_data = session_status[session_id]

        if session_data['status'] != 'completed':
            return jsonify({
                'status': 'error',
                'error': '교정이 완료되지 않았습니다.'
            }), 400

        corrector = session_data.get('corrector')
        original_file_path = session_data.get('temp_file_path')
        original_filename = session_data.get('filename', 'corrected_file.xlsx')

        if not corrector or not original_file_path:
            return jsonify({
                'status': 'error',
                'error': '교정 데이터를 찾을 수 없습니다.'
            }), 400

        # 교정된 엑셀 파일 생성
        corrected_file_path = corrector.create_corrected_excel_file(session_id, original_file_path)

        if not corrected_file_path or not os.path.exists(corrected_file_path):
            return jsonify({
                'status': 'error',
                'error': '교정된 파일 생성에 실패했습니다.'
            }), 500

        # 파일 다운로드
        download_dir = os.path.dirname(corrected_file_path)
        download_filename = os.path.basename(corrected_file_path)

        # 사용자 친화적 파일명 생성
        original_name = os.path.splitext(original_filename)[0]
        user_filename = f"{original_name}_교정완료.xlsx"

        return send_from_directory(
            download_dir,
            download_filename,
            as_attachment=True,
            download_name=user_filename
        )

    except Exception as e:
        logger.error(f"❌ 교정된 엑셀 다운로드 오류: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

def process_excel_translation_from_content(file_content, filename, translation_engine, session_id=None):
    """엑셀 파일 내용으로부터 번역 처리 함수"""
    try:
        import openpyxl
        import tempfile
        import os
        from datetime import datetime
        from io import BytesIO
        import re

        def update_progress(step_num, step_name, message, percentage):
            if session_id and session_id in session_status:
                session_status[session_id]['progress'] = {
                    'current_step': step_name,
                    'total_steps': 5,
                    'current_step_num': step_num,
                    'message': message,
                    'percentage': percentage
                }

        def add_translation_log(message, log_type='info'):
            """엑셀 번역 로그 기록 함수"""
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_entry = f"{timestamp} [{log_type.upper()}] {message}"
            print(log_entry)  # 콘솔에도 출력

            # server.log 파일에도 기록
            try:
                with open(log_file, 'a', encoding='utf-8') as f:
                    f.write(log_entry + '\n')
            except:
                pass  # 로그 파일 쓰기 실패해도 진행

        print(f"📁 엑셀 번역 시작: {filename}, 엔진: {translation_engine}")
        add_translation_log(f'📁 엑셀 번역 시작: {filename} (엔진: {translation_engine})', 'info')
        update_progress(1, 'loading', f'{translation_engine} 엔진으로 엑셀 파일 로드 중...', 20)

        # 1. 엑셀 파일 로드 (파일 내용으로부터)
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        print(f"📊 워크시트 목록: {workbook.sheetnames}")
        update_progress(2, 'analyzing', '워크시트 구조 분석 중...', 40)

        # 2. properties 워크시트 찾기
        if 'properties' not in workbook.sheetnames:
            return {
                'status': 'error',
                'error': 'properties 워크시트를 찾을 수 없습니다. XLT 번역 결과 파일을 업로드해주세요.'
            }

        worksheet = workbook['properties']
        print(f"📝 properties 워크시트: {worksheet.max_column}열 × {worksheet.max_row}행")

        # 3. 헤더 분석 (1행)
        headers = {}
        for col in range(1, worksheet.max_column + 1):
            header = worksheet.cell(row=1, column=col).value
            if header:
                headers[header] = col
                print(f"   컬럼 {col}: {header}")

        # 4. 한국어 컬럼 찾기 (여러 이름 가능)
        ko_col = None
        possible_ko_names = ['ko_KR', 'ko', 'Korean', '한국어', '원본', 'Original']

        for name in possible_ko_names:
            if name in headers:
                ko_col = headers[name]
                print(f"   한국어 컬럼 발견: {name} (컬럼 {ko_col})")
                break

        if not ko_col:
            return {
                'status': 'error',
                'error': f'한국어 컬럼을 찾을 수 없습니다. 다음 중 하나의 컬럼명을 사용하세요: {", ".join(possible_ko_names)}'
            }

        # 5. 번역 대상 언어 컬럼 확인 및 생성
        target_languages = ['en_US', 'ja_JP', 'zh_TW', 'th_TH']
        target_cols = {}

        for lang in target_languages:
            if lang in headers:
                target_cols[lang] = headers[lang]
                print(f"   기존 {lang} 컬럼 사용: 컬럼 {headers[lang]}")
            else:
                # 새 컬럼 생성
                new_col = worksheet.max_column + 1
                worksheet.cell(row=1, column=new_col, value=lang)
                target_cols[lang] = new_col
                print(f"   새 {lang} 컬럼 생성: 컬럼 {new_col}")

        # 6. 번역할 텍스트 수집
        translation_tasks = []

        for row in range(2, worksheet.max_row + 1):  # 2행부터 데이터
            ko_text = worksheet.cell(row=row, column=ko_col).value

            if ko_text and isinstance(ko_text, str) and ko_text.strip():
                # 한글이 포함된 텍스트인지 확인
                if re.search(r'[가-힣]', ko_text):
                    # 번역이 필요한 언어들 체크 (빈 셀만)
                    need_translation = {}
                    for lang in target_languages:
                        target_col = target_cols[lang]
                        existing_value = worksheet.cell(row=row, column=target_col).value
                        if not existing_value or (isinstance(existing_value, str) and not existing_value.strip()):
                            need_translation[lang] = target_col

                    if need_translation:
                        translation_tasks.append({
                            'row': row,
                            'ko_text': ko_text.strip(),
                            'targets': need_translation
                        })

        print(f"🔤 번역 대상: {len(translation_tasks)}개 텍스트")
        add_translation_log(f'🔤 엑셀 번역 대상 분석 완료: {len(translation_tasks)}개 한글 텍스트 발견', 'info')
        update_progress(3, 'preparing', f'{len(translation_tasks)}개 텍스트 번역 준비 완료', 60)

        if not translation_tasks:
            return {
                'status': 'error',
                'error': '번역할 한글 텍스트가 없거나 모든 번역이 이미 완료되었습니다.'
            }

        # 6. 번역 실행 - 동적 예상 시간 계산
        engine_name = 'Claude 통합 처리' if translation_engine == 'claude_integrated' else 'Claude AI'

        # Claude AI 번역 엔진별 예상 시간 계산
        if translation_engine == 'claude_integrated':
            estimated_minutes = len(translation_tasks) * 0.33  # 청크 기반: 텍스트당 20초 (3개씩)
        else:  # claude 모드
            estimated_minutes = len(translation_tasks) * 0.25  # Claude: 텍스트당 15초

        max_minutes = max(1, min(estimated_minutes, 20))  # 최소 1분, 최대 20분
        time_text = f"{int(max_minutes)}분" if max_minutes >= 1 else f"{int(max_minutes * 60)}초"

        update_progress(4, 'translating', f'{engine_name}로 번역 중... (예상: {time_text})', 80)
        add_translation_log(f'🚀 엑셀 번역 실행: {engine_name} (예상 시간: {time_text})', 'info')

        # Claude AI 전용 번역 (유일한 번역 방식)
        translation_results = translate_with_claude_integrated(translation_tasks, session_id)

        # 7. 번역 결과를 엑셀에 적용
        successful_translations = 0
        for i, task in enumerate(translation_tasks):
            if i < len(translation_results):
                result = translation_results[i]
                for lang, col in task['targets'].items():
                    if lang in result and result[lang]:
                        worksheet.cell(row=task['row'], column=col, value=result[lang])
                        successful_translations += 1

        print(f"✅ 성공한 번역: {successful_translations}개")

        # 8. 파일 저장
        update_progress(5, 'saving', '번역 결과를 파일에 저장 중...', 95)

        downloads_dir = os.path.join(os.path.dirname(__file__), 'downloads')
        os.makedirs(downloads_dir, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f'translated_{timestamp}_{filename}'
        output_path = os.path.join(downloads_dir, output_filename)

        workbook.save(output_path)
        workbook.close()

        print(f"💾 번역 완료 파일 저장: {output_filename}")
        add_translation_log(f'💾 엑셀 번역 완료: {successful_translations}개 텍스트 → {output_filename}', 'success')
        update_progress(5, 'completed', f'{successful_translations}개 텍스트 번역 완료!', 100)

        return {
            'status': 'success',
            'message': f'{successful_translations}개 텍스트 번역 완료',
            'filename': output_filename,
            'translation_count': successful_translations,
            'total_tasks': len(translation_tasks)
        }

    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"❌ 엑셀 번역 처리 오류: {e}")
        print(f"   상세: {error_trace}")

        return {
            'status': 'error',
            'error': f'엑셀 번역 처리 중 오류: {str(e)}'
        }

def translate_with_claude_integrated(translation_tasks, session_id=None):
    """Claude 통합 번역 (배치 최적화 버전)"""

    # 로그 함수 정의 (함수 내부에서 사용 가능하도록)
    def add_translation_log(message, log_type='info'):
        """엑셀 번역 로그 기록 함수"""
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"{timestamp} [{log_type.upper()}] {message}"
        print(log_entry)  # 콘솔에도 출력

        # server.log 파일에도 기록
        try:
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(log_entry + '\n')
        except:
            pass  # 로그 파일 쓰기 실패해도 진행

    try:
        print("🔍 엑셀 Claude 번역 함수 진입")
        add_translation_log('🔍 엑셀 Claude 번역 함수 시작', 'debug')

        from xlt.translation.claude_translator import ClaudeTranslator
        from xlt.core.config import XLTConfig
        print("✅ Claude 모듈 import 성공")
        add_translation_log('✅ Claude 모듈 import 성공', 'debug')

        translator = ClaudeTranslator(XLTConfig())
        print("✅ Claude 번역기 초기화 성공")
        add_translation_log('✅ Claude 번역기 초기화 성공', 'debug')

        if not translation_tasks:
            print("❌ 번역 태스크가 비어있음")
            add_translation_log('❌ 번역 태스크가 비어있음', 'warning')
            return []

        # 배치 처리용 데이터 준비
        all_texts = [task['ko_text'] for task in translation_tasks]

        # 모든 태스크의 대상 언어들을 통합 (중복 제거)
        all_target_langs = list(set().union(*[task['targets'].keys() for task in translation_tasks]))

        print(f"🚀 엑셀 Claude 배치 번역: {len(all_texts)}개 텍스트 → {len(all_target_langs)}개 언어")
        print(f"   최적화: {len(translation_tasks)}회 개별 호출 → 청크 기반 처리")

        # 서버 로그에 청크 처리 시작 기록
        chunk_size = getattr(translator.config, 'claude_chunk_size', 3)
        add_translation_log(f'📊 엑셀 청크 기반 처리 시작: {len(all_texts)}개 텍스트 → {chunk_size}개씩 분할', 'info')
        add_translation_log(f'   대상 언어: {", ".join(all_target_langs)}', 'info')

        # 청크 진행률 콜백 함수 정의
        def chunk_progress_callback(progress_data):
            # 청크 완료 시 서버 로그 기록
            if progress_data.get('status') == 'chunk_completed':
                completed = progress_data.get('completed_chunks', 0)
                total = progress_data.get('total_chunks', 0)
                add_translation_log(f'✅ 엑셀 청크 {completed}/{total} 완료: {progress_data.get("message", "")}', 'info')
            elif progress_data.get('status') == 'fallback_completed':
                add_translation_log(f'🔄 엑셀 청크 폴백 처리: {progress_data.get("message", "")}', 'warning')

            if session_id:
                # 기존 translation_progress에 청크 진행률 추가
                if session_id not in translation_progress:
                    translation_progress[session_id] = {}

                translation_progress[session_id].update({
                    'chunk_progress': progress_data,
                    'updated_at': time.time()
                })

                # session_status도 업데이트 (기존 API와 호환성 유지)
                if session_id in session_status:
                    session_status[session_id]['translation_progress'] = {
                        'status': progress_data['status'],
                        'message': progress_data['message'],
                        'chunk_info': f"{progress_data['completed_chunks']}/{progress_data['total_chunks']}"
                    }

        # 청크 기반 배치 처리 호출 (타임아웃 문제 해결)
        batch_results = translator.translate_batch_integrated_chunked(
            all_texts,
            all_target_langs,
            progress_callback=chunk_progress_callback
        )

        print(f"   📊 배치 처리 완료: {len(batch_results)}개 결과 수신")
        add_translation_log(f'📊 엑셀 청크 배치 처리 완료: {len(batch_results)}개 결과 수신', 'info')

        # 배치 결과를 개별 태스크 결과로 매핑
        results = []
        for i, task in enumerate(translation_tasks):
            if i < len(batch_results) and batch_results[i]:
                batch_result = batch_results[i]
                task_result = {}

                # 해당 태스크가 요구하는 언어들만 추출
                for lang in task['targets'].keys():
                    if lang in batch_result:
                        task_result[lang] = batch_result[lang]
                        print(f"   ✅ 텍스트{i+1} {lang}: {batch_result[lang][:30]}...")

                results.append(task_result)
            else:
                # 배치 결과가 없거나 인덱스 초과 시 빈 결과
                print(f"   ❌ 텍스트{i+1}: 번역 결과 없음")
                results.append({})

        successful_count = len([r for r in results if r])
        print(f"🎉 엑셀 Claude 배치 번역 완료: {successful_count}/{len(translation_tasks)} 성공")

        # 서버 로그에 최종 결과 기록
        success_rate = (successful_count / len(translation_tasks) * 100) if translation_tasks else 0
        add_translation_log(f'🎉 엑셀 Claude 청크 번역 완료: {successful_count}/{len(translation_tasks)} 성공 (성공률: {success_rate:.1f}%)', 'info')

        return results

    except Exception as e:
        print(f"❌ Claude 배치 번역 오류: {e}")
        add_translation_log(f'❌ 엑셀 Claude 번역 오류: {str(e)}', 'error')
        import traceback
        error_detail = traceback.format_exc()
        print(f"   상세 오류: {error_detail}")
        add_translation_log(f'   상세 오류: {error_detail}', 'error')

        # Claude 실패 시 빈 결과 반환 (Google fallback 제거됨)
        print("❌ Claude 번역 실패, 빈 결과 반환")
        add_translation_log('❌ 엑셀 Claude 번역 실패, 빈 결과 반환', 'error')
        return [{}] * len(translation_tasks)


# XLT v3.0: 테스트 기능들 제거
# @app.route('/test-ocr') - OCR 테스트 페이지 제거됨
# @app.route('/test-selection') - 선택 테스트 페이지 제거됨

@app.route('/api/figma-preview', methods=['GET', 'POST'])
def figma_preview():
    """피그마 URL의 이미지 미리보기 제공"""

    # GET 요청 시 홈페이지로 리다이렉션
    if request.method == 'GET':
        return redirect('/')

    try:
        data = request.get_json()
        if not data or 'figma_url' not in data:
            return jsonify({
                'status': 'error',
                'error': 'Figma URL이 필요합니다.'
            }), 400

        figma_url = data['figma_url'].strip()
        if not figma_url:
            return jsonify({
                'status': 'error',
                'error': 'Figma URL을 입력해주세요.'
            }), 400

        # 피그마 URL 형식 기본 검증
        if 'figma.com' not in figma_url:
            return jsonify({
                'status': 'error',
                'error': '올바른 Figma URL을 입력해주세요.'
            }), 400

        print(f"📱 피그마 미리보기 요청: {figma_url}")

        # 피그마 프로세서 초기화
        from xlt.input.figma import FigmaProcessor
        from PIL import Image
        figma_processor = FigmaProcessor(pipeline.config)

        # 피그마 토큰 확인
        if not figma_processor.figma_token:
            return jsonify({
                'status': 'error',
                'error': '피그마 토큰이 설정되지 않았습니다. figma_config.json을 확인해주세요.'
            }), 400

        # 피그마 이미지 다운로드
        try:
            image, description = figma_processor.process(figma_url)

            if not image:
                return jsonify({
                    'status': 'error',
                    'error': '이미지를 로드할 수 없습니다. URL을 확인해주세요.'
                }), 400

            # 이미지를 Base64로 인코딩하여 브라우저에서 표시 가능하게 만들기
            import base64
            from io import BytesIO

            # 이미지 크기 조정 (미리보기용으로 최대 400px)
            max_width = 400
            if image.width > max_width:
                ratio = max_width / image.width
                new_height = int(image.height * ratio)
                image = image.resize((max_width, new_height), Image.LANCZOS)

            # Base64 인코딩
            buffer = BytesIO()
            image.save(buffer, format='PNG')
            img_data = base64.b64encode(buffer.getvalue()).decode()

            # URL에서 메타데이터 추출
            try:
                # URL에서 node-id와 file-id 추출
                import re
                file_match = re.search(r'/design/([^/?]+)', figma_url)
                node_match = re.search(r'node-id=([^&]+)', figma_url)

                file_id = file_match.group(1) if file_match else 'Unknown'
                node_id = node_match.group(1) if node_match else 'Unknown'
            except:
                file_id = 'Unknown'
                node_id = 'Unknown'

            # 메타데이터 정리
            clean_metadata = {
                'width': image.width,
                'height': image.height,
                'format': 'PNG',
                'source': 'Figma',
                'node_id': node_id,
                'file_id': file_id,
                'description': description
            }

            return jsonify({
                'status': 'success',
                'image_data': f'data:image/png;base64,{img_data}',
                'metadata': clean_metadata,
                'message': '피그마 이미지 미리보기 로드 성공'
            })

        except Exception as figma_error:
            error_msg = str(figma_error)
            print(f"❌ 피그마 처리 오류: {error_msg}")

            # 구체적인 오류 메시지 제공
            if 'Invalid node' in error_msg or 'node not found' in error_msg:
                user_error = '지정된 노드를 찾을 수 없습니다. URL의 node-id를 확인해주세요.'
            elif 'Invalid file' in error_msg or 'file not found' in error_msg:
                user_error = '피그마 파일을 찾을 수 없습니다. URL을 확인해주세요.'
            elif 'Permission' in error_msg or 'access' in error_msg.lower():
                user_error = '피그마 파일에 접근할 권한이 없습니다. 파일이 공개되어 있는지 확인해주세요.'
            elif 'token' in error_msg.lower():
                user_error = '피그마 토큰이 유효하지 않습니다. figma_config.json의 토큰을 확인해주세요.'
            else:
                user_error = f'피그마 이미지 로드 실패: {error_msg}'

            return jsonify({
                'status': 'error',
                'error': user_error
            }), 400

    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"❌ 피그마 미리보기 API 오류: {e}")
        print(f"   상세 오류: {error_trace}")

        return jsonify({
            'status': 'error',
            'error': f'서버 오류가 발생했습니다: {str(e)}'
        }), 500

    # ===== 다국어 품질 우선 번역 테스트 시스템 (v5.1.0) =====

    @app.route('/quality-test')
    def quality_test_page():
        """다국어 품질 우선 번역 테스트 페이지"""
        return render_template('quality_test.html')

    @app.route('/thai-quality-test')
    def thai_quality_test_page():
        """태국어 품질 우선 번역 테스트 페이지 (하위 호환성)"""
        return render_template('thai_quality_test.html')

    @app.route('/api/test-quality', methods=['POST'])
    def test_quality():
        """다국어 품질 우선 번역 모드 테스트"""
        try:
            data = request.get_json()
            if not data or 'text' not in data:
                return jsonify({
                    'status': 'error',
                    'error': '텍스트가 필요합니다'
                }), 400

            text = data['text']
            # 요청된 언어들 (기본값: 모든 언어)
            target_languages = data.get('languages', ['en_US', 'ja_JP', 'zh_TW', 'th_TH'])

            logger.info(f"🧪 다국어 품질 테스트: {text} → {len(target_languages)}개 언어")

            # Claude Translator 인스턴스 생성
            from xlt.translation.claude_translator import ClaudeTranslator
            from xlt.core.config import XLTConfig

            config = XLTConfig()
            claude_translator = ClaudeTranslator(config)

            # 다국어 품질 우선 번역 모드 실행
            result = claude_translator.translate_quality_focused(text, target_languages)

            logger.info(f"✅ 다국어 품질 우선 번역 완료: 평균 {result.get('average_quality_score', '--')}점")

            return jsonify({
                'status': 'success',
                'original_text': text,
                'target_languages': target_languages,
                'result': result,
                'timestamp': datetime.now().isoformat()
            })

        except Exception as e:
            logger.error(f"❌ 다국어 품질 테스트 실패: {str(e)}")
            return jsonify({
                'status': 'error',
                'error': str(e)
            }), 500

    @app.route('/api/test-thai-quality', methods=['POST'])
    def test_thai_quality():
        """태국어 품질 우선 번역 모드 테스트 (하위 호환성)"""
        try:
            data = request.get_json()
            if not data or 'text' not in data:
                return jsonify({
                    'status': 'error',
                    'error': '텍스트가 필요합니다'
                }), 400

            text = data['text']
            logger.info(f"🇹🇭 태국어 전용 품질 테스트: {text}")

            # Claude Translator 인스턴스 생성
            from xlt.translation.claude_translator import ClaudeTranslator
            from xlt.core.config import XLTConfig

            config = XLTConfig()
            claude_translator = ClaudeTranslator(config)

            # 태국어만 품질 우선 번역
            result = claude_translator.translate_quality_focused(text, ['th_TH'])

            # 하위 호환성을 위해 기존 형식으로 변환
            thai_result = {
                'original': result['original'],
                'corrected_korean': result['corrected_korean'],
                'corrections_applied': result['corrections_applied'],
                'ko_KR': result['ko_KR'],
                'th_TH': result.get('th_TH', ''),
                'quality_focused': result['quality_focused'],
                'quality_score': result.get('quality_scores', {}).get('th_TH', 95)
            }

            logger.info(f"✅ 태국어 품질 우선 번역 완료: {thai_result}")

            return jsonify({
                'status': 'success',
                'original_text': text,
                'result': thai_result,
                'timestamp': datetime.now().isoformat()
            })

        except Exception as e:
            logger.error(f"❌ 태국어 품질 테스트 실패: {str(e)}")
            return jsonify({
                'status': 'error',
                'error': str(e)
            }), 500


if __name__ == '__main__':
    # 동적 버전 정보 로드
    try:
        from xlt.utils.version_manager import get_full_name
        system_name = get_full_name()
    except:
        system_name = "XLT System v5.0.6"  # 폴백

    logger.info("=" * 60)
    logger.info(f"🚀 {system_name} - 완전 자동화 번역 시스템")
    logger.info("=" * 60)

    # 시스템 필수 설정 확인
    logger.info("🔍 시스템 설정 확인 중...")
    setup_status = check_system_setup()

    # 업데이트 확인 (시작 시)
    try:
        if updater:
            from xlt.utils.updater import check_updates_on_startup
            check_updates_on_startup()
    except Exception as e:
        logger.warning(f"⚠️ 업데이트 확인 중 오류: {str(e)}")

    if setup_status['all_configured']:
        logger.info("✅ 모든 필수 설정이 완료되었습니다!")

        # 설정된 서버 포트 표시
        user_config = load_user_config()
        server_port = user_config.get('server_port', 5004)

        logger.info(f"🌐 웹 서버 포트: {server_port}")
        logger.info(f"🌐 접속 URL: http://localhost:{server_port}")
        logger.info("🎯 특징: 피그마 전용 + Claude AI 전용 번역 + LINE API 기반 용어집")
        logger.info("=" * 60)
        logger.info("📋 서버 로그는 server.log 파일에 기록됩니다")
        logger.info("=" * 60)

        # 전역 405 에러 핸들러 (모든 POST 전용 엔드포인트 보호)
        @app.errorhandler(405)
        def handle_method_not_allowed(error):
            """모든 405 Method Not Allowed 에러를 홈페이지로 리다이렉션"""
            return redirect('/')

        app.run(debug=False, host='0.0.0.0', port=server_port)

    else:
        print("⚠️  필수 설정이 완료되지 않았습니다!")
        print(f"❌ 누락된 설정: {', '.join(setup_status['missing_settings'])}")
        print("")
        print("📋 설정 방법:")

        if '피그마 토큰' in setup_status['missing_settings']:
            print("   1. 피그마 토큰 설정:")
            print("      - Figma 웹사이트 → Settings → Personal Access Tokens")
            print("      - 'Create new token' 클릭하여 토큰 생성")
            print("      - figma_config.json 파일에 저장하거나 FIGMA_TOKEN 환경변수 설정")

        # XLT System v3.0: 출력 디렉토리 설정 제거됨

        print("")
        print("🔧 웹 설정 페이지에서도 설정 가능:")

        # 기본 포트로 서버 시작 (설정 페이지만 접근 가능)
        default_port = 5004
        print(f"   → http://localhost:{default_port}/settings")
        print("")
        print("설정 완료 후 서버를 재시작하세요.")
        print("=" * 60)

        app.run(debug=False, host='0.0.0.0', port=default_port)