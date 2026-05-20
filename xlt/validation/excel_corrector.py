"""
Claude AI 자동 교정 엔진

검증에서 발견된 문제들을 Claude AI로 자동 교정하는 시스템
"""

import json
import logging
import openpyxl
from openpyxl.styles import PatternFill
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
import time
import traceback
import copy

from ..translation.claude_translator import ClaudeTranslator
from ..translation.unifi_translator import UnifiTranslator
from .claude_prompts import ClaudePrompts

logger = logging.getLogger(__name__)

class ExcelCorrector:
    """Claude AI 기반 엑셀 자동 교정기"""

    def __init__(self, config):
        """초기화"""
        self.config = config
        self.claude_translator = ClaudeTranslator(config)
        self.unifi_translator = UnifiTranslator(config)
        self.prompts = ClaudePrompts()

        # 교정 결과 저장
        self.correction_results = {}
        self.correction_progress = {}

        logger.info("✅ ExcelCorrector 초기화 완료")

    def auto_correct_excel(self, session_id: str, validation_results: Dict[str, Any],
                          original_excel_data: Dict[str, Dict[str, str]]) -> Dict[str, Any]:
        """
        Claude AI 기반 엑셀 자동 교정 실행

        Args:
            session_id: 세션 ID
            validation_results: 검증 결과
            original_excel_data: 원본 엑셀 데이터

        Returns:
            교정 결과 딕셔너리
        """
        try:
            logger.info(f"🔧 Claude AI 자동 교정 시작: {session_id}")

            # 진행 상태 초기화
            self.correction_progress[session_id] = {
                'status': 'starting',
                'progress': 0,
                'current_step': '교정 준비 중...',
                'total_steps': 0,
                'corrected_items': 0,
                'start_time': time.time()
            }

            # 교정이 필요한 항목들 식별
            items_to_correct = self._identify_correction_items(validation_results, original_excel_data)

            if not items_to_correct:
                return {
                    'status': 'no_correction_needed',
                    'message': '교정이 필요한 항목이 없습니다.',
                    'corrected_data': original_excel_data
                }

            total_items = len(items_to_correct)
            self.correction_progress[session_id]['total_steps'] = total_items
            self.correction_progress[session_id]['status'] = 'correcting'

            # 교정된 데이터 초기화 (원본 복사)
            corrected_data = copy.deepcopy(original_excel_data)

            # 교정 히스토리
            correction_history = []

            # 용어집 데이터 로드
            terminology_data = self.unifi_translator.line_terminology

            # 각 항목별 교정 실행
            for i, correction_item in enumerate(items_to_correct, 1):
                try:
                    # 진행 상태 업데이트
                    self.correction_progress[session_id]['progress'] = int((i / total_items) * 100)
                    self.correction_progress[session_id]['current_step'] = f"{correction_item['key_id']} 교정 중..."

                    logger.info(f"🔧 {i}/{total_items} 교정 중: {correction_item['key_id']}")

                    # Claude AI로 교정 실행
                    correction_result = self._correct_single_item(
                        correction_item, terminology_data
                    )

                    if correction_result and correction_result.get('success', False):
                        # 교정된 데이터 적용
                        key_id = correction_item['key_id']
                        translations = correction_result['translations']

                        # 원본과 다른 경우만 교정 적용
                        for lang, new_translation in translations.items():
                            if lang in corrected_data[key_id]:
                                old_translation = corrected_data[key_id][lang]
                                if old_translation != new_translation:
                                    corrected_data[key_id][lang] = new_translation

                                    # 교정 히스토리 기록
                                    correction_history.append({
                                        'key_id': key_id,
                                        'language': lang,
                                        'original': old_translation,
                                        'corrected': new_translation,
                                        'reason': correction_item.get('reason', 'Claude AI 품질 개선'),
                                        'used_terminology': correction_result.get('used_terminology', [])
                                    })

                        self.correction_progress[session_id]['corrected_items'] += 1

                    else:
                        logger.warning(f"⚠️ 교정 실패: {correction_item['key_id']}")

                except Exception as e:
                    logger.error(f"❌ 개별 항목 교정 오류: {e}")
                    continue

            # 교정 완료 처리
            self.correction_progress[session_id]['status'] = 'validating'
            self.correction_progress[session_id]['current_step'] = '교정 결과 검증 중...'

            # 교정된 결과 검증
            validation_result = self._validate_correction_results(
                original_excel_data, corrected_data
            )

            # 최종 결과 생성
            final_result = {
                'session_id': session_id,
                'status': 'completed',
                'corrected_data': corrected_data,
                'correction_summary': {
                    'total_items_processed': total_items,
                    'successfully_corrected': len(correction_history),
                    'improvement_rate': len(correction_history) / total_items if total_items > 0 else 0
                },
                'correction_history': correction_history,
                'validation_result': validation_result,
                'completed_at': time.time(),
                'processing_time': time.time() - self.correction_progress[session_id]['start_time']
            }

            # 결과 저장
            self.correction_results[session_id] = final_result

            # 진행 상태 완료 처리
            self.correction_progress[session_id]['status'] = 'completed'
            self.correction_progress[session_id]['progress'] = 100
            self.correction_progress[session_id]['current_step'] = '교정 완료'

            logger.info(f"✅ Claude AI 자동 교정 완료: {len(correction_history)}개 항목 교정")
            return final_result

        except Exception as e:
            logger.error(f"❌ 자동 교정 오류: {e}")
            logger.error(traceback.format_exc())

            # 오류 상태 업데이트
            if session_id in self.correction_progress:
                self.correction_progress[session_id]['status'] = 'error'
                self.correction_progress[session_id]['error'] = str(e)

            return {
                'status': 'error',
                'error': str(e),
                'session_id': session_id
            }

    def _identify_correction_items(self, validation_results: Dict[str, Any],
                                 excel_data: Dict[str, Dict[str, str]]) -> List[Dict[str, Any]]:
        """교정이 필요한 항목들 식별"""
        try:
            correction_items = []

            # 검증 결과에서 문제 항목들 수집
            detailed_results = validation_results.get('detailed_results', {})

            # 1. 맞춤법 오류 항목들
            spelling_issues = detailed_results.get('spelling_validation', {}).get('issues', [])
            for issue in spelling_issues:
                # 해당 텍스트를 가진 Key ID 찾기
                key_id = self._find_key_id_by_text(issue['text'], excel_data, 'ko_KR')
                if key_id:
                    correction_items.append({
                        'key_id': key_id,
                        'issue_type': 'spelling',
                        'korean_text': issue['text'],
                        'reason': f"맞춤법 오류: {issue['error']}"
                    })

            # 2. 용어집 불일치 항목들
            terminology_issues = detailed_results.get('terminology_validation', {}).get('issues', [])
            for issue in terminology_issues:
                key_id = self._find_key_id_by_text(issue['text'], excel_data, 'ko_KR')
                if key_id:
                    correction_items.append({
                        'key_id': key_id,
                        'issue_type': 'terminology',
                        'korean_text': issue['text'],
                        'reason': f"용어 불일치: {issue['wrong_term']} -> {issue['suggested_term']}"
                    })

            # 3. 언어 불일치 항목들 (해당 언어만 재번역)
            language_issues = detailed_results.get('language_validation', {}).get('issues', [])
            for issue in language_issues:
                key_id = self._find_key_id_by_text(issue['text'], excel_data, issue['column'])
                if key_id:
                    correction_items.append({
                        'key_id': key_id,
                        'issue_type': 'language_mismatch',
                        'korean_text': excel_data[key_id].get('ko_KR', ''),
                        'target_language': issue['column'],
                        'reason': f"언어 불일치: {issue['detected_language']}"
                    })

            # 4. 다국어 용어 오류 항목들
            multilingual_issues = detailed_results.get('multilingual_validation', {}).get('issues', [])
            for issue in multilingual_issues:
                key_id = self._find_key_id_by_text(issue['text'], excel_data, issue['language'])
                if key_id:
                    correction_items.append({
                        'key_id': key_id,
                        'issue_type': 'multilingual_terminology',
                        'korean_text': excel_data[key_id].get('ko_KR', ''),
                        'target_language': issue['language'],
                        'reason': f"다국어 용어 오류: {issue['wrong_term']}"
                    })

            # 중복 제거 (Key ID 기준)
            unique_items = {}
            for item in correction_items:
                key_id = item['key_id']
                if key_id not in unique_items:
                    unique_items[key_id] = item
                else:
                    # 여러 이슈가 있는 경우 reason 합치기
                    existing_reason = unique_items[key_id]['reason']
                    new_reason = item['reason']
                    if new_reason not in existing_reason:
                        unique_items[key_id]['reason'] = f"{existing_reason}, {new_reason}"

            logger.info(f"📋 교정 대상 식별 완료: {len(unique_items)}개 항목")
            return list(unique_items.values())

        except Exception as e:
            logger.error(f"❌ 교정 항목 식별 오류: {e}")
            return []

    def _find_key_id_by_text(self, text: str, excel_data: Dict[str, Dict[str, str]],
                            language: str) -> Optional[str]:
        """텍스트로 Key ID 찾기"""
        try:
            for key_id, row_data in excel_data.items():
                if row_data.get(language, '') == text:
                    return key_id
            return None
        except Exception:
            return None

    def _correct_single_item(self, correction_item: Dict[str, Any],
                           terminology_data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """단일 항목 Claude AI 교정"""
        try:
            korean_text = correction_item['korean_text']
            if not korean_text or not korean_text.strip():
                return None

            # 목표 언어 설정
            if 'target_language' in correction_item:
                # 특정 언어만 교정
                target_languages = [correction_item['target_language']]
            else:
                # 전체 언어 교정
                target_languages = ['en_US', 'ja_JP', 'zh_TW', 'th_TH']

            # Claude AI 교정 프롬프트 생성
            prompt = self.prompts.get_auto_correction_prompt(
                korean_text, terminology_data, target_languages
            )

            # Claude AI 호출
            response = self._call_claude_ai(prompt)

            if response and 'translations' in response:
                return {
                    'success': True,
                    'original_text': korean_text,
                    'translations': response['translations'],
                    'used_terminology': response.get('used_terminology', [])
                }

            return None

        except Exception as e:
            logger.error(f"❌ 단일 항목 교정 오류: {e}")
            return None

    def _validate_correction_results(self, original_data: Dict[str, Dict[str, str]],
                                   corrected_data: Dict[str, Dict[str, str]]) -> Dict[str, Any]:
        """교정 결과 검증"""
        try:
            # Claude AI로 교정 결과 검증
            prompt = self.prompts.get_correction_validation_prompt(
                original_data, corrected_data
            )

            response = self._call_claude_ai(prompt)

            if response:
                return {
                    'status': 'completed',
                    'validation_result': response.get('validation_result', {}),
                    'detailed_feedback': response.get('detailed_feedback', [])
                }

            return {'status': 'validation_error', 'message': 'Claude AI 검증 실패'}

        except Exception as e:
            logger.error(f"❌ 교정 결과 검증 오류: {e}")
            return {'status': 'error', 'error': str(e)}

    def _call_claude_ai(self, prompt: str) -> Optional[Dict[str, Any]]:
        """Claude AI 호출"""
        try:
            # claude_translator를 통해 Claude AI 호출
            response = self.claude_translator._call_claude_api(prompt)

            if response and 'text' in response:
                # JSON 응답 파싱
                text = response['text']

                # JSON 블록 추출
                if '```json' in text:
                    json_start = text.find('```json') + 7
                    json_end = text.find('```', json_start)
                    json_text = text[json_start:json_end].strip()
                else:
                    json_text = text.strip()

                try:
                    return json.loads(json_text)
                except json.JSONDecodeError as e:
                    logger.error(f"❌ Claude AI 응답 JSON 파싱 실패: {e}")
                    logger.error(f"응답 텍스트: {text[:500]}...")
                    return None

            return None

        except Exception as e:
            logger.error(f"❌ Claude AI 호출 오류: {e}")
            return None

    def create_corrected_excel_file(self, session_id: str, original_file_path: str) -> Optional[str]:
        """교정된 엑셀 파일 생성"""
        try:
            correction_result = self.correction_results.get(session_id)
            if not correction_result:
                return None

            corrected_data = correction_result['corrected_data']
            correction_history = correction_result['correction_history']

            # 원본 파일 로드
            wb = openpyxl.load_workbook(original_file_path)
            ws = wb.active

            # 교정된 데이터 적용
            corrected_cells = set()
            for history in correction_history:
                key_id = history['key_id']
                language = history['language']

                # Key ID로 행 찾기
                for row in range(2, ws.max_row + 1):
                    if str(ws.cell(row=row, column=1).value) == key_id:
                        # 언어에 해당하는 열 찾기
                        for col in range(2, ws.max_column + 1):
                            header = ws.cell(row=1, column=col).value
                            if header == language:
                                # 교정된 텍스트 적용
                                ws.cell(row=row, column=col).value = history['corrected']
                                # 교정된 셀 하이라이트 (노란색 배경)
                                ws.cell(row=row, column=col).fill = PatternFill(
                                    start_color='FFFF00', end_color='FFFF00', fill_type='solid'
                                )
                                corrected_cells.add((row, col))
                                break
                        break

            # 교정 히스토리 시트 생성
            if correction_history:
                history_ws = wb.create_sheet("교정 히스토리")

                # 헤더 작성
                headers = ['Key ID', '언어', '교정 전', '교정 후', '교정 사유']
                for col, header in enumerate(headers, 1):
                    history_ws.cell(row=1, column=col).value = header

                # 교정 히스토리 데이터 작성
                for row, history in enumerate(correction_history, 2):
                    history_ws.cell(row=row, column=1).value = history['key_id']
                    history_ws.cell(row=row, column=2).value = history['language']
                    history_ws.cell(row=row, column=3).value = history['original']
                    history_ws.cell(row=row, column=4).value = history['corrected']
                    history_ws.cell(row=row, column=5).value = history['reason']

            # 파일 저장
            original_path = Path(original_file_path)
            corrected_file_path = original_path.parent / f"{original_path.stem}_corrected_{session_id}.xlsx"
            wb.save(corrected_file_path)

            logger.info(f"✅ 교정된 엑셀 파일 생성: {corrected_file_path}")
            return str(corrected_file_path)

        except Exception as e:
            logger.error(f"❌ 교정된 엑셀 파일 생성 오류: {e}")
            return None

    def get_correction_progress(self, session_id: str) -> Optional[Dict[str, Any]]:
        """교정 진행 상태 조회"""
        return self.correction_progress.get(session_id)

    def get_correction_results(self, session_id: str) -> Optional[Dict[str, Any]]:
        """교정 결과 조회"""
        return self.correction_results.get(session_id)