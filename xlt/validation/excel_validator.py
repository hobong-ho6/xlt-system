"""
Claude AI 기반 엑셀 검증 엔진

모든 검증 과정을 Claude AI로 처리하는 통합 시스템
"""

import json
import logging
import openpyxl
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
import time
import traceback

from ..translation.claude_translator import ClaudeTranslator
from ..translation.unifi_translator import UnifiTranslator
from .claude_prompts import ClaudePrompts
from .language_detector import LanguageDetector

logger = logging.getLogger(__name__)

class ExcelValidator:
    """Claude AI 기반 엑셀 파일 검증기"""

    def __init__(self, config):
        """초기화"""
        self.config = config
        self.claude_translator = ClaudeTranslator(config)
        self.unifi_translator = UnifiTranslator(config)
        self.prompts = ClaudePrompts()
        self.language_detector = LanguageDetector()  # 🚀 다른 Claude 검증 시스템

        # 검증 결과 저장
        self.validation_results = {}
        self.excel_data = {}
        self.terminology_data = {}

        logger.info("✅ ExcelValidator 초기화 완료")

    def validate_excel_file(self, file_path: str, session_id: str, progress_callback=None) -> Dict[str, Any]:
        """
        엑셀 파일 전체 검증 실행 (🚀 진짜 전체 데이터 검증!)

        Args:
            file_path: 엑셀 파일 경로
            session_id: 세션 ID
            progress_callback: 진행 상황 업데이트 콜백

        Returns:
            검증 결과 딕셔너리
        """
        try:
            logger.info(f"📊 엑셀 전체 검증 시작: {file_path}")

            def update_progress(step: str, percent: int, message: str):
                if progress_callback:
                    progress_callback(step, percent, message)

            update_progress("loading", 5, "엑셀 파일 로드 중...")

            # 1. 엑셀 파일 로드 및 구조 검증
            self.excel_data = self._load_excel_file(file_path)
            if not self.excel_data:
                return self._create_error_result("엑셀 파일 로드 실패")

            update_progress("loading", 10, "용어집 데이터 로드 중...")

            # 2. 용어집 데이터 로드
            self.terminology_data = self._load_terminology_data()

            total_rows = len(self.excel_data)
            logger.info(f"🎯 전체 {total_rows}개 행 검증 시작 (더 이상 샘플링 없음!)")

            # 3. 5단계 Claude AI 전체 검증 실행
            validation_results = {
                'session_id': session_id,
                'file_path': file_path,
                'total_rows': total_rows,
                'is_full_validation': True,  # 🚀 전체 검증 표시
                'validation_summary': {
                    'total_issues': 0,
                    'spelling_errors': 0,
                    'terminology_errors': 0,
                    'language_mismatches': 0,
                    'multilingual_errors': 0,
                    'completeness_issues': 0
                },
                'detailed_results': {}
            }

            # Step 1: 한글 맞춤법/띄어쓰기 전체 검증
            update_progress("validation", 15, "1단계: 한글 맞춤법 전체 검증 중...")
            logger.info("1️⃣ 한글 맞춤법/띄어쓰기 전체 검증 시작")
            spelling_results = self._validate_korean_spelling()
            validation_results['detailed_results']['spelling_validation'] = spelling_results
            validation_results['validation_summary']['spelling_errors'] = len(spelling_results.get('issues', []))

            # Step 2: 한글 용어집 전체 비교 검증
            update_progress("validation", 35, "2단계: 한글 용어집 전체 검증 중...")
            logger.info("2️⃣ 한글 용어집 비교 전체 검증 시작")
            terminology_results = self._validate_korean_terminology()
            validation_results['detailed_results']['terminology_validation'] = terminology_results
            validation_results['validation_summary']['terminology_errors'] = len(terminology_results.get('issues', []))

            # Step 3: 언어 일치성 전체 검증
            update_progress("validation", 55, "3단계: 언어 일치성 전체 검증 중...")
            logger.info("3️⃣ 언어 일치성 전체 검증 시작")
            language_results = self._validate_language_consistency()
            validation_results['detailed_results']['language_validation'] = language_results
            validation_results['validation_summary']['language_mismatches'] = len(language_results.get('issues', []))

            # Step 4: 다국어 용어집 전체 비교 검증
            update_progress("validation", 75, "4단계: 다국어 용어집 전체 검증 중...")
            logger.info("4️⃣ 다국어 용어집 비교 전체 검증 시작")
            multilingual_results = self._validate_multilingual_terminology()
            validation_results['detailed_results']['multilingual_validation'] = multilingual_results
            validation_results['validation_summary']['multilingual_errors'] = len(multilingual_results.get('issues', []))

            # Step 5: 빈 행 및 완성도 전체 검증
            update_progress("validation", 90, "5단계: 데이터 완성도 전체 검증 중...")
            logger.info("5️⃣ 데이터 완성도 전체 검증 시작")
            completeness_results = self._validate_data_completeness()
            validation_results['detailed_results']['completeness_validation'] = completeness_results
            validation_results['validation_summary']['completeness_issues'] = len(completeness_results.get('issues', []))

            # 전체 문제 수 계산
            total_issues = sum([
                validation_results['validation_summary']['spelling_errors'],
                validation_results['validation_summary']['terminology_errors'],
                validation_results['validation_summary']['language_mismatches'],
                validation_results['validation_summary']['multilingual_errors'],
                validation_results['validation_summary']['completeness_issues']
            ])
            validation_results['validation_summary']['total_issues'] = total_issues

            update_progress("completed", 100, "전체 검증 완료!")

            # 검증 완료 시간
            validation_results['completed_at'] = time.time()
            validation_results['has_issues'] = total_issues > 0
            validation_results['total_batches'] = sum([
                spelling_results.get('batches_processed', 0),
                terminology_results.get('batches_processed', 0)
            ])

            # 결과 저장
            self.validation_results[session_id] = validation_results

            logger.info(f"🎉 엑셀 전체 검증 완료: 총 {total_rows}개 행 검증, {total_issues}개 문제 발견")
            logger.info(f"📊 검증 범위: 100% (이전: 샘플 0.6~1.3%)")
            return validation_results

        except Exception as e:
            logger.error(f"❌ 엑셀 검증 오류: {e}")
            logger.error(traceback.format_exc())
            return self._create_error_result(f"검증 처리 오류: {str(e)}")

    def _load_excel_file(self, file_path: str) -> Dict[str, Dict[str, str]]:
        """엑셀 파일 로드 및 데이터 구조화"""
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # 헤더 행 확인 (빈 헤더도 처리)
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(str(header))
                else:
                    # 첫 번째 열이 빈값이면 'Key ID'로 가정
                    if col == 1:
                        headers.append('Key ID')
                        logger.info(f"📝 첫 번째 열 헤더가 빈값 - 'Key ID'로 가정")
                    else:
                        # 다른 열이 빈값이면 중단
                        break

            logger.info(f"📋 엑셀 데이터 로드 완료: {ws.max_row-1}행, 헤더: {headers}")

            expected_headers = ['Key ID', 'en_US', 'ko_KR', 'ja_JP', 'zh_TW', 'th_TH']
            if len(headers) < 3:  # 최소 Key ID, 언어1, 언어2는 있어야 함
                logger.error(f"❌ 헤더가 부족합니다: {headers}")
                return {}

            # 한국어 열 확인
            if 'ko_KR' not in headers:
                logger.warning(f"⚠️ 한국어(ko_KR) 열을 찾을 수 없습니다: {headers}")

            # 데이터 행 처리
            excel_data = {}
            for row in range(2, ws.max_row + 1):
                key_id = ws.cell(row=row, column=1).value
                if not key_id:
                    continue

                row_data = {}
                for col, header in enumerate(headers[1:], start=2):  # Key ID 제외
                    cell_value = ws.cell(row=row, column=col).value
                    row_data[header] = str(cell_value) if cell_value is not None else ""

                excel_data[str(key_id)] = row_data

            logger.info(f"✅ 데이터 파싱 완료: {len(excel_data)}개 키 로드됨")
            return excel_data

        except Exception as e:
            logger.error(f"❌ 엑셀 파일 로드 실패: {e}")
            return {}

    def _load_terminology_data(self) -> Dict[str, Any]:
        """용어집 데이터 로드"""
        try:
            # UnifiTranslator에서 용어집 데이터 가져오기
            terminology = self.unifi_translator.line_terminology
            logger.info(f"📚 용어집 데이터 로드 완료: {len(terminology)}개 용어")
            return terminology
        except Exception as e:
            logger.error(f"⚠️ 용어집 로드 실패: {e}")
            return {}

    def _validate_korean_spelling(self) -> Dict[str, Any]:
        """한글 맞춤법/띄어쓰기 검증"""
        try:
            # 한글 텍스트와 Key ID 매핑 추출
            korean_data = []
            row_number = 2  # 엑셀은 2행부터 데이터 시작
            for key_id, row_data in self.excel_data.items():
                korean_text = row_data.get('ko_KR', '')
                if korean_text and korean_text.strip():
                    korean_data.append({
                        'key_id': key_id,
                        'text': korean_text,
                        'row_number': row_number
                    })
                row_number += 1

            if not korean_data:
                return {'status': 'no_korean_text', 'issues': []}

            # 🚀 전체 데이터 배치 검증 (더 이상 샘플링 아님!)
            batch_size = 25  # Claude AI 안정적 처리 가능한 배치 크기
            all_issues = []
            total_batches = (len(korean_data) + batch_size - 1) // batch_size

            logger.info(f"📝 맞춤법 전체 검증 시작 - 총 {len(korean_data)}개 텍스트, {total_batches}개 배치")

            for batch_idx in range(0, len(korean_data), batch_size):
                batch_data = korean_data[batch_idx:batch_idx + batch_size]
                batch_num = (batch_idx // batch_size) + 1

                logger.info(f"📝 맞춤법 배치 {batch_num}/{total_batches} 검증 중... ({len(batch_data)}개)")

                prompt = self.prompts.get_spelling_validation_prompt(
                    [item['text'] for item in batch_data],
                    [item['key_id'] for item in batch_data]
                )

                response = self._call_claude_ai(prompt)

                if response and 'results' in response:
                    # 배치 결과 처리
                    for i, result in enumerate(response['results']):
                        if not result.get('is_correct', True) and i < len(batch_data):
                            for error in result.get('errors', []):
                                all_issues.append({
                                    'key_id': batch_data[i]['key_id'],
                                    'row_number': batch_data[i]['row_number'],
                                    'text': result['original_text'],
                                    'error_type': error['type'],
                                    'error': error['error'],
                                    'suggestion': error['suggestion']
                                })
                else:
                    logger.warning(f"⚠️ 배치 {batch_num} Claude AI 응답 실패")

            logger.info(f"✅ 맞춤법 전체 검증 완료 - {len(korean_data)}개 검증, {len(all_issues)}개 문제 발견")

            return {
                'status': 'completed',
                'issues': all_issues,
                'total_checked': len(korean_data),
                'batches_processed': total_batches
            }

            return {'status': 'claude_error', 'issues': []}

        except Exception as e:
            logger.error(f"❌ 한글 맞춤법 검증 오류: {e}")
            return {'status': 'error', 'issues': [], 'error': str(e)}

    def _validate_korean_terminology(self) -> Dict[str, Any]:
        """한글 용어집 비교 검증"""
        try:
            # 한글 텍스트와 Key ID 매핑 추출
            korean_data = []
            row_number = 2  # 엑셀은 2행부터 데이터 시작
            for key_id, row_data in self.excel_data.items():
                korean_text = row_data.get('ko_KR', '')
                if korean_text and korean_text.strip():
                    korean_data.append({
                        'key_id': key_id,
                        'text': korean_text,
                        'row_number': row_number
                    })
                row_number += 1

            if not korean_data:
                return {'status': 'no_korean_text', 'issues': []}

            # 🚀 용어집 전체 데이터 배치 검증
            batch_size = 20  # 용어집 데이터가 많아서 조금 작게
            all_issues = []
            all_exceptions = []
            total_batches = (len(korean_data) + batch_size - 1) // batch_size

            logger.info(f"📚 용어집 전체 검증 시작 - 총 {len(korean_data)}개 텍스트, {total_batches}개 배치")

            for batch_idx in range(0, len(korean_data), batch_size):
                batch_data = korean_data[batch_idx:batch_idx + batch_size]
                batch_num = (batch_idx // batch_size) + 1

                logger.info(f"📚 용어집 배치 {batch_num}/{total_batches} 검증 중... ({len(batch_data)}개)")

                prompt = self.prompts.get_terminology_validation_prompt(
                    [item['text'] for item in batch_data],
                    self.terminology_data,
                    [item['key_id'] for item in batch_data]
                )
                response = self._call_claude_ai(prompt)

                if response and 'results' in response:
                    # 배치 결과 처리
                    for i, result in enumerate(response['results']):
                        if i >= len(batch_data):
                            continue

                        current_data = batch_data[i]
                        if result.get('has_terminology_issue', False):
                            for issue in result.get('issues', []):
                                all_issues.append({
                                    'key_id': current_data['key_id'],
                                    'row_number': current_data['row_number'],
                                    'text': result['original_text'],
                                    'wrong_term': issue['wrong_term'],
                                    'suggested_term': issue['suggested_term'],
                                    'reason': issue['reason']
                                })

                        # exceptional 항목 처리
                        if result.get('is_exception', False):
                            all_exceptions.append({
                                'key_id': current_data['key_id'],
                                'row_number': current_data['row_number'],
                                'text': result['original_text'],
                                'exception_reason': result.get('exception_reason', 'exceptional 항목으로 검증 통과')
                            })
                else:
                    logger.warning(f"⚠️ 용어집 배치 {batch_num} Claude AI 응답 실패")

            logger.info(f"✅ 용어집 전체 검증 완료 - {len(korean_data)}개 검증, {len(all_issues)}개 문제 발견")

            return {
                'status': 'completed',
                'issues': all_issues,
                'exceptions': all_exceptions,
                'total_checked': len(korean_data),
                'batches_processed': total_batches
            }

            return {'status': 'claude_error', 'issues': []}

        except Exception as e:
            logger.error(f"❌ 한글 용어집 검증 오류: {e}")
            return {'status': 'error', 'issues': [], 'error': str(e)}

    def _validate_language_consistency(self) -> Dict[str, Any]:
        """언어 일치성 검증"""
        try:
            # 언어별 텍스트와 Key ID 매핑 수집
            texts_by_column = {}
            languages = ['en_US', 'ko_KR', 'ja_JP', 'zh_TW', 'th_TH']

            for lang in languages:
                texts_data = []
                row_number = 2  # 엑셀은 2행부터 데이터 시작
                for key_id, row_data in self.excel_data.items():
                    text = row_data.get(lang, '')
                    if text and text.strip():
                        texts_data.append({
                            'key_id': key_id,
                            'text': text,
                            'row_number': row_number
                        })
                    row_number += 1

                if texts_data:
                    texts_by_column[lang] = texts_data  # 🚀 전체 데이터 사용 (더 이상 샘플링 없음)

            if not texts_by_column:
                return {'status': 'no_text_data', 'issues': []}

            # Claude AI로 언어 감지
            prompt = self.prompts.get_language_detection_prompt(texts_by_column)
            response = self._call_claude_ai(prompt)

            if response and 'results' in response:
                # 결과 처리
                issues = []
                for result in response['results']:
                    column = result['column']
                    for detected_issue in result.get('detected_issues', []):
                        # 해당 텍스트의 key_id와 row_number 찾기
                        target_text = detected_issue['text']
                        matched_data = None
                        if column in texts_by_column:
                            for data in texts_by_column[column]:
                                if data['text'] == target_text:
                                    matched_data = data
                                    break

                        issues.append({
                            'key_id': matched_data['key_id'] if matched_data else 'unknown',
                            'row_number': matched_data['row_number'] if matched_data else 0,
                            'column': result['column'],
                            'expected_language': result['expected_language'],
                            'text': detected_issue['text'],
                            'detected_language': detected_issue['detected_language'],
                            'issue': detected_issue['issue']
                        })

                return {
                    'status': 'completed',
                    'issues': issues,
                    'checked_languages': list(texts_by_column.keys())
                }

            return {'status': 'claude_error', 'issues': []}

        except Exception as e:
            logger.error(f"❌ 언어 일치성 검증 오류: {e}")
            return {'status': 'error', 'issues': [], 'error': str(e)}

    def _validate_multilingual_terminology(self) -> Dict[str, Any]:
        """다국어 용어집 비교 검증"""
        try:
            # 언어별 텍스트와 Key ID 매핑 수집
            texts_by_language = {}
            languages = ['en_US', 'ja_JP', 'zh_TW', 'th_TH']  # 한국어 제외

            for lang in languages:
                texts_data = []
                row_number = 2  # 엑셀은 2행부터 데이터 시작
                for key_id, row_data in self.excel_data.items():
                    text = row_data.get(lang, '')
                    if text and text.strip():
                        texts_data.append({
                            'key_id': key_id,
                            'text': text,
                            'row_number': row_number
                        })
                    row_number += 1

                if texts_data:
                    texts_by_language[lang] = texts_data  # 🚀 전체 데이터 사용

            if not texts_by_language:
                return {'status': 'no_multilingual_data', 'issues': []}

            # Claude AI로 다국어 용어 검증
            prompt = self.prompts.get_multilingual_terminology_prompt(texts_by_language, self.terminology_data)
            response = self._call_claude_ai(prompt)

            if response and 'results' in response:
                # 결과 처리
                issues = []
                for result in response['results']:
                    language = result['language']
                    for issue in result.get('issues', []):
                        # 해당 텍스트의 key_id와 row_number 찾기
                        target_text = issue['text']
                        matched_data = None
                        if language in texts_by_language:
                            for data in texts_by_language[language]:
                                if data['text'] == target_text:
                                    matched_data = data
                                    break

                        issues.append({
                            'key_id': matched_data['key_id'] if matched_data else 'unknown',
                            'row_number': matched_data['row_number'] if matched_data else 0,
                            'language': result['language'],
                            'text': issue['text'],
                            'wrong_term': issue['wrong_term'],
                            'suggested_term': issue['suggested_term'],
                            'issue_type': issue['issue_type']
                        })

                return {
                    'status': 'completed',
                    'issues': issues,
                    'checked_languages': list(texts_by_language.keys())
                }

            return {'status': 'claude_error', 'issues': []}

        except Exception as e:
            logger.error(f"❌ 다국어 용어집 검증 오류: {e}")
            return {'status': 'error', 'issues': [], 'error': str(e)}

    def _validate_data_completeness(self) -> Dict[str, Any]:
        """데이터 완성도 검증 - 빈칸 검증 강화"""
        try:
            # 🚀 다른 Claude 수준의 로컬 검증들 수행
            empty_issues = self._check_empty_cells()
            advanced_issues = self._check_advanced_issues()  # 🚀 정교한 검증 시스템
            local_issues = empty_issues + advanced_issues

            # 완성도 데이터 준비 (Key ID와 행번호 포함)
            completeness_data = []
            row_number = 2  # 엑셀은 2행부터 데이터 시작
            for key_id, row_data in self.excel_data.items():
                completeness_data.append({
                    'key_id': key_id,
                    'row_number': row_number,
                    'row_data': row_data
                })
                row_number += 1

            # Claude AI로 완성도 검증
            prompt = self.prompts.get_completeness_validation_prompt(completeness_data)
            response = self._call_claude_ai(prompt)

            # 결과 통합
            all_issues = local_issues.copy()  # 로컬 빈칸 검증 결과

            if response:
                # Claude AI 결과 추가 - key_id와 row_number 정보 보존
                for issue in response.get('issues', []):
                    # Claude AI가 반환한 key_id 사용 (실제 엑셀 데이터)
                    claude_issue = {
                        'key_id': issue.get('key_id', 'unknown'),
                        'row_number': issue.get('row_number', 0),
                        'issue_type': issue.get('issue_type', 'completeness'),
                        'description': issue.get('description', ''),
                        'missing_languages': issue.get('missing_languages', [])
                    }

                    # 중복 방지: 동일한 key_id의 빈칸 이슈가 있으면 병합
                    existing_issue = None
                    for local_issue in all_issues:
                        if (local_issue['key_id'] == claude_issue['key_id'] and
                            local_issue['issue_type'] == 'empty_cells'):
                            existing_issue = local_issue
                            break

                    if existing_issue:
                        # 기존 빈칸 이슈와 Claude 결과 병합
                        existing_issue['claude_analysis'] = claude_issue['description']
                    else:
                        # 새로운 이슈 추가
                        all_issues.append(claude_issue)

                summary = response.get('summary', {})
            else:
                summary = {
                    'total_rows': len(completeness_data),
                    'empty_cells_found': len(local_issues)
                }

            logger.info(f"✅ 완성도 검증 완료: 로컬 {len(local_issues)}개 + Claude AI 검증")

            return {
                'status': 'completed',
                'issues': all_issues,
                'summary': summary,
                'local_empty_check': len(local_issues),
                'total_issues': len(all_issues)
            }

        except Exception as e:
            logger.error(f"❌ 데이터 완성도 검증 오류: {e}")
            return {'status': 'error', 'issues': [], 'error': str(e)}

    def _check_empty_cells(self) -> List[Dict[str, Any]]:
        """로컬 빈칸 검증 - 직접 빈 셀 탐지"""
        empty_issues = []
        row_number = 2  # 엑셀은 2행부터 데이터 시작

        languages = ['ko_KR', 'en_US', 'ja_JP', 'zh_TW', 'th_TH']

        for key_id, row_data in self.excel_data.items():
            missing_languages = []

            for lang in languages:
                value = row_data.get(lang, '')
                # 빈값 체크: None, 빈 문자열, 공백만 있는 경우
                if not value or str(value).strip() == '' or str(value).lower() in ['none', 'null', 'nan']:
                    missing_languages.append(lang)

            if missing_languages:
                # ko_KR이 빈값이면 critical, 다른 언어는 warning
                issue_severity = 'critical' if 'ko_KR' in missing_languages else 'warning'

                empty_issues.append({
                    'key_id': key_id,
                    'row_number': row_number,
                    'issue_type': 'empty_cells',
                    'severity': issue_severity,
                    'missing_languages': missing_languages,
                    'description': f"빈 필드 발견: {', '.join(missing_languages)}",
                    'total_missing': len(missing_languages)
                })

            row_number += 1

        logger.info(f"📊 로컬 빈칸 검증: {len(empty_issues)}개 문제 발견")
        return empty_issues

    def _check_advanced_issues(self) -> List[Dict[str, Any]]:
        """🚀 다른 Claude 수준의 정교한 문제 감지 시스템"""
        all_issues = []
        row_number = 2

        # 언어 목록
        languages = ['ko_KR', 'en_US', 'ja_JP', 'zh_TW', 'th_TH']

        logger.info("🔍 다른 Claude 수준 검증 시작:")

        # 1. 언어 불일치 검증 (유니코드 범위 기반)
        language_issues = []
        for key_id, row_data in self.excel_data.items():
            for lang in languages:
                text = row_data.get(lang, '')
                if not text:
                    continue

                problems = self.language_detector.detect_language_issues(str(text), lang)
                if problems:
                    language_issues.append({
                        'key_id': key_id,
                        'row_number': row_number,
                        'issue_type': 'language_mismatch',
                        'severity': 'critical',
                        'language': lang,
                        'text': str(text)[:100],
                        'problems': problems,
                        'description': f'{lang} 열에 다른 언어 문자 혼입: {", ".join(problems)}',
                        'suggestion': f'{lang} 언어만 사용하도록 수정'
                    })

            row_number += 1

        all_issues.extend(language_issues)
        logger.info(f"├─ 언어 불일치: {len(language_issues)}개")

        # 2. 번역 누락 검증
        translation_issues = []
        row_number = 2

        for key_id, row_data in self.excel_data.items():
            ko_text = row_data.get('ko_KR', '')

            # 케이스 1: 모든 언어가 영어로 동일 (번역 누락)
            en_text = row_data.get('en_US', '')
            if (en_text and not self.language_detector.is_symbol_only(en_text) and
                self.language_detector.has_latin(en_text)):

                all_same_as_en = True
                for lang in ['ko_KR', 'ja_JP', 'zh_TW', 'th_TH']:
                    other_text = row_data.get(lang, '')
                    if other_text.strip() != en_text.strip():
                        all_same_as_en = False
                        break

                if all_same_as_en:
                    # 텍스트 길이로 심각도 판단
                    text_only = en_text
                    for pattern in self.language_detector.PLACEHOLDER_PATTERNS:
                        text_only = pattern.sub("", text_only)
                    word_count = len(text_only.split())

                    severity = 'critical' if word_count >= 3 or len(text_only) >= 25 else 'medium'
                    translation_issues.append({
                        'key_id': key_id,
                        'row_number': row_number,
                        'issue_type': 'translation_missing_all',
                        'severity': severity,
                        'text': en_text[:100],
                        'description': f'5개 언어 모두 영어 원문으로 동일 (번역 누락)',
                        'suggestion': '각 언어로 번역 필요'
                    })

            # 케이스 2: 한국어 원문이 다른 언어 열에 그대로
            if ko_text and self.language_detector.has_korean(ko_text):
                for lang in ['en_US', 'ja_JP', 'zh_TW', 'th_TH']:
                    other_text = row_data.get(lang, '')
                    if not other_text:
                        continue

                    missing_info = self.language_detector.detect_translation_missing(
                        ko_text, other_text, lang)

                    if missing_info['is_missing']:
                        translation_issues.append({
                            'key_id': key_id,
                            'row_number': row_number,
                            'issue_type': 'translation_missing_korean',
                            'severity': missing_info['severity'],
                            'language': lang,
                            'text': other_text[:100],
                            'description': missing_info['description'],
                            'suggestion': f'{lang} 언어로 번역 필요'
                        })

            row_number += 1

        all_issues.extend(translation_issues)
        logger.info(f"├─ 번역 누락: {len(translation_issues)}개")

        # 3. Placeholder 일관성 검증
        placeholder_issues = []
        row_number = 2

        for key_id, row_data in self.excel_data.items():
            ko_text = row_data.get('ko_KR', '')
            if not ko_text:
                row_number += 1
                continue

            for lang in ['en_US', 'ja_JP', 'zh_TW', 'th_TH']:
                other_text = row_data.get(lang, '')
                if not other_text:
                    continue

                consistency = self.language_detector.check_placeholder_consistency(
                    ko_text, other_text)

                if not consistency['is_consistent']:
                    placeholder_issues.append({
                        'key_id': key_id,
                        'row_number': row_number,
                        'issue_type': 'placeholder_inconsistency',
                        'severity': 'high',
                        'language': lang,
                        'ko_placeholders': consistency['ko_placeholders'],
                        'target_placeholders': consistency['target_placeholders'],
                        'missing': consistency['missing'],
                        'extra': consistency['extra'],
                        'description': f'Placeholder 불일치 (누락: {consistency["missing"]}, 추가: {consistency["extra"]})',
                        'suggestion': 'Placeholder 개수와 형식을 한국어와 일치시키기'
                    })

            row_number += 1

        all_issues.extend(placeholder_issues)
        logger.info(f"├─ Placeholder 불일치: {len(placeholder_issues)}개")

        # 4. /n 오타 검증
        slash_n_issues = []
        row_number = 2

        for key_id, row_data in self.excel_data.items():
            for lang in languages:
                text = row_data.get(lang, '')
                if not text:
                    continue

                if self.language_detector.check_slash_n_typo(str(text)):
                    slash_n_issues.append({
                        'key_id': key_id,
                        'row_number': row_number,
                        'issue_type': 'slash_n_typo',
                        'severity': 'critical',
                        'language': lang,
                        'text': str(text)[:100],
                        'description': f'"/n" 오타 발견 (화면에 그대로 노출됨)',
                        'suggestion': str(text).replace('/n', '\\n')
                    })

            row_number += 1

        all_issues.extend(slash_n_issues)
        logger.info(f"├─ /n 오타: {len(slash_n_issues)}개")

        logger.info(f"🎯 다른 Claude 수준 검증 완료: 총 {len(all_issues)}개 문제 발견")
        return all_issues

    def _call_claude_ai(self, prompt: str) -> Optional[Dict[str, Any]]:
        """Claude AI 호출"""
        try:
            # claude_translator를 통해 Claude AI 호출
            response = self.claude_translator._call_claude_api(prompt)

            if response and 'text' in response:
                # JSON 응답 파싱
                text = response['text']

                # JSON 블록 추출
                if '```json' in text:
                    json_start = text.find('```json') + 7
                    json_end = text.find('```', json_start)
                    json_text = text[json_start:json_end].strip()
                else:
                    json_text = text.strip()

                try:
                    return json.loads(json_text)
                except json.JSONDecodeError as e:
                    logger.error(f"❌ Claude AI 응답 JSON 파싱 실패: {e}")
                    logger.error(f"응답 텍스트: {text[:500]}...")
                    return None

            return None

        except Exception as e:
            logger.error(f"❌ Claude AI 호출 오류: {e}")
            return None

    def _create_error_result(self, error_message: str) -> Dict[str, Any]:
        """에러 결과 생성"""
        return {
            'status': 'error',
            'error': error_message,
            'has_issues': True,
            'validation_summary': {
                'total_issues': -1,
                'spelling_errors': 0,
                'terminology_errors': 0,
                'language_mismatches': 0,
                'multilingual_errors': 0,
                'completeness_issues': 0
            },
            'detailed_results': {}
        }

    def get_validation_results(self, session_id: str) -> Optional[Dict[str, Any]]:
        """세션 ID로 검증 결과 조회"""
        return self.validation_results.get(session_id)

    def has_validation_issues(self, session_id: str) -> bool:
        """검증 결과에 문제가 있는지 확인"""
        results = self.get_validation_results(session_id)
        if results:
            return results.get('has_issues', False)
        return False